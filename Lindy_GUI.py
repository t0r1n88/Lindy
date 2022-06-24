import tkinter
import numpy as np
import pandas as pd
import os
from dateutil.parser import ParserError
from docxtpl import DocxTemplate
from docxcompose.composer import Composer
from docx import Document
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import time
import datetime
from datetime import date
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, Reference, PieChart, PieChart3D, Series
pd.options.display.max_colwidth = 100
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import re
import tempfile



def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and f  or PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_file_template_doc():
    """
    Функция для выбора файла шаблона
    :return: Путь к файлу шаблона
    """
    global name_file_template_doc
    name_file_template_doc = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))


def select_file_template_table():
    """
    Функция для выбора шаблона для создания общей таблицы
    :return:
    """
    global name_file_template_table
    # Получаем путь к файлу
    name_file_template_table = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_file_data_doc():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global name_file_data_doc
    # Получаем путь к файлу
    name_file_data_doc = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_doc():
    """
    Функция для выбора папки куда будут генерироваться файлы
    :return:
    """
    global path_to_end_folder_doc
    path_to_end_folder_doc = filedialog.askdirectory()


def select_file_data_report():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global name_file_data_report
    # Получаем путь к файлу
    name_file_data_report = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_report():
    """
    Функция для выбора папки куда будут генерироваться файлы
    :return:
    """
    global path_to_end_folder_report
    path_to_end_folder_report = filedialog.askdirectory()


def select_files_data_groups():
    """
    Функция для выбора файлов с данными при выполнении прочих операций
    :return:
    """
    # Создаем глобальную переменную, дада я знаю что надо все сделать в виде классов.Потом когда нибудь
    global path_to_files_groups
    path_to_files_groups = filedialog.askdirectory()


def select_file_params_calculate_data():
    """
    Функция для выбора файла c ячейками которые нужно подсчитать
    :return: Путь к файлу
    """
    global name_file_params_calculate_data
    name_file_params_calculate_data = filedialog.askopenfilename(
        filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_files_data_calculate_data():
    """
    Функция для выбора файлов с данными параметры из которых нужно подсчитать
    :return: Путь к файлам с данными
    """
    global names_files_calculate_data
    # Получаем путь к файлу
    names_files_calculate_data = filedialog.askopenfilenames(
        filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_calculate_data():
    """
    Функция для выбора папки куда будут генерироваться файл  с результатом подсчета и файл с проверочной инфомрацией
    :return:
    """
    global path_to_end_folder_calculate_data
    path_to_end_folder_calculate_data = filedialog.askdirectory()


def calculate_age(born):
    """
    Функция для расчета текущего возраста взято с https://stackoverflow.com/questions/2217488/age-from-birthdate-in-python/9754466#9754466
    :param born: дата рождения
    :return: возраст
    """

    try:
        today = date.today()
        return today.year - born.year - ((today.month, today.day) < (born.month, born.day))
    except TypeError:
        print(born)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()
    except ValueError:
        print(f' Ошибка при подсчете текущего возраста ячейки {born}')
        messagebox.showerror('ЦОПП Бурятия', 'Пустая ячейка с датой или некорректная запись!!!')
        quit()
    except:
        print(f' Ошибка при подсчете текущего возраста ячейки {born}')
        messagebox.showerror('ЦОПП Бурятия', 'Отсутствует или некорректная дата \nПроверьте файл!')
        quit()

def check_date_columns(i, value):
    """
    Функция для проверки типа колонки. Необходимо найти колонки с датой
    :param i:
    :param value:
    :return:
    """
    #  Да да это просто
    if '00:00:00' in str(value):
        try:
            itog = pd.to_datetime(str(value),infer_datetime_format=True)

        except ParserError:
            pass
        except ValueError:
            pass
        except TypeError:
            pass
        else:
            return i

def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return ''
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()

def convert_date(cell):
    """
    Функция для конвертации даты в формате 1957-05-10 в формат 10.05.1957(строковый)
    """

    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except TypeError:
        print(f' Ошибка при конвертации ячейки {cell}')
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()
    except ValueError:
        print(f' Ошибка при конвертации ячейки {cell}')
        messagebox.showerror('ЦОПП Бурятия', 'Пустая ячейка с датой или некорректная запись!!!')
        quit()

def extract_date_begin_course(cell:str):
    """
    Функция для извлечения даты начала курса
    """

    try:
        # Находим обе даты
        match = re.findall(r'\d\d.\d\d.\d\d\d\d', cell)
        # date_course = datetime.datetime.strptime(match[0], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return match[0]
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()



def extract_date_end_course(cell:str):
    """
    Функция для извлечения даты окончания курса
    """
    try:
        # Находим обе даты
        match = re.findall(r'\d\d.\d\d.\d\d\d\d', cell)
        # Конвертируем строку
        # date_course = datetime.datetime.strptime(match[1], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return match[1]
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()

def extract_month_begin_course(cell:str):
    """
    Функция для извлечения месяца начала курса в формате от 1 до 12
    """
    try:
        # Находим оба месяца выделив месяц круглыми скобками
        match = re.findall(r'\d\d.(\d\d).\d\d\d\d', cell)
        # Конвертируем строку
        # date_course = datetime.datetime.strptime(match[1], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return int(match[0])
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()


def extract_month_end_course(cell:str):
    """
    Функция для извлечения месяца окончания курса в формате от 1 до 12
    """
    try:
        # Находим обе даты
        match = re.findall(r'\d\d.(\d\d).\d\d\d\d', cell)
        # Конвертируем строку
        # date_course = datetime.datetime.strptime(match[1], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return int(match[1])
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()


def create_initials(fio):
    """
    Функция для создания инициалов для использования в договорах
    формат фио -Будаев Олег Тимурович выходной формат О.Т. Будаев
    """
    # Создаем 3 переменные

    initials_firstname = ''
    initials_middlename = ''
    initials_lastname = ''

    # Сплитим по пробелу
    if type(fio) == float:
        print(fio)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность написания ФИО в столбце ФИО_именительный')
        quit()

    lst_fio = fio.split()
    # Если ФИО стандартное
    if len(lst_fio) == 3:

        lastname = lst_fio[0]
        firstname = lst_fio[1]
        middlename = lst_fio[2]
        # Создаем инициалы
        initials_firstname = firstname[0].upper()
        initials_middlename = middlename[0].upper()
        initials_lastname = lastname
        # Возвращаем полученную строку
        return f'{initials_firstname}.{initials_middlename}. {initials_lastname}'
    elif len(lst_fio) == 2:
        lastname = lst_fio[0]
        firstname = lst_fio[1]

        initials_firstname = firstname[0].upper()
        initials_lastname = lastname
        return f'{initials_firstname}. {initials_lastname}'
    elif len(lst_fio) == 1:
        lastname = lst_fio[0]
        initials_lastname = lastname
        return f'{initials_lastname}'
    else:
        print(fio)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность написания ФИО в столбце ФИО_именительный\n'
                                             'Фамилии,имена,отчества состоящие из нескольких слов пишите через дефис!')
        quit()

def combine_all_docx(filename_master, files_lst,name_doc):
    """
    Функция для объединения файлов Word взято отсюда
    https://stackoverflow.com/questions/24872527/combine-word-document-using-python-docx
    :param filename_master: базовый файл
    :param files_list: список с созданными файлами
    :param name_doc :желаемое название файла
    :return: итоговый файл
    """
    #Получаем текущее время
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    number_of_sections = len(files_lst)
    # Открываем и обрабатываем базовый файл
    master = Document(filename_master)
    composer = Composer(master)
    # Перебираем и добавляем файлы к базовому
    for i in range(0, number_of_sections):
        doc_temp = Document(files_lst[i])
        composer.append(doc_temp)
    # Сохраняем файл
    composer.save(f"{path_to_end_folder_doc}/{name_doc} Объединеный файл от {current_time}.docx")


def generate_docs_dpo():
    """
    Функция для создания ддокументов по ДПО
    :return:
    """
    try:
        # Получаем название для создаваемых документов
        name_doc = entry_name_file.get()
        # Считываем данные с листа ДПО в указанной таблице
        df = pd.read_excel(name_file_data_doc, sheet_name='ДПО')
        # Преобразуем столбцы с датой в правильный формат день.месяц.год, так пандас при считывании приводит к формату год месяц день
        df['Дата_рождения_получателя'] = pd.to_datetime(df['Дата_рождения_получателя'],dayfirst=True,errors='coerce')
        df['Дата_выдачи_паспорта'] = pd.to_datetime(df['Дата_выдачи_паспорта'],dayfirst=True,errors='coerce')
        # Конвертируем

        df['Дата_рождения_получателя'] = df['Дата_рождения_получателя'].apply(convert_date)

        df['Дата_выдачи_паспорта'] = df['Дата_выдачи_паспорта'].apply(convert_date)
        #Добавляем столбец инициалы
        df['Инициалы'] = df['ФИО_именительный'].apply(create_initials)

        # Добавляем столбцы дата начала и дата окончания обучения


        df['Дата_начала_обучения'] = df['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_begin_course)
        df['Дата_окончания_обучения'] = df['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_end_course)

        # Конвертируем датафрейм в список словарей
        data = df.to_dict('records')

        # Создаем переменную для типа создаваемого документа
        status_rb_type_doc = group_rb_type_doc.get()
        # Создаем переменную для получения состояния чекбокса объединения файлов.
        mode_combine = mode_combine_value.get()
        # если статус == 0 то создаем индивидуальные приказы по количеству строк.30 строк-30 документов

        if status_rb_type_doc == 0:
            if mode_combine == 'No':
                # Создаем отдельные файлы
                for row in data:
                    doc = DocxTemplate(name_file_template_doc)
                    context = row
                    # Превращаем строку в список кортежей, где первый элемент кортежа это ключ а второй данные

                    doc.render(context)
                    t = time.localtime()
                    current_time = time.strftime('%H_%M_%S', t)

                    doc.save(f'{path_to_end_folder_doc}/{name_doc} {row["ФИО_именительный"]} от {current_time}.docx')

            else:
                # Список с созданными файлами
                files_lst = []
                # Создаем временную папку
                with tempfile.TemporaryDirectory() as tmpdirname:
                    print('created temporary directory', tmpdirname)
                    # Создаем и сохраняем во временную папку созданные документы Word
                    for row in data:
                        doc = DocxTemplate(name_file_template_doc)
                        context = row
                        doc.render(context)
                        # Сохраняем файл
                        doc.save(f'{tmpdirname}/{row["ФИО_именительный"]}.docx')
                        # Добавляем путь к файлу в список
                        files_lst.append(f'{tmpdirname}/{row["ФИО_именительный"]}.docx')
                    # Получаем базовый файл
                    main_doc = files_lst.pop(0)
                    # Запускаем функцию
                    combine_all_docx(main_doc, files_lst,name_doc)

        else:

            # Получаем первую строку таблицы, предполагая что раз это групповой список то и данные будут совпадать
            context = data[0]
            # # Добавляем в словарь context словарь со списками значений, формата Список_Название колонки:[значения]
            # context.update(lst_main_dict)
            # Добавляем в словарь context полностью весь список словарей data ,чтобы реализовать добавление в одну таблицу данных из разных ключей
            context['lst_items'] = data
            # Загружаем шаблон

            doc = DocxTemplate(name_file_template_doc)
            # Создаем документ
            doc.render(context)
            t = time.localtime()
            current_time = time.strftime('%H_%M_%S', t)
            # сохраняем документ
            # очищаем название под которым будем сохранять документ от кавычек и двоеточий
            name_file_dpo = context["Наименование_дополнительной_профессиональной_программы"].replace(':','').replace('"','').replace("'",'')

            doc.save(
                f'{path_to_end_folder_doc}/{name_doc} {name_file_dpo} от {current_time}.docx')

    except NameError:

        messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    else:
        messagebox.showinfo('ЦОПП Бурятия', 'Создание документов успешно завершено!')


def generate_docs_po():
    """
    Функция для создания документов ПО
    :return:
    """
    try:
        # Получаем название для создаваемых документов
        name_doc = entry_name_file.get()
        # Считываем данные с листа ПО в указанной таблице
        df = pd.read_excel(name_file_data_doc, sheet_name='ПО')
        # Преобразуем столбцы с датой в правильный формат день.месяц.год, так пандас при считывании приводит к формату год месяц день
        df['Дата_рождения_получателя'] = pd.to_datetime(df['Дата_рождения_получателя'],dayfirst=True,errors='coerce')
        df['Дата_выдачи_паспорта'] = pd.to_datetime(df['Дата_выдачи_паспорта'],dayfirst=True,errors='coerce')
        if 'Дата_выдачи_свидетельства_о_рождении' in df.columns:
            # Это сделано чтобы не добавлять в прежние документы эту колонку
            df['Дата_выдачи_свидетельства_о_рождении'] = pd.to_datetime(df['Дата_выдачи_свидетельства_о_рождении'],dayfirst=True,errors='coerce')
            df['Дата_выдачи_свидетельства_о_рождении'] = df['Дата_выдачи_свидетельства_о_рождении'].apply(convert_date)



        df['Дата_рождения_получателя'] = df['Дата_рождения_получателя'].apply(convert_date)
        df['Дата_выдачи_паспорта'] = df['Дата_выдачи_паспорта'].apply(convert_date)

        # Добавляем столбец инициалы
        df['Инициалы'] = df['ФИО_именительный'].apply(create_initials)

        # Добавляем столбцы дата начала и дата окончания обучения
        df['Дата_начала_обучения'] = df['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_begin_course)
        df['Дата_окончания_обучения'] = df['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_end_course)

        # Конвертируем датафрейм в список словарей
        data = df.to_dict('records')
        # Создаем переменную для получения состояния чекбокса объединения файлов.
        mode_combine = mode_combine_value.get()
        # Создаем переменную для типа создаваемого документа
        status_rb_type_doc = group_rb_type_doc.get()
        # если статус == 0 то создаем индивидуальные приказы по количеству строк.30 строк-30 документов
        if status_rb_type_doc == 0:

            try:
                if mode_combine == 'No':
                    # Создаем отдельные файлы
                    for row in data:
                        doc = DocxTemplate(name_file_template_doc)
                        context = row
                        # Превращаем строку в список кортежей, где первый элемент кортежа это ключ а второй данные

                        doc.render(context)
                        t = time.localtime()
                        current_time = time.strftime('%H_%M_%S', t)

                        doc.save(
                            f'{path_to_end_folder_doc}/{name_doc} {row["ФИО_именительный"]} от {current_time}.docx')

                else:
                    # Список с созданными файлами
                    files_lst = []
                    # Создаем временную папку
                    with tempfile.TemporaryDirectory() as tmpdirname:
                        print('created temporary directory', tmpdirname)
                        # Создаем и сохраняем во временную папку созданные документы Word
                        for row in data:
                            doc = DocxTemplate(name_file_template_doc)
                            context = row
                            doc.render(context)
                            # Сохраняем файл
                            doc.save(f'{tmpdirname}/{row["ФИО_именительный"]}.docx')
                            # Добавляем путь к файлу в список
                            files_lst.append(f'{tmpdirname}/{row["ФИО_именительный"]}.docx')
                        # Получаем базовый файл
                        main_doc = files_lst.pop(0)
                        # Запускаем функцию
                        combine_all_docx(main_doc, files_lst,name_doc)

            except KeyError:
                messagebox.showerror('ЦОПП Бурятия', 'Колонка с ФИО должна называться ФИО_именительный')
                quit()
            except:
                messagebox.showerror('ЦОПП Бурятия',
                                     'Проверьте содержимое шаблона\nНе допускаются любые символы кроме _ в словах внутри фигурных скобок\nСлова должны могут быть разделены нижним подчеркиванием\n'
                                     'Шаблоны для документов ДПО и ПО отличаются!!! Таблицы для них отличаются!!!')
                quit()

            else:
                messagebox.showinfo('ЦОПП Бурятия', 'Создание документов успешно завершено!')

        else:

            # Итеруемся по списку словарей, чтобы получить список ФИО
            try:
                # Получаем первую строку таблицы, предполагая что раз это групповой список то и данные будут совпадать
                context = data[0]
                # Добавляем в словарь context полностью весь список словарей data ,чтобы реализовать добавление в одну таблицу данных из разных ключей
                context['lst_items'] = data

                # Загружаем шаблон
                doc = DocxTemplate(name_file_template_doc)
                # Создаем документ
                doc.render(context)
                # Получаем текущее время
                t = time.localtime()
                current_time = time.strftime('%H_%M_%S', t)
                # сохраняем документ
                # очищаем название под которым будем сохранять документ от кавычек и двоеточий
                name_file_po = context["Наименование_программы_профессионального_обучения"].replace(':','').replace('"', '').replace("'",'')

                doc.save(
                    f'{path_to_end_folder_doc}/{name_doc} {name_file_po} от {current_time}.docx')
            except KeyError:
                messagebox.showerror('ЦОПП Бурятия,Колонка с ФИО должна называться ФИО_именительный')
                quit()

            except OSError:
                messagebox.showerror('ЦОПП Бурятия', 'Закройте открытый файл Word')
                quit()
            except:
                messagebox.showerror('ЦОПП Бурятия',
                                     'Проверьте содержимое шаблона\nНе допускаются любые символы кроме _ в словах внутри фигурных скобок\nСлова должны могут быть разделены нижним подчеркиванием')
                quit()
            else:
                messagebox.showinfo('ЦОПП Бурятия', 'Создание документов успешно завершено!')

    except NameError as e:
        messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')


def generate_docs_other():
    """
    Функция для создания документов из произвольных таблиц(т.е. отличающихся от структуры базы данных ЦОПП Бурятия)
    :return:
    """
    try:
        # Получаем название для создаваемых документов
        name_doc = entry_name_file.get()
        # Считываем данные
        df = pd.read_excel(name_file_data_doc,dtype=str)
        # # Обрабатываем колонки с датами, чтобы они отображались корректно
        # получаем первую строку датафрейма
        first_row = df.iloc[0, :]
        lst_first_row = list(first_row)
        lst_date_columns = []
        # Перебираем
        for idx, value in enumerate(lst_first_row):
            result = check_date_columns(idx, value)
            if result:
                lst_date_columns.append(result)
            else:
                continue
        # Конвертируем в пригодный строковый формат
        for i in lst_date_columns:
            df.iloc[:, i] = pd.to_datetime(df.iloc[:, i],errors='coerce',dayfirst=True)
            df.iloc[:, i] = df.iloc[:, i].apply(create_doc_convert_date)


        # Конвертируем датафрейм в список словарей
        data = df.to_dict('records')
        # Создаем счетчик для названий файлов в случае если нет колонки ФИО
        count = 0

        # Создаем переменную для получения состояния чекбокса объединения файлов.
        mode_combine = mode_combine_value.get()
        # Создаем переменную для типа создаваемого документа
        status_rb_type_doc = group_rb_type_doc.get()
        # если статус == 0 то создаем индивидуальные приказы по количеству строк.30 строк-30 документов
        if status_rb_type_doc == 0:
            if mode_combine == 'No':
                # Создаем в цикле документы
                for row in data:
                    doc = DocxTemplate(name_file_template_doc)
                    context = row
                    count += 1

                    t = time.localtime()
                    current_time = time.strftime('%H_%M_%S', t)
                    try:
                        if 'ФИО' in row:
                            doc.render(context)

                            doc.save(f'{path_to_end_folder_doc}/{name_doc} {row["ФИО"]} от {current_time}.docx')
                        else:
                            doc.render(context)

                            doc.save(f'{path_to_end_folder_doc}/{name_doc} {count} от {current_time}.docx')
                    except:
                        messagebox.showerror('ЦОПП Бурятия',
                                             'Проверьте содержимое шаблона\nНе допускаются любые символы кроме _ в словах внутри фигурных скобок\nСлова должны могут быть разделены нижним подчеркиванием')
                        exit()
            else:
                # Список с созданными файлами
                files_lst = []
                # Создаем временную папку
                with tempfile.TemporaryDirectory() as tmpdirname:
                    print('created temporary directory', tmpdirname)
                    # счетчик
                    temp_count= 0
                    # Создаем и сохраняем во временную папку созданные документы Word
                    for row in data:
                        doc = DocxTemplate(name_file_template_doc)
                        context = row
                        doc.render(context)
                        # Сохраняем файл
                        doc.save(f'{tmpdirname}/{temp_count}.docx')
                        # Добавляем путь к файлу в список
                        files_lst.append(f'{tmpdirname}/{temp_count}.docx')
                        temp_count +=1
                    # Получаем базовый файл
                    main_doc = files_lst.pop(0)
                    # Запускаем функцию
                    combine_all_docx(main_doc, files_lst,name_doc)
        else:
            context = data[0]
            # Добавляем в словарь context полностью весь список словарей data ,чтобы реализовать добавление в одну таблицу данных из разных ключей
            context['lst_items'] = data
            doc = DocxTemplate(name_file_template_doc)
            # Создаем документ
            doc.render(context)
            # сохраняем документ
            # генерируем название
            t = time.localtime()
            current_time = time.strftime('%H_%M_%S', t)
            doc.save(
                f'{path_to_end_folder_doc}/{name_doc} {current_time}.docx')
        messagebox.showinfo('ЦОПП Бурятия', 'Создание документов успешно завершено!')
    except NameError as e:
        messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    else:
        messagebox.showinfo('ЦОПП Бурятия', 'Создание документов успешно завершено!')


# Функции для создания отчетов
def create_report_one_pk():
    """
    Функция для создания отчета 1-ПК
    :return:
    """

    try:

        df_dpo = pd.read_excel(name_file_data_report, sheet_name='ДПО')
        # проверяем на наличие данных
        if df_dpo.shape[0] == 0:
            messagebox.showerror('ЦОПП Бурятия', 'Лист с данными ПК не заполнен!')

        # Создаем шрифт которым будем выделять названия таблиц
        font_name_table = Font(name='Arial Black', size=15, italic=True)

        # Создаем файл excel
        wb = openpyxl.Workbook()
        # Создаем лист раздела 1.3
        wb.create_sheet(title='Раздел 1.3', index=0)
        wb.create_sheet(title='2.1 По категориям,ПК и ПП', index=1)
        wb.create_sheet(title='Раздел 2.1 Модульные', index=2)
        wb.create_sheet(title='Раздел 2.1 Женщины', index=3)
        wb.create_sheet(title='Раздел 2.1 28,29,30', index=4)
        wb.create_sheet(title='Раздел 2.2 Общая чис', index=5)
        wb.create_sheet(title='Раздел 2.2 Бюджеты', index=6)
        wb.create_sheet(title='Раздел 2.2 Источник средств', index=7)
        wb.create_sheet(title='Раздел 2.3.1 Програм', index=8)
        wb.create_sheet(title='Раздел 2.3.1 Всего', index=9)
        wb.create_sheet(title='2.3.1 По видам и категориям', index=10)
        wb.create_sheet(title='2.3.1 По видам и образованию', index=11)
        wb.create_sheet(title='2.3.1 По видам и форме обучения', index=12)
        wb.create_sheet(title='Раздел 2.3.2 Програм', index=13)
        wb.create_sheet(title='Раздел 2.3.2 Всего', index=14)
        wb.create_sheet(title='2.3.2 По видам и категориям', index=15)
        wb.create_sheet(title='2.3.2 По видам и образованию', index=16)
        wb.create_sheet(title='2.3.2 По видам и форме обучения', index=17)
        wb.create_sheet(title='Раздел 2.4', index=18)
        wb.create_sheet(title='Раздел 2.5', index=19)
        # Удаляем пустой лист
        del wb['Sheet']

        # Раздел 1.3
        # Количество программ по каждому виду обучения
        # группируем. Так как нам нужны текстовые данные то применяем создаем строку с помощью join
        quantity_program_on_type_provisional = df_dpo.groupby(
            'Наименование_дополнительной_профессиональной_программы').agg({
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка': lambda
                x: ','.join(
                x)})

        # Применяем к полученной серии функцию разделения по запятой. Предполо
        quantity_program_on_type_provisional[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'] = \
            quantity_program_on_type_provisional[
                'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'].apply(
                lambda x: x.split(',')[0])

        df_1_3 = quantity_program_on_type_provisional[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'].value_counts().to_frame()
        # Раздел 1.3

        # переименовываем первую колонку
        df_1_3.rename(columns={
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка': 'Число реализованных программ'},
            inplace=True)

        group_quantity_student_program = df_dpo.groupby(
            ['Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка']).size()

        # Добавляем колонку 5- всего слушателей
        df_1_3['Всего слушателей'] = group_quantity_student_program

        # Подсчет сетевой формы
        df_dpo_network = df_dpo[df_dpo['Использование_сетевой_формы_обучения'] == 'Сетевая форма']

        group_quantity_network_program_provisional = df_dpo_network.groupby(
            'Наименование_дополнительной_профессиональной_программы').agg({
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка': lambda
                x: ','.join(x)})
        # Применяем к полученной серии функцию разделения по запятой. Предполо
        group_quantity_network_program_provisional[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'] = \
            group_quantity_network_program_provisional[
                'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'].apply(
                lambda x: x.split(',')[0])
        df_1_3['Число программ сетевая форма'] = group_quantity_network_program_provisional[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'].value_counts().to_frame()

        # Считаем число слушателей на сетевых программах
        df_1_3['Численость слушателей сетевый программ'] = df_dpo_network.groupby(
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка').agg(
            {'ФИО_именительный': 'count'})

        # Считаем дистанционные технологии. Создаем датафрейм по условию использования ДОТ и ЭО
        df_dpo_distant = df_dpo[
            (df_dpo['Использование_ЭО'] != 'Без применения ЭО') & (df_dpo['Использование_ДОТ'] != 'Без применения ДОТ')]

        group_quantity_distant = df_dpo_distant.groupby('Наименование_дополнительной_профессиональной_программы').agg({
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка': lambda
                x: ','.join(
                x)})

        # Применяем к полученной серии функцию разделения по запятой.
        group_quantity_distant[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'] = \
            group_quantity_distant[
                'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'].apply(
                lambda x: x.split(',')[0])
        df_1_3['Число программ ЭО и ДОТ'] = group_quantity_distant[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'].value_counts().to_frame()

        # Считаем количество слушателей
        df_1_3['Численность слушателей ЭО И ДОТ'] = df_dpo_distant.groupby(
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка').agg(
            {'ФИО_именительный': 'count'})

        # Считаем только тех кто учился на программах исключительно ЭО и ДОТ
        df_dpo_distant_only_dot = df_dpo_distant[(df_dpo_distant['Использование_ЭО'] == 'Исключительно с ЭО') & (
                df_dpo_distant['Использование_ДОТ'] == 'Исключительно с ДОТ')]
        df_1_3['Численность слушателей обученных исключительно только ЭО и ДОТ'] = df_dpo_distant_only_dot.groupby(
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка').agg(
            {'ФИО_именительный': 'count'})

        finish_df_1_3 = df_1_3.reset_index()

        finish_df_1_3.columns = ['Вид образовательных программ',
                                 'Число реализованных образовательных программ всего,единиц',
                                 'Всего слушателей, обученных по программам, человек',
                                 'число программ(из графы 3) реализованных с использованием сетевой формы',
                                 'численность слушателей обученных(из графы 5) по программам с использованием сетевой формы-всего',
                                 'число программ(из графы 3),реализованных с применением ЭО или ДОТ',
                                 'Численность слушателей обученных(из графы 5) по программам с применением ЭО и ДОТ',
                                 'в том числе(из графы 14) с применением исключительно ЭО и ДОТ']

        for r in dataframe_to_rows(finish_df_1_3, index=False, header=True):
            wb['Раздел 1.3'].append(r)

        # Устанавливаем ширину колоноки устанавливаем перенос
        wb['Раздел 1.3'].column_dimensions['A'].width = 30
        wb['Раздел 1.3']['A1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['B'].width = 30
        wb['Раздел 1.3']['B1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['C'].width = 30
        wb['Раздел 1.3']['C1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['D'].width = 30
        wb['Раздел 1.3']['D1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['E'].width = 30
        wb['Раздел 1.3']['E1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['F'].width = 30
        wb['Раздел 1.3']['F1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['G'].width = 30
        wb['Раздел 1.3']['G1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['H'].width = 30
        wb['Раздел 1.3']['H1'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3'].column_dimensions['I'].width = 30
        wb['Раздел 1.3']['I1'].alignment = Alignment(wrap_text=True)

        # раздел 2.1  Сведения о численности слушателей обученных по доп профессиональным программам. Создаем копию исходного датафрема
        df_2_1 = df_dpo.copy()
        # Создаем колонку для удобства подсчета
        df_2_1['for_counting'] = 1.0

        # Считаем раздел 2.1 По категориям,ПК и ПП Название листа Раздел 2.1 По категориям,ПК и ПП

        df_2_1_pkpo = pd.pivot_table(df_2_1, index=['Категория_слушателя', 'Место_работы_слушателя',
                                                    'Является_ли_слушатель_руководителем'],
                                     columns=[
                                         'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                     values=['for_counting'],
                                     aggfunc='sum')

        # Удаляем мультиколонку for counting
        df_2_1_pkpo.columns = df_2_1_pkpo.columns.droplevel()

        # заполняем NaN нулями чтобы просуммировать
        df_2_1_pkpo.fillna(0.0, inplace=True)

        # Считаем в зависимости от количества колонок
        if len(df_2_1_pkpo.columns) == 0:
            df_2_1_pkpo['Всего слушателей_temp'] = np.NaN

        elif len(df_2_1_pkpo.columns) == 1:
            df_2_1_pkpo['Всего слушателей_temp'] = df_2_1_pkpo.iloc[:, 0]
        else:
            df_2_1_pkpo['Всего слушателей_temp'] = df_2_1_pkpo['Повышение квалификации'] + df_2_1_pkpo[
                'Профессиональная переподготовка']

        # Перемещаем колонку Всего слушателей
        df_2_1_pkpo.insert(0, 'Всего слушателей', df_2_1_pkpo['Всего слушателей_temp'])

        # Удаляем временную колонку
        df_2_1_pkpo.drop(columns='Всего слушателей_temp', axis=1, inplace=True)

        df_2_1_pkpo = df_2_1_pkpo.reset_index()

        # Заменяем нули пустыми значениями
        df_2_1_pkpo.replace(0.0, np.NaN, inplace=True)

        wb['2.1 По категориям,ПК и ПП'][f'A1'] = 'Слушателей обученных по дополнительным профессиональным программам'
        wb['2.1 По категориям,ПК и ПП'][f'A1'].font = font_name_table

        # Сохраняем датафрейм с подсчетом по пк и пп в лист Раздел 2.1 По категориям,ПК и ПП
        for r in dataframe_to_rows(df_2_1_pkpo, index=False, header=True):
            if len(r) != 1:
                wb['2.1 По категориям,ПК и ПП'].append(r)

        # Устанавливаем размер колонок листа Раздел 2.1 По категориям,ПК и ПП
        wb['2.1 По категориям,ПК и ПП'].column_dimensions['A'].width = 50
        wb['2.1 По категориям,ПК и ПП'].column_dimensions['B'].width = 50
        wb['2.1 По категориям,ПК и ПП'].column_dimensions['C'].width = 30
        wb['2.1 По категориям,ПК и ПП'].column_dimensions['D'].width = 30
        wb['2.1 По категориям,ПК и ПП'].column_dimensions['E'].width = 30
        wb['2.1 По категориям,ПК и ПП'].column_dimensions['F'].width = 30

        # Считаем слушателей по модульным программам с вариативным выбором название листа
        # Раздел 2.1 Модульные по категориям,ПК и ПП

        df_2_1_module_pkpo_yes = df_2_1[df_2_1['Модульная_программа_с_вариативным_выбором'] == 'да']

        df_2_1_module_pkpo = pd.pivot_table(df_2_1_module_pkpo_yes,
                                            index=['Категория_слушателя', 'Место_работы_слушателя',
                                                   'Является_ли_слушатель_руководителем'],
                                            columns=[
                                                'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                            values=['for_counting'],
                                            aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_module_pkpo.columns = df_2_1_module_pkpo.columns.droplevel()

        # Заполняем NaN
        df_2_1_module_pkpo.fillna(0.0, inplace=True)

        # Считаем в зависимости от количества колонок
        if len(df_2_1_module_pkpo.columns) == 0:
            df_2_1_module_pkpo['Всего слушателей_temp'] = np.NaN
        elif len(df_2_1_module_pkpo.columns) == 1:
            df_2_1_module_pkpo['Всего слушателей_temp'] = df_2_1_module_pkpo.iloc[:, 0]
        else:
            df_2_1_module_pkpo['Всего слушателей_temp'] = df_2_1_module_pkpo['Повышение квалификации'] + \
                                                          df_2_1_module_pkpo[
                                                              'Профессиональная переподготовка']

        # перемещаем колонку всего слушателй_temp
        df_2_1_module_pkpo.insert(0, 'Всего слушателей', df_2_1_module_pkpo['Всего слушателей_temp'])

        # Удаляем лишнюю колонку
        df_2_1_module_pkpo.drop(columns='Всего слушателей_temp', axis=1, inplace=True)

        df_2_1_module_pkpo = df_2_1_module_pkpo.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_module_pkpo.replace(0.0, np.NaN, inplace=True)

        wb['Раздел 2.1 Модульные'][f'A1'] = 'Обучено слушателей по модульным программам с вариативным выбором'
        wb['Раздел 2.1 Модульные'][f'A1'].font = font_name_table

        # Сохраняем датафрейм с подсчетом слушателей модульных программ  пк и пп  в лист Раздел 2.1
        for r in dataframe_to_rows(df_2_1_module_pkpo, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 Модульные'].append(r)
        wb['Раздел 2.1 Модульные'].column_dimensions['A'].width = 50
        wb['Раздел 2.1 Модульные'].column_dimensions['B'].width = 50
        wb['Раздел 2.1 Модульные'].column_dimensions['C'].width = 50
        wb['Раздел 2.1 Модульные'].column_dimensions['D'].width = 50
        wb['Раздел 2.1 Модульные'].column_dimensions['E'].width = 50
        wb['Раздел 2.1 Модульные'].column_dimensions['F'].width = 50

        # Считаем слушателей женщин
        # Раздел 2.1 Женщины

        wb['Раздел 2.1 Женщины'][f'A1'] = 'Обучено женщин слушателей'
        wb['Раздел 2.1 Женщины'][f'A1'].font = font_name_table

        # Считаем женщин
        df_2_1_women_yes = df_2_1[df_2_1['Пол_получателя'] == 'Жен']

        df_2_1_women = pd.pivot_table(df_2_1_women_yes,
                                      index=['Категория_слушателя', 'Место_работы_слушателя',
                                             'Является_ли_слушатель_руководителем'],
                                      columns=[
                                          'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                      values=['for_counting'],
                                      aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_women.columns = df_2_1_women.columns.droplevel()

        # Заполняем NaN
        df_2_1_women.fillna(0.0, inplace=True)

        # Считаем в зависимости от количества колонок
        if len(df_2_1_women.columns) == 0:
            df_2_1_women['Всего слушателей_temp'] = np.NaN

        elif len(df_2_1_women.columns) == 1:
            df_2_1_women['Всего слушателей_temp'] = df_2_1_women.iloc[:, 0]
        else:
            df_2_1_women['Всего слушателей_temp'] = df_2_1_women['Повышение квалификации'] + df_2_1_women[
                'Профессиональная переподготовка']

        # перемещаем колонку всего слушателй_temp
        df_2_1_women.insert(0, 'Всего женщин-слушателей', df_2_1_women['Всего слушателей_temp'])

        # Удаляем лишнюю колонку
        df_2_1_women.drop(columns='Всего слушателей_temp', axis=1, inplace=True)

        df_2_1_women = df_2_1_women.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_women.replace(0.0, np.NaN, inplace=True)

        # Сохраняем датафрейм с подсчетом женщин  по пк и пп в лист Раздел 2.1
        for r in dataframe_to_rows(df_2_1_women, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 Женщины'].append(r)

        # Устанавливаем размер колонок листа Раздел 2.1 Женщины
        wb['Раздел 2.1 Женщины'].column_dimensions['A'].width = 50
        wb['Раздел 2.1 Женщины'].column_dimensions['B'].width = 50
        wb['Раздел 2.1 Женщины'].column_dimensions['C'].width = 30
        wb['Раздел 2.1 Женщины'].column_dimensions['D'].width = 30
        wb['Раздел 2.1 Женщины'].column_dimensions['E'].width = 30
        wb['Раздел 2.1 Женщины'].column_dimensions['F'].width = 30

        # Создаем датафрейм для строк 28 и 29 по обычным программам

        # Раздел 2.1 Стр 28,29,30
        # Создаем список категорий слушателей которых нужно посчитать
        lst_cat = ['работник предприятия или организации', 'работник образовательной организации',
                   'лицо, замещающее государственную должность или должность ГГС'
            , 'лицо,замещающее муниципальную должность или должность муниципальной службы',
                   'лицо,уволенное с военной службы',
                   'незанятое лицо по направлению СЗ', 'другое']

        df_2_1_28_29_base = df_2_1.loc[df_2_1['Категория_слушателя'].isin(lst_cat)]

        df_2_1_28_29 = pd.pivot_table(df_2_1_28_29_base, index=['Уровень_образования_ВО_СПО'],
                                      columns=[
                                          'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                      values=['for_counting'],
                                      aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_28_29.columns = df_2_1_28_29.columns.droplevel()

        # Заполняем NaN
        df_2_1_28_29.fillna(0.0, inplace=True)
        # Считаем в зависимости от количества колонок
        if len(df_2_1_28_29.columns) == 0:
            df_2_1_28_29['Всего слушателей_temp'] = np.NaN
        elif len(df_2_1_28_29.columns) == 1:
            df_2_1_28_29['Всего слушателей_temp'] = df_2_1_28_29.iloc[:, 0]
        else:
            df_2_1_28_29['Всего слушателей_temp'] = df_2_1_28_29['Повышение квалификации'] + df_2_1_28_29[
                'Профессиональная переподготовка']

        # перемещаем колонку всего слушателй_temp
        df_2_1_28_29.insert(0, 'Всего слушателей', df_2_1_28_29['Всего слушателей_temp'])
        # Удаляем лишнюю колонку
        df_2_1_28_29.drop(columns='Всего слушателей_temp', axis=1, inplace=True)
        df_2_1_28_29 = df_2_1_28_29.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_28_29.replace(0.0, np.NaN, inplace=True)
        # Добавляем заголовок
        wb['Раздел 2.1 28,29,30'][
            f'A1'] = 'Образование слушателей обученных по дополнительным профессиональным программам'
        wb['Раздел 2.1 28,29,30'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_1_28_29, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 28,29,30'].append(r)

        # # Создаем датафрейм для строк 28 и 29 по модульным программам с вариативным выбором
        df_2_1_28_29_module_base = df_2_1_28_29_base[
            df_2_1_28_29_base['Модульная_программа_с_вариативным_выбором'] == 'да']

        df_2_1_28_29_module = pd.pivot_table(df_2_1_28_29_module_base, index=['Уровень_образования_ВО_СПО'],
                                             columns=[
                                                 'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                             values=['for_counting'],
                                             aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_28_29_module.columns = df_2_1_28_29_module.columns.droplevel()

        # Заполняем NaN
        df_2_1_28_29_module.fillna(0.0, inplace=True)
        # Считаем в зависимости от количества колонок
        if len(df_2_1_28_29_module.columns) == 0:
            df_2_1_28_29_module['Всего слушателей_temp'] = np.NaN
        elif len(df_2_1_28_29_module.columns) == 1:
            df_2_1_28_29_module['Всего слушателей_temp'] = df_2_1_28_29_module.iloc[:, 0]
        else:
            df_2_1_28_29_module['Всего слушателей_temp'] = df_2_1_28_29_module['Повышение квалификации'] + \
                                                           df_2_1_28_29_module[
                                                               'Профессиональная переподготовка']
        # перемещаем колонку всего слушателй_temp
        df_2_1_28_29_module.insert(0, 'Всего слушателей', df_2_1_28_29_module['Всего слушателей_temp'])
        # Удаляем лишнюю колонку
        df_2_1_28_29_module.drop(columns='Всего слушателей_temp', axis=1, inplace=True)
        df_2_1_28_29_module = df_2_1_28_29_module.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_28_29_module.replace(0.0, np.NaN, inplace=True)
        # Создаем промежуток
        max_row_28_29_module = wb['Раздел 2.1 28,29,30'].max_row
        wb['Раздел 2.1 28,29,30'][
            f'A{max_row_28_29_module + 2}'] = 'Образование слушателей обученных по модульным программам с вариативным выбором'
        wb['Раздел 2.1 28,29,30'][f'A{max_row_28_29_module + 2}'].font = font_name_table

        for r in dataframe_to_rows(df_2_1_28_29_module, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 28,29,30'].append(r)

        # Создаем датафрейм 28 и 29 для женщин
        df_2_1_28_29_women_base = df_2_1_28_29_base[df_2_1_28_29_base['Пол_получателя'] == 'Жен']

        df_2_1_28_29_women = pd.pivot_table(df_2_1_28_29_women_base, index=['Уровень_образования_ВО_СПО'],
                                            columns=[
                                                'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                            values=['for_counting'],
                                            aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_28_29_women.columns = df_2_1_28_29_women.columns.droplevel()

        # Заполняем NaN
        df_2_1_28_29_women.fillna(0.0, inplace=True)
        # Считаем в зависимости от количества колонок
        if len(df_2_1_28_29_women.columns) == 0:
            df_2_1_28_29_women['Всего слушателей_temp'] = np.NaN
        elif len(df_2_1_28_29_women.columns) == 1:
            df_2_1_28_29_women['Всего слушателей_temp'] = df_2_1_28_29_women.iloc[:, 0]
        else:
            df_2_1_28_29_women['Всего слушателей_temp'] = df_2_1_28_29_women['Повышение квалификации'] + \
                                                          df_2_1_28_29_women[
                                                              'Профессиональная переподготовка']
        # перемещаем колонку всего слушателй_temp
        df_2_1_28_29_women.insert(0, 'Всего женщин-слушателей', df_2_1_28_29_women['Всего слушателей_temp'])
        # Удаляем лишнюю колонку
        df_2_1_28_29_women.drop(columns='Всего слушателей_temp', axis=1, inplace=True)
        df_2_1_28_29_women = df_2_1_28_29_women.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_28_29_women.replace(0.0, np.NaN, inplace=True)

        # Создаем промежуток
        max_row_28_29_women = wb['Раздел 2.1 28,29,30'].max_row
        wb['Раздел 2.1 28,29,30'][
            f'A{max_row_28_29_women + 2}'] = 'Образование женщин обученных по дополнительным профессиональным программам'
        wb['Раздел 2.1 28,29,30'][f'A{max_row_28_29_women + 2}'].font = font_name_table

        for r in dataframe_to_rows(df_2_1_28_29_women, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 28,29,30'].append(r)

        # Создаем датафреймы для строки 30
        # Создаем список категорий слушателей которых нужно посчитать
        lst_cat_30 = ['работник предприятия или организации', 'работник образовательной организации',
                      'лицо, замещающее государственную должность или должность ГГС'
            , 'лицо,замещающее муниципальную должность или должность муниципальной службы',
                      'лицо,уволенное с военной службы',
                      'незанятое лицо по направлению СЗ', 'студент ВО', 'другое']

        df_2_1_30_base = df_2_1.loc[df_2_1['Категория_слушателя'].isin(lst_cat_30)]

        df_2_1_30_base = df_2_1_30_base[df_2_1_30_base['Для_освоения_ДПП_требуется_наличие_ВО'] == 'требуется ВО']

        df_2_1_30 = pd.pivot_table(df_2_1_30_base, index=['Уровень_образования_ВО_СПО'],
                                   columns=[
                                       'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                   values=['for_counting'],
                                   aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_30.columns = df_2_1_30.columns.droplevel()

        # Заполняем NaN
        df_2_1_30.fillna(0.0, inplace=True)
        # Считаем в зависимости от количества колонок
        if len(df_2_1_30.columns) == 0:
            df_2_1_30['Всего слушателей_temp'] = np.NaN
        elif len(df_2_1_30.columns) == 1:
            df_2_1_30['Всего слушателей_temp'] = df_2_1_30.iloc[:, 0]
        else:
            df_2_1_30['Всего слушателей_temp'] = df_2_1_30['Повышение квалификации'] + df_2_1_30[
                'Профессиональная переподготовка']

        # перемещаем колонку всего слушателй_temp
        df_2_1_30.insert(0, 'Всего слушателей', df_2_1_30['Всего слушателей_temp'])
        # Удаляем лишнюю колонку
        df_2_1_30.drop(columns='Всего слушателей_temp', axis=1, inplace=True)
        df_2_1_30 = df_2_1_30.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_30.replace(0.0, np.NaN, inplace=True)

        # Создаем промежуток
        max_row_30 = wb['Раздел 2.1 28,29,30'].max_row
        wb['Раздел 2.1 28,29,30'][
            f'A{max_row_30 + 5}'] = 'Всего слушателей обученных по программам для освоения которых требуется ВО'
        wb['Раздел 2.1 28,29,30'][f'A{max_row_30 + 5}'].font = font_name_table

        for r in dataframe_to_rows(df_2_1_30, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 28,29,30'].append(r)

        # создаем датафрейм строки 30 для модульных программ
        df_2_1_30_module_base = df_2_1_30_base[df_2_1_30_base['Модульная_программа_с_вариативным_выбором'] == 'да']

        df_2_1_30_module = pd.pivot_table(df_2_1_30_module_base, index=['Уровень_образования_ВО_СПО'],
                                          columns=[
                                              'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                          values=['for_counting'],
                                          aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_30_module.columns = df_2_1_30_module.columns.droplevel()

        # Заполняем NaN
        df_2_1_30_module.fillna(0.0, inplace=True)
        # создаем датафрейм строки 30 для модульных программ
        df_2_1_30_module_base = df_2_1_30_base[df_2_1_30_base['Модульная_программа_с_вариативным_выбором'] == 'да']

        df_2_1_30_module = pd.pivot_table(df_2_1_30_module_base, index=['Уровень_образования_ВО_СПО'],
                                          columns=[
                                              'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                          values=['for_counting'],
                                          aggfunc='sum')

        # Удаляем мультиколонку
        df_2_1_30_module.columns = df_2_1_30_module.columns.droplevel()

        # Заполняем NaN
        df_2_1_30_module.fillna(0.0, inplace=True)
        # Считаем в зависимости от количества колонок
        if len(df_2_1_30_module.columns) == 0:
            df_2_1_30_module['Всего слушателей_temp'] = np.NaN

        elif len(df_2_1_30_module.columns) == 1:
            df_2_1_30_module['Всего слушателей_temp'] = df_2_1_30_module.iloc[:, 0]
        else:
            df_2_1_30_module['Всего слушателей_temp'] = df_2_1_30_module['Повышение квалификации'] + df_2_1_30_module[
                'Профессиональная переподготовка']

        # перемещаем колонку всего слушателй_temp
        df_2_1_30_module.insert(0, 'Всего слушателей', df_2_1_30_module['Всего слушателей_temp'])
        # Удаляем лишнюю колонку
        df_2_1_30_module.drop(columns='Всего слушателей_temp', axis=1, inplace=True)
        df_2_1_30_module = df_2_1_30_module.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_30_module.replace(0.0, np.NaN, inplace=True)

        # Создаем промежуток
        max_row_30_module = wb['Раздел 2.1 28,29,30'].max_row
        wb['Раздел 2.1 28,29,30'][
            f'A{max_row_30_module + 2}'] = 'Всего слушателей обученных по модульным программам для освоения которых требуется ВО'
        wb['Раздел 2.1 28,29,30'][f'A{max_row_30_module + 2}'].font = font_name_table

        for r in dataframe_to_rows(df_2_1_30_module, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 28,29,30'].append(r)
        # создаем датафрейм женщин по строке 30
        df_2_1_30_women_base = df_2_1_30_base[df_2_1_30_base['Пол_получателя'] == 'Жен']

        df_2_1_30_women = pd.pivot_table(df_2_1_30_women_base, index=['Уровень_образования_ВО_СПО'],
                                         columns=[
                                             'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                         values=['for_counting'],
                                         aggfunc='sum')

        # Удаляем мультиколонку

        # Заполняем NaN
        df_2_1_30_women.fillna(0.0, inplace=True)
        # Считаем в зависимости от количества колонок
        if len(df_2_1_30_women.columns) == 0:
            df_2_1_30_women['Всего слушателей_temp'] = np.NaN
        elif len(df_2_1_30_women.columns) == 1:
            df_2_1_30_women['Всего слушателей_temp'] = df_2_1_30_women.iloc[:, 0]
        else:
            df_2_1_30_women['Всего слушателей_temp'] = df_2_1_30_women['Повышение квалификации'] + df_2_1_30_women[
                'Профессиональная переподготовка']

        # перемещаем колонку всего слушателй_temp
        df_2_1_30_women.insert(0, 'Всего женщин-слушателей', df_2_1_30_women['Всего слушателей_temp'])
        # Удаляем лишнюю колонку
        df_2_1_30_women.drop(columns='Всего слушателей_temp', axis=1, inplace=True)
        df_2_1_30_women = df_2_1_30_women.reset_index()
        # Заменяем нули пустыми значениями
        df_2_1_30_women.replace(0.0, np.NaN, inplace=True)

        # Создаем промежуток
        max_row_30_women = wb['Раздел 2.1 28,29,30'].max_row
        wb['Раздел 2.1 28,29,30'][
            f'A{max_row_30_women + 2}'] = 'Всего женщин обученных по  программам для освоения которых требуется ВО'
        wb['Раздел 2.1 28,29,30'][f'A{max_row_30_women + 2}'].font = font_name_table

        for r in dataframe_to_rows(df_2_1_30_women, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.1 28,29,30'].append(r)

        # Устанавливаем размер колонок раздела 2.1 строки 28 и 29
        wb['Раздел 2.1 28,29,30'].column_dimensions['A'].width = 50
        wb['Раздел 2.1 28,29,30'].column_dimensions['B'].width = 50
        wb['Раздел 2.1 28,29,30'].column_dimensions['C'].width = 50
        wb['Раздел 2.1 28,29,30'].column_dimensions['D'].width = 50

        # Считаем раздел 2.2
        df_2_2 = df_dpo.copy()
        df_2_2['for_counting'] = 1

        # Создаем  Раздел 2.2 Общая чис

        df_2_2_all = pd.pivot_table(df_2_2,
                                    index='Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка',
                                    columns=['Источник_финансирования_обучения'],
                                    values=['for_counting'],
                                    aggfunc='sum')

        # Удаляем мультиколонку для общей таблицы
        df_2_2_all.columns = df_2_2_all.columns.droplevel()
        # Исправляем индексы
        df_2_2_all = df_2_2_all.reset_index()

        # Записываем в раздел 2.2
        wb['Раздел 2.2 Общая чис'][
            f'A1'] = 'Общая численность обученных по источнику финансирования по колонкам 4,5,6,7,12,13,14,15'
        wb['Раздел 2.2 Общая чис'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_2_all, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.2 Общая чис'].append(r)

        # Считаем общую сумму по колонкам 8,9,10
        df_2_2_all_8910 = pd.pivot_table(df_2_2, index=[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                         columns=[
                                             'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                         values=['for_counting'],
                                         aggfunc='sum')

        # Удаляем мультиколонку для общей таблицы
        df_2_2_all_8910.columns = df_2_2_all_8910.columns.droplevel()
        # Исправляем индексы
        df_2_2_all_8910 = df_2_2_all_8910.reset_index()

        # Записываем в раздел 2.2
        # Создаем промежуток
        max_row_2_2 = wb['Раздел 2.2 Общая чис'].max_row
        wb['Раздел 2.2 Общая чис'][f'A{max_row_2_2 + 2}'] = 'Общая численность обученных по колонкам 8,9,10,16,17,18'
        wb['Раздел 2.2 Общая чис'][f'A{max_row_2_2 + 2}'].font = font_name_table

        for r in dataframe_to_rows(df_2_2_all_8910, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.2 Общая чис'].append(r)

        wb['Раздел 2.2 Общая чис'].column_dimensions['A'].width = 50
        wb['Раздел 2.2 Общая чис'].column_dimensions['B'].width = 50
        wb['Раздел 2.2 Общая чис'].column_dimensions['C'].width = 50
        wb['Раздел 2.2 Общая чис'].column_dimensions['D'].width = 50
        wb['Раздел 2.2 Общая чис'].column_dimensions['E'].width = 50
        wb['Раздел 2.2 Общая чис'].column_dimensions['F'].width = 50

        # Создаем лист для Раздел 2.2 Бюджеты
        # Создаем сводную таблицу для колонок 4,5,6,7,12,13,14,15
        df_2_2_budget = pd.pivot_table(df_2_2, index=[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка',
            'Категория_слушателя', 'Является_ли_слушатель_руководителем'],
                                       columns=['Источник_финансирования_обучения'],
                                       values=['for_counting'],
                                       aggfunc='sum')

        # Удаляем мультиколонку
        df_2_2_budget.columns = df_2_2_budget.columns.droplevel()

        df_2_2_budget = df_2_2_budget.reset_index()

        # Записываем в раздел 2.2

        wb['Раздел 2.2 Бюджеты'][f'A1'] = 'Численность обученных по категориям и по колонкам 4,5,6,7,12,13,14,15 '
        wb['Раздел 2.2 Бюджеты'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_2_budget, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.2 Бюджеты'].append(r)
        wb['Раздел 2.2 Бюджеты'].column_dimensions['A'].width = 50
        wb['Раздел 2.2 Бюджеты'].column_dimensions['B'].width = 50
        wb['Раздел 2.2 Бюджеты'].column_dimensions['C'].width = 50
        wb['Раздел 2.2 Бюджеты'].column_dimensions['D'].width = 50
        wb['Раздел 2.2 Бюджеты'].column_dimensions['E'].width = 50
        wb['Раздел 2.2 Бюджеты'].column_dimensions['F'].width = 50
        wb['Раздел 2.2 Бюджеты'].column_dimensions['G'].width = 50

        # Раздел 2.2 Источник средств

        # Считаем по колонкам 8,9,10,16,17,18
        df_2_2_budget_8910 = pd.pivot_table(df_2_2, index=[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка',
            'Категория_слушателя', 'Является_ли_слушатель_руководителем'],
                                            columns=[
                                                'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                            values=['for_counting'],
                                            aggfunc='sum')

        # Удаляем мультиколонку
        df_2_2_budget_8910.columns = df_2_2_budget_8910.columns.droplevel()

        df_2_2_budget_8910 = df_2_2_budget_8910.reset_index()

        wb['Раздел 2.2 Источник средств'][f'A1'] = 'Численность обученных по категориям и по колонкам 8,9,10,16,17,18'
        wb['Раздел 2.2 Источник средств'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_2_budget_8910, index=False, header=True):
            if len(r) != 1:
                wb['Раздел 2.2 Источник средств'].append(r)

        # Устанавливаем размер колонок в разделе 2.2
        wb['Раздел 2.2 Источник средств'].column_dimensions['A'].width = 50
        wb['Раздел 2.2 Источник средств'].column_dimensions['B'].width = 50
        wb['Раздел 2.2 Источник средств'].column_dimensions['C'].width = 50
        wb['Раздел 2.2 Источник средств'].column_dimensions['D'].width = 50
        wb['Раздел 2.2 Источник средств'].column_dimensions['E'].width = 50
        wb['Раздел 2.2 Источник средств'].column_dimensions['F'].width = 50
        wb['Раздел 2.2 Источник средств'].column_dimensions['G'].width = 50

        # Создаем раздел 2.3.1
        df_2_3_1_base = df_dpo.copy()
        df_2_3_1_base['for_counting'] = 1
        df_2_3_1 = df_2_3_1_base[df_2_3_1_base[
                                     'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'] == 'Повышение квалификации']

        # Создаем лист Раздел 2.3.1 Програм
        # считаем количество программ
        # группируем. Так как нам нужны текстовые данные то применяем создаем строку с помощью join
        quantity_program_on_econom = df_2_3_1.groupby('Наименование_дополнительной_профессиональной_программы').agg(
            {'Вид_экономической_деятельности_дополнительной_профессиональной_программы': lambda x: ','.join(x)})

        # Применяем к полученной серии функцию разделения по запятой. Предполо
        quantity_program_on_econom['Вид_экономической_деятельности_дополнительной_профессиональной_программы'] = \
            quantity_program_on_econom[
                'Вид_экономической_деятельности_дополнительной_профессиональной_программы'].apply(
                lambda x: x.split(',')[0])

        df_2_3_1_quantity_program = quantity_program_on_econom[
            'Вид_экономической_деятельности_дополнительной_профессиональной_программы'].value_counts().to_frame()

        # переименовываем индекс и колонку
        df_2_3_1_quantity_program.index.name = 'Вид экономической деятельности'
        df_2_3_1_quantity_program.columns = ['Количество программ']

        # Записываем в раздел 2.3.1
        wb['Раздел 2.3.1 Програм'][
            f'A1'] = 'Число реализованных программ повышения квалификации по видам экономической деятельности'
        wb['Раздел 2.3.1 Програм'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_3_1_quantity_program, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.3.1 Програм'].append(r)

        # Устанавливаем размер колонок в листе 2.3.1 Програм
        wb['Раздел 2.3.1 Програм'].column_dimensions['A'].width = 70

        # Лист Раздел 2.3.1 Всего
        # Считаем общую сумму колонка 4.

        df_2_3_1_category_sum_all = pd.pivot_table(df_2_3_1, index=[
            'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                                   columns=['Категория_слушателя',
                                                            'Является_ли_слушатель_руководителем'],
                                                   values=['for_counting'],
                                                   aggfunc='sum')

        df_2_3_1_category_sum_all.fillna(0.0, inplace=True)

        # Последовательно убираем 2 мультииндекса
        df_2_3_1_category_sum_all.columns = df_2_3_1_category_sum_all.columns.droplevel()
        df_2_3_1_category_sum_all.columns = df_2_3_1_category_sum_all.columns.droplevel()

        # Заменяем имена колонок
        df_2_3_1_category_sum_all.columns = range(len(df_2_3_1_category_sum_all.columns))

        # Считаем сумму
        df_2_3_1_category_sum_all['Всего'] = df_2_3_1_category_sum_all.iloc[:, :].sum(axis=1)

        df_2_3_1_category_sum_all_out = df_2_3_1_category_sum_all['Всего'].to_frame()

        df_2_3_1_category_sum_all_out.columns = ['Всего слушателей,по видам экономической деятельности']

        # Записываем в раздел 2.3.1

        wb['Раздел 2.3.1 Всего'][f'A1'] = 'Численность слушателей обученных по каждому виду экономической деятельности'
        wb['Раздел 2.3.1 Всего'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_3_1_category_sum_all_out, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.3.1 Всего'].append(r)
        # Устанавливаем размер колонок в листе 2.3.1 Всего
        wb['Раздел 2.3.1 Всего'].column_dimensions['A'].width = 70

        # Раздел 2.3.1 По видам и категориям

        # Считаем суммы по видам и категориям слушателей
        df_2_3_1_category = pd.pivot_table(df_2_3_1,
                                           index=[
                                               'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                           columns=['Категория_слушателя', 'Является_ли_слушатель_руководителем'],
                                           values=['for_counting'],
                                           aggfunc='sum')

        df_2_3_1_category.columns = df_2_3_1_category.columns.droplevel()

        # Записываем в раздел 2.3.1
        if df_2_3_1_category.shape[0] == 0:
            wb['2.3.1 По видам и категориям'][
                f'A1'] = 'Численность КАТЕГОРИЙ слушателей обученных по каждому виду экономической деятельности'
            wb['2.3.1 По видам и категориям'][f'A1'].font = font_name_table
        else:
            wb['2.3.1 По видам и категориям'][
                f'A1'] = 'Численность КАТЕГОРИЙ слушателей обученных по каждому виду экономической деятельности'
            wb['2.3.1 По видам и категориям'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_2_3_1_category, index=True, header=True):
                if len(r) != 1:
                    wb['2.3.1 По видам и категориям'].append(r)
        # Устанавливаем размер колонок в листе 2.3.1 По видам и категориям
        wb['2.3.1 По видам и категориям'].column_dimensions['A'].width = 70

        # Считаем по уровню образованию
        # Раздел 2.3.1 По видам и образованию
        # Создаем список категорий слушателей которых нужно посчитать
        lst_2_3_1_obraz = ['работник предприятия или организации', 'работник образовательной организации',
                           'лицо, замещающее государственную должность или должность ГГС'
            , 'лицо,замещающее муниципальную должность или должность муниципальной службы',
                           'лицо,уволенное с военной службы',
                           'незанятое лицо по направлению СЗ', 'безработный по направлению СЗ', 'другое']

        df_2_3_1_obraz_base = df_2_3_1.loc[df_2_1['Категория_слушателя'].isin(lst_2_3_1_obraz)]

        df_2_3_1_obraz = pd.pivot_table(df_2_3_1_obraz_base,
                                        index=[
                                            'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                        columns=['Уровень_образования_ВО_СПО'],
                                        values=['for_counting'],
                                        aggfunc='sum')

        # Убираем мультииндекс
        df_2_3_1_obraz.columns = df_2_3_1_obraz.columns.droplevel()

        # Записываем в раздел 2.3.1

        wb['2.3.1 По видам и образованию'][
            f'A1'] = 'Численность  слушателей по уровню образования, обученных по каждому виду экономической деятельности'
        wb['2.3.1 По видам и образованию'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_3_1_obraz, index=True, header=True):
            if len(r) != 1:
                wb['2.3.1 По видам и образованию'].append(r)
        wb['2.3.1 По видам и образованию'].column_dimensions['A'].width = 70

        # Раздел 2.3.1 По видам и форме обучения

        # Считаем слушателей по форме обучения
        df_2_3_1_forma_obuch = pd.pivot_table(df_2_3_1, index=[
            'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                              columns=['Форма_обучения'],
                                              values=['for_counting'],
                                              aggfunc='sum')

        # Убираем мультииндекс
        df_2_3_1_forma_obuch.columns = df_2_3_1_forma_obuch.columns.droplevel()

        # Записываем в Раздел 2.3.1 По видам и образованию

        wb['2.3.1 По видам и форме обучения'][
            f'A1'] = 'Численность  слушателей по форме обучения, обученных по каждому виду экономической деятельности'
        wb['2.3.1 По видам и форме обучения'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_2_3_1_forma_obuch, index=True, header=True):
            if len(r) != 1:
                wb['2.3.1 По видам и форме обучения'].append(r)

        # Устанавливаем размер колонок в разделе 2.3.1
        wb['2.3.1 По видам и форме обучения'].column_dimensions['A'].width = 50

        # Создаем раздел 2.3.2
        df_2_3_2_base = df_dpo.copy()
        df_2_3_2_base['for_counting'] = 1
        df_2_3_2 = df_2_3_2_base[df_2_3_2_base[
                                     'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'] == 'Профессиональная переподготовка']
        # Если программ профессиональной переподготовки нет
        if df_2_3_2.shape[0] == 0:
            wb['Раздел 2.3.2 Програм'][
                f'A1'] = 'Число реализованных программ профессиональной переподготовки по видам экономической деятельности'
            wb['Раздел 2.3.2 Програм'][f'A1'].font = font_name_table
            wb['Раздел 2.3.2 Програм'].column_dimensions['A'].width = 50

            wb['Раздел 2.3.2 Всего'][f'A1'] = 'Слушателей обученных по программам профессиональной переподготовки'
            wb['Раздел 2.3.2 Всего'][f'A1'].font = font_name_table
            wb['Раздел 2.3.2 Всего'].column_dimensions['A'].width = 50

            wb['2.3.2 По видам и категориям'][
                f'A1'] = 'Численность КАТЕГОРИЙ слушателей обученных по каждому виду экономической деятельности'
            wb['2.3.2 По видам и категориям'][f'A1'].font = font_name_table
            wb['2.3.2 По видам и категориям'].column_dimensions['A'].width = 50

            wb['2.3.2 По видам и образованию'][
                f'A1'] = 'Численность  слушателей по уровню образования, обученных по каждому виду экономической деятельности'
            wb['2.3.2 По видам и образованию'][f'A1'].font = font_name_table
            wb['2.3.2 По видам и образованию'].column_dimensions['A'].width = 50

            wb['2.3.2 По видам и форме обучения'][
                f'A1'] = 'Численность  слушателей по форме обучения, обученных по каждому виду экономической деятельности'
            wb['2.3.2 По видам и форме обучения'][f'A1'].font = font_name_table
            wb['2.3.2 По видам и форме обучения'].column_dimensions['A'].width = 50

        else:
            # Создаем лист Раздел 2.3.2 Програм

            # считаем количество программ
            # группируем. Так как нам нужны текстовые данные то применяем создаем строку с помощью join
            quantity_program_on_econom = df_2_3_2.groupby('Наименование_дополнительной_профессиональной_программы').agg(
                {'Вид_экономической_деятельности_дополнительной_профессиональной_программы': lambda x: ','.join(x)})

            # Применяем к полученной серии функцию разделения по запятой. Предполо
            quantity_program_on_econom['Вид_экономической_деятельности_дополнительной_профессиональной_программы'] = \
                quantity_program_on_econom[
                    'Вид_экономической_деятельности_дополнительной_профессиональной_программы'].apply(
                    lambda x: x.split(',')[0])

            df_2_3_2_quantity_program = quantity_program_on_econom[
                'Вид_экономической_деятельности_дополнительной_профессиональной_программы'].value_counts().to_frame()

            # переименовываем индекс и колонку
            df_2_3_2_quantity_program.index.name = 'Вид экономической деятельности'
            df_2_3_2_quantity_program.columns = ['Количество программ']

            # Записываем в раздел 2.3.2
            wb['Раздел 2.3.2 Програм'][
                f'A1'] = 'Число реализованных программ профессиональной переподготовки по видам экономической деятельности'
            wb['Раздел 2.3.2 Програм'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_2_3_2_quantity_program, index=True, header=True):
                if len(r) != 1:
                    wb['Раздел 2.3.2 Програм'].append(r)
            wb['Раздел 2.3.2 Програм'].column_dimensions['A'].width = 50

            # Создаем лист Раздел 2.3.2 Всего
            # Считаем общую сумму колонка 4
            df_2_3_2_category_sum_all = pd.pivot_table(df_2_3_2, index=[
                'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                                       columns=['Категория_слушателя',
                                                                'Является_ли_слушатель_руководителем'],
                                                       values=['for_counting'],
                                                       aggfunc='sum')

            df_2_3_2_category_sum_all.fillna(0.0, inplace=True)

            # Последовательно убираем 2 мультииндекса
            df_2_3_2_category_sum_all.columns = df_2_3_2_category_sum_all.columns.droplevel()
            df_2_3_2_category_sum_all.columns = df_2_3_2_category_sum_all.columns.droplevel()

            # Заменяем имена колонок
            df_2_3_2_category_sum_all.columns = range(len(df_2_3_2_category_sum_all.columns))

            # Считаем сумму
            df_2_3_2_category_sum_all['Всего'] = df_2_3_2_category_sum_all.iloc[:, :].sum(axis=1)

            df_2_3_2_category_sum_all_out = df_2_3_2_category_sum_all['Всего'].to_frame()

            df_2_3_2_category_sum_all_out.columns = ['Всего слушателей,по видам экономической деятельности']

            # Записываем в раздел 2.3.2
            # Создаем промежуток

            wb['Раздел 2.3.2 Всего'][
                f'A1'] = 'Численность слушателей обученных по каждому виду экономической деятельности'
            wb['Раздел 2.3.2 Всего'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_2_3_2_category_sum_all_out, index=True, header=True):
                if len(r) != 1:
                    wb['Раздел 2.3.2 Всего'].append(r)
            wb['Раздел 2.3.2 Всего'].column_dimensions['A'].width = 50

            # Создаем лист Раздел 2.3.2 По видам и категориям
            # Считаем суммы по категориям
            df_2_3_2_category = pd.pivot_table(df_2_3_2, index=[
                'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                               columns=['Категория_слушателя', 'Является_ли_слушатель_руководителем'],
                                               values=['for_counting'],
                                               aggfunc='sum')

            df_2_3_2_category.columns = df_2_3_2_category.columns.droplevel()

            # Записываем в раздел 2.3.2

            wb['2.3.2 По видам и категориям'][
                f'A1'] = 'Численность КАТЕГОРИЙ слушателей обученных по каждому виду экономической деятельности'
            wb['2.3.2 По видам и категориям'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_2_3_2_category, index=True, header=True):
                if len(r) != 1:
                    wb['2.3.2 По видам и категориям'].append(r)
            wb['2.3.2 По видам и категориям'].column_dimensions['A'].width = 50

            # Создаем лист Раздел 2.3.2 По видам и образованию

            # Создаем список категорий слушателей которых нужно посчитать
            lst_2_2_obraz = ['работник предприятия или организации', 'работник образовательной организации',
                             'лицо, замещающее государственную должность или должность ГГС'
                , 'лицо,замещающее муниципальную должность или должность муниципальной службы',
                             'лицо,уволенное с военной службы', 'незанятое лицо по направлению СЗ',
                             'безработный по направлению СЗ', 'другое']

            df_2_3_2_obraz_base = df_2_3_2.loc[df_2_1['Категория_слушателя'].isin(lst_2_2_obraz)]

            df_2_3_2_obraz = pd.pivot_table(df_2_3_2_obraz_base,
                                            index=[
                                                'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                            columns=['Уровень_образования_ВО_СПО'],
                                            values=['for_counting'],
                                            aggfunc='sum')

            # Убираем мультииндекс
            df_2_3_2_obraz.columns = df_2_3_2_obraz.columns.droplevel()

            # Записываем в раздел 2.3.1

            wb['2.3.2 По видам и образованию'][
                f'A1'] = 'Численность  слушателей по уровню образования, обученных по каждому виду экономической деятельности'
            wb['2.3.2 По видам и образованию'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_2_3_2_obraz, index=True, header=True):
                if len(r) != 1:
                    wb['2.3.2 По видам и образованию'].append(r)
            wb['2.3.2 По видам и образованию'].column_dimensions['A'].width = 50

            # Создаем Раздел 2.3.2 По видам и форме обучения
            # Считаем слушателей по форме обучения
            df_2_3_2_forma_obuch = pd.pivot_table(df_2_3_2, index=[
                'Вид_экономической_деятельности_дополнительной_профессиональной_программы'],
                                                  columns=['Форма_обучения'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')

            # Убираем мультииндекс
            df_2_3_2_forma_obuch.columns = df_2_3_2_forma_obuch.columns.droplevel()

            # Записываем в раздел 2.3.1

            wb['2.3.2 По видам и форме обучения'][
                f'A1'] = 'Численность  слушателей по форме обучения, обученных по каждому виду экономической деятельности'
            wb['2.3.2 По видам и форме обучения'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_2_3_2_forma_obuch, index=True, header=True):
                if len(r) != 1:
                    wb['2.3.2 По видам и форме обучения'].append(r)

            # Устанавливаем размер колонок в разделе 2.3.1
            wb['2.3.2 По видам и форме обучения'].column_dimensions['A'].width = 50

        # Создаем дополнительную числовую колонку где каждое значение это 1, для удобства агрегирования
        df_2_4 = df_dpo.copy()
        # Добавляем колонку с 1
        df_2_4['for_counting'] = 1

        # Считаем в общем сколько обучено
        df_2_4_all = pd.pivot_table(df_2_4, columns=[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка',
            'Пол_получателя'],
                                    values=['for_counting'],
                                    aggfunc='sum')
        df_2_4_all.index = ['Всего обучено']

        df_2_4_by_age = pd.pivot_table(df_2_4, index=['Возрастная_категория_1ПК'],
                                       columns=[
                                           'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка',
                                           'Пол_получателя'],
                                       values=['for_counting'],
                                       aggfunc='sum')

        df_2_4_by_age.columns = df_2_4_by_age.columns.droplevel()

        # Соединяем 2 датафрейма
        df_2_4_out = pd.concat([df_2_4_all, df_2_4_by_age])

        wb['Раздел 2.4'][f'A1'] = 'Распределение слушателей по возрасту,полу и программам'
        wb['Раздел 2.4'][f'A1'].font = font_name_table
        wb['Раздел 2.4'][
            f'A2'] = 'В таблице отображаются только те возрастные категории которые ЕСТЬ в исходной таблице!'
        wb['Раздел 2.4'][f'A2'].font = font_name_table

        # Сохраняем в лист Раздел 2.4
        for r in dataframe_to_rows(df_2_4_out, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.4'].append(r)
        wb['Раздел 2.4'].column_dimensions['A'].width = 50

        # раздел 2.5 Инвалиды
        df_dpo_2_5 = df_dpo.copy()
        df_dpo_2_5['for_counting'] = 1
        df_dpo_2_5.fillna('Не заполнено')
        df_dpo_2_5 = df_dpo_2_5[(df_dpo_2_5['Сведения_об_ограничении_возможностей_здоровья'] != 'нет ОВЗ') & (
                df_dpo_2_5['Сведения_об_ограничении_возможностей_здоровья'] != 'Не заполнено')]

        if df_dpo_2_5.shape[0] == 0:
            wb['Раздел 2.5'][f'A1'] = 'Обучение лиц с ограниченными возможностями здоровья и инвалидов'
            wb['Раздел 2.5'][f'A1'].font = font_name_table
            wb['Раздел 2.5'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.5'][f'A2'].font = font_name_table
            wb['Раздел 2.5']['A2'].alignment = Alignment(wrap_text=True)
        else:
            # создаем сводную таблицу
            df_dpo_2_5_out = pd.pivot_table(df_dpo_2_5, index=['Сведения_об_ограничении_возможностей_здоровья'],
                                            columns=[
                                                'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'],
                                            values=['for_counting'],
                                            aggfunc='sum')
            df_dpo_2_5_out.columns = df_dpo_2_5_out.columns.droplevel()
            df_dpo_2_5_out['Всего обучено'] = df_dpo_2_5_out.sum(axis=1)
            # Записываем в лист
            wb['Раздел 2.5'][f'A1'] = 'Обучение лиц с ограниченными возможностями здоровья и инвалидов'
            wb['Раздел 2.5'][f'A1'].font = font_name_table
            wb['Раздел 2.5'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.5'][f'A2'].font = font_name_table
            wb['Раздел 2.5']['A2'].alignment = Alignment(wrap_text=True)
            for r in dataframe_to_rows(df_dpo_2_5_out, index=True, header=True):
                if len(r) != 1:
                    wb['Раздел 2.5'].append(r)

            wb['Раздел 2.5'][f'A3'] = 'Наименование показателей'
            wb['Раздел 2.5'].column_dimensions['A'].width = 80
            wb['Раздел 2.5'].column_dimensions['B'].width = 30
            wb['Раздел 2.5']['B3'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.5'].column_dimensions['C'].width = 30
            wb['Раздел 2.5']['C3'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.5'].column_dimensions['D'].width = 30
            wb['Раздел 2.5']['D3'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.5'].column_dimensions['D'].width = 30
            wb['Раздел 2.5']['E3'].alignment = Alignment(wrap_text=True)

        # Считаем колонку с женщинами инвалидами
        df_2_5_women = df_dpo_2_5[df_dpo_2_5['Пол_получателя'] == 'Жен'].groupby(
            'Сведения_об_ограничении_возможностей_здоровья').agg({'for_counting': 'sum'})

        df_2_5_women.columns = ['Количество женщин']

        sum_woman = df_2_5_women['Количество женщин'].sum()

        if sum_woman == 0:
            # т.е если женщин инвалидов нет, то создаем датафрейм с пустыми значениями
            df_2_5_women_all = pd.DataFrame(data=[np.NaN, np.NaN, np.NaN], index=['Всего', 'инвалид', 'Лицо с ОВЗ'],
                                            columns=['Количество женщин'])
        else:
            # в противном случае, считаем общее количество женщин-инвалидов. Создаем маленький датафрейм
            df_2_5_women_all = pd.DataFrame(data=sum_woman, index=['Всего женщин инвалидов'],
                                            columns=['Количество женщин'])
            # Объединяем столбы с женщинами инвалидами
            df_2_5_women_all = pd.concat([df_2_5_women_all, df_2_5_women])

        wb['Раздел 2.5'][f'A9'] = 'Женщины( инвалиды,ОВЗ,ребенок-инвалид)'
        wb['Раздел 2.5'][f'A9'].font = font_name_table

        # Добавляем в раздел 2.5 информацию о женщинах
        for r in dataframe_to_rows(df_2_5_women_all, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.5'].append(r)

        # Получаем текущее время для того чтобы использовать в названии
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder_report}/Часть отчета 1-ПК.xlsx {current_time}.xlsx')

    except NameError:
        messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    else:
        messagebox.showinfo('ЦОПП Бурятия', 'Создание части отчета 1-ПК\nЗавершено!')


def create_report_one_po():
    """
    Функция для создания отчета 1-ПО
    :return:
    """
    try:
        df_po = pd.read_excel(name_file_data_report, sheet_name='ПО',
                              dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str})
        if df_po.shape[0] == 0:
            messagebox.showerror('ЦОПП Бурятия','Лист с данными ПО не заполнен!')

        # Создаем шрифт которым будем выделять названия таблиц
        font_name_table = Font(name='Arial Black', size=15, italic=True)

        # Создаем файл excel
        wb = openpyxl.Workbook()
        # Создаем листы
        wb.create_sheet(title='Раздел 1.3', index=0)
        wb.create_sheet(title='Раздел 2.1.1', index=1)
        wb.create_sheet(title='Раздел 2.1.2', index=2)
        wb.create_sheet(title='Раздел 2.1.3', index=3)
        wb.create_sheet(title='Раздел 2.2', index=4)
        wb.create_sheet(title='Раздел 2.3', index=5)
        wb.create_sheet(title='Раздел 2.4', index=6)
        wb.create_sheet(title='Раздел 2.5', index=7)
        wb.create_sheet(title='Раздел 2.6', index=8)
        # Удаляем пустой лист
        del wb['Sheet']

        df_po_1_3_base = df_po.copy()
        df_po_1_3_base['for_counting'] = 1
        df_po_1_3_base.fillna('Не заполнено', inplace=True)

        # Считаем лист 1.3
        # Считаем количество реализованных программ
        # группируем. Так как нам нужны текстовые данные то применяем создаем строку с помощью join
        po_quantity_program_on_type_provisional = df_po_1_3_base.groupby(
            'Наименование_программы_профессионального_обучения').agg(
            {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})
        # Применяем к полученной серии функцию разделения по запятой. Предполо
        po_quantity_program_on_type_provisional['Программа_профессионального_обучения_направление_подготовки'] = \
            po_quantity_program_on_type_provisional[
                'Программа_профессионального_обучения_направление_подготовки'].apply(
                lambda x: x.split(';')[0])
        df_po_1_3 = po_quantity_program_on_type_provisional[
            'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

        # переименовываем колонку с
        df_po_1_3.rename(
            columns={'Программа_профессионального_обучения_направление_подготовки': 'Число реализованных программ'},
            inplace=True)

        # Сортирируем индекс чтобы было легче добавлять столбцы
        df_po_1_3.sort_index(ascending=False, inplace=True)

        # Считаем количество обученных по каждой программе
        df_po_1_3_quantity_students = df_po_1_3_base.groupby(
            ['Программа_профессионального_обучения_направление_подготовки']).agg({'for_counting': 'sum'})

        df_po_1_3_quantity_students.sort_index(ascending=False, inplace=True)

        # Соединяем 2 датафрейма
        df_po_1_3['Всего обучено'] = df_po_1_3_quantity_students

        # Считаем количество программ с сетевой формой
        # Создаем датафрейм с теми данными о сетевой форме
        df_po_1_3_network_base = df_po_1_3_base[
            df_po_1_3_base['Использование_сетевой_формы_обучения'] == 'Сетевая форма']

        # Считаем количество реализованных программ
        # группируем. Так как нам нужны текстовые данные то применяем создаем строку с помощью join
        po_quantity_program_on_type_network = df_po_1_3_network_base.groupby(
            'Наименование_программы_профессионального_обучения').agg(
            {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})
        # Применяем к полученной серии функцию разделения по запятой. Предполо
        po_quantity_program_on_type_network['Программа_профессионального_обучения_направление_подготовки'] = \
            po_quantity_program_on_type_network['Программа_профессионального_обучения_направление_подготовки'].apply(
                lambda x: x.split(';')[0])
        df_po_1_3_network = po_quantity_program_on_type_network[
            'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

        df_po_1_3_network.rename(
            columns={'Программа_профессионального_обучения_направление_подготовки': 'Число реализованных программ'},
            inplace=True)

        df_po_1_3_network.sort_index(ascending=False, inplace=True)

        df_po_1_3['Число программ с сетевой формой'] = df_po_1_3_network

        # Считаем число слушателей на сетевых программах
        df_po_1_3['Численность слушателей сетевых программ'] = df_po_1_3_network_base.groupby(
            'Программа_профессионального_обучения_направление_подготовки').agg({'for_counting': 'sum'})

        # Считаем электронное обучение
        df_po_1_3_distant_ao = df_po_1_3_base[(df_po_1_3_base['Использование_ЭО'] != 'Без применения ЭО') & (
                    df_po_1_3_base['Использование_ЭО'] != 'Не заполнено')]

        po_group_quantity_distant_ao = df_po_1_3_distant_ao.groupby(
            'Наименование_программы_профессионального_обучения').agg(
            {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})

        # Применяем к полученной серии функцию разделения по запятой.
        po_group_quantity_distant_ao['Программа_профессионального_обучения_направление_подготовки'] = \
            po_group_quantity_distant_ao['Программа_профессионального_обучения_направление_подготовки'].apply(
                lambda x: x.split(';')[0])
        df_po_1_3['Число программ реализуемых с помощью ЭО'] = po_group_quantity_distant_ao[
            'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

        # Считаем количество слушателей
        df_po_1_3['Численность слушателей обученных с применением ЭО'] = df_po_1_3_distant_ao.groupby(
            'Программа_профессионального_обучения_направление_подготовки').agg({'for_counting': 'sum'})

        # Считаем пользователей с исключительно ЭО
        df_po_1_3_distant_only_ao = df_po_1_3_base[(df_po_1_3_base['Использование_ЭО'] == 'Исключительно с ЭО')]

        df_po_1_3['Численность слушателей обученных только с ЭО '] = df_po_1_3_distant_only_ao.groupby(
            'Программа_профессионального_обучения_направление_подготовки').agg({'for_counting': 'sum'})

        df_po_1_3_distant_dot = df_po_1_3_base[(df_po_1_3_base['Использование_ДОТ'] != 'Без применения ДОТ') & (
                    df_po_1_3_base['Использование_ДОТ'] != 'Не заполнено')]

        po_group_quantity_distant_dot = df_po_1_3_distant_dot.groupby(
            'Наименование_программы_профессионального_обучения').agg(
            {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})

        # Применяем к полученной серии функцию разделения по запятой.
        po_group_quantity_distant_dot['Программа_профессионального_обучения_направление_подготовки'] = \
            po_group_quantity_distant_dot['Программа_профессионального_обучения_направление_подготовки'].apply(
                lambda x: x.split(';')[0])
        df_po_1_3['Число программ реализуемых с помощью ДОТ'] = po_group_quantity_distant_dot[
            'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

        df_po_1_3['Численность слушателей обученных с применением ДОТ'] = df_po_1_3_distant_dot.groupby(
            ['Программа_профессионального_обучения_направление_подготовки']).agg({'for_counting': 'sum'})

        # Считаем слушателей обученных только с ДОТ
        df_po_1_3_distant_dot_only = df_po_1_3_base[df_po_1_3_base['Использование_ДОТ'] == 'Исключительно с ДОТ']

        df_po_1_3['Численность слушателей обученных только с  ДОТ'] = df_po_1_3_distant_dot_only.groupby(
            ['Программа_профессионального_обучения_направление_подготовки']).agg({'for_counting': 'sum'})

        df_po_1_3.index.name = 'Вид образовательных программ'

        # Записываем датафрейм на лист
        wb['Раздел 1.3'][f'A1'] = 'Сведения о образовательных программах,реализуемых организацией'
        wb['Раздел 1.3'][f'A1'].font = font_name_table

        for r in dataframe_to_rows(df_po_1_3, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 1.3'].append(r)
        wb['Раздел 1.3'][f'A2'] = 'Наименование образовательных программ'
        wb['Раздел 1.3'].column_dimensions['A'].width = 60
        wb['Раздел 1.3'].column_dimensions['B'].width = 20
        wb['Раздел 1.3'].column_dimensions['D'].width = 20
        wb['Раздел 1.3'].column_dimensions['E'].width = 20
        wb['Раздел 1.3'].column_dimensions['F'].width = 20
        wb['Раздел 1.3'].column_dimensions['G'].width = 20
        wb['Раздел 1.3'].column_dimensions['H'].width = 20
        wb['Раздел 1.3'].column_dimensions['I'].width = 20
        wb['Раздел 1.3'].column_dimensions['J'].width = 20
        wb['Раздел 1.3'].column_dimensions['K'].width = 20

        wb['Раздел 1.3']['B2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['C2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['D2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['E2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['F2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['G2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['H2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['I2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['J2'].alignment = Alignment(wrap_text=True)
        wb['Раздел 1.3']['K2'].alignment = Alignment(wrap_text=True)

        # 2.1.1
        # Создаем раздел 2.1.1
        df_po_2_1_1_base = df_po.copy()
        df_po_2_1_1_base['for_counting'] = 1
        # Отбираем слушателей обученных по программам проф подготовки по профессиям рабочих,должностям служащих
        df_po_2_1_1_base = df_po_2_1_1_base[df_po_2_1_1_base[
                                                'Программа_профессионального_обучения_направление_подготовки'] == 'Программа профессиональной подготовки по профессии рабочего, должности служащего']

        # Проверяем есть ли подходяшие данные
        if df_po_2_1_1_base.shape[0] == 0:
            wb['Раздел 2.1.1'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.1.1'][f'A2'].font = font_name_table
            wb['Раздел 2.1.1']['A2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1'][f'A3'] = 'Код профессии при наличии'
            wb['Раздел 2.1.1'].column_dimensions['A'].width = 50
            wb['Раздел 2.1.1'].column_dimensions['C'].width = 15
            wb['Раздел 2.1.1'].column_dimensions['D'].width = 15

        else:
            # Считаем лист  2.3.1
            # Создаем сводную таблицу для колонок 6,7,8,9
            df_po_2_1_1_6_9 = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                             columns=['Источник_финансирования_обучения'],
                                             values=['for_counting'],
                                             aggfunc='sum')
            df_po_2_1_1_6_9.columns = df_po_2_1_1_6_9.columns.droplevel()
            # Суммируем по столбцам
            df_po_2_1_1_6_9['Всего по источникам финансирования'] = df_po_2_1_1_6_9.sum(axis=1)
            # Создаем сводную таблицу для колонок 13,14,15
            df_po_2_1_1_13_15 = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                               columns=['Форма_обучения'],
                                               values=['for_counting'],
                                               aggfunc='sum')
            df_po_2_1_1_13_15.columns = df_po_2_1_1_13_15.columns.droplevel()

            df_po_2_1_1_13_15['Всего по форме обучения'] = df_po_2_1_1_13_15.sum(axis=1)

            df_po_2_1_1 = pd.concat([df_po_2_1_1_6_9, df_po_2_1_1_13_15], axis=1)
            # Считаем по категориям финансирования
            df_po_2_1_1_cat_finance = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                                     columns=[
                                                         'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                                     values=['for_counting'],
                                                     aggfunc='sum')
            df_po_2_1_1_cat_finance.columns = df_po_2_1_1_cat_finance.columns.droplevel()

            df_po_2_1_1 = pd.concat([df_po_2_1_1, df_po_2_1_1_cat_finance], axis=1)

            # Считаем женщин

            df_po_2_1_1_women = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                               columns=['Пол_получателя'],
                                               values=['for_counting'],
                                               aggfunc='sum')
            df_po_2_1_1_women.columns = df_po_2_1_1_women.columns.droplevel()

            df_po_2_1_1 = pd.concat([df_po_2_1_1, df_po_2_1_1_women], axis=1)

            # Считаем тех кто прошел по индвидуальным учебным планам

            df_po_2_1_1_ind_up = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                                columns=[
                                                    'прошли_ускоренное_обучение_по_индивидуаль-ным_учебным_планам'],
                                                values=['for_counting'],
                                                aggfunc='sum')
            df_po_2_1_1_ind_up.columns = df_po_2_1_1_ind_up.columns.droplevel()

            df_po_2_1_1 = pd.concat([df_po_2_1_1, df_po_2_1_1_ind_up], axis=1)

            # Создаем проверку
            df_po_2_1_1['temp'] = df_po_2_1_1['Всего по источникам финансирования'] == df_po_2_1_1[
                'Всего по форме обучения']

            df_po_2_1_1['Совпадение сумм столбцов 6-9 и 13-15'] = df_po_2_1_1['temp'].apply(
                lambda x: 'СОВПАДАЕТ' if x is True else 'НЕ СОВПАДАЕТ!!!')
            df_po_2_1_1.drop(columns='temp', inplace=True)
            df_po_2_1_1.index.name = 'Код профессии при наличии'

            # Записываем датафрейм на лист
            wb['Раздел 2.1.1'][
                f'A1'] = 'Распределение слушателей,обученных по программам профессиональной подготовки по професииям рабочих,должностям служащих'
            wb['Раздел 2.1.1'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_po_2_1_1, index=True, header=True):
                if len(r) != 1:
                    wb['Раздел 2.1.1'].append(r)
            wb['Раздел 2.1.1'][
                f'A2'] = 'Код професии'
            wb['Раздел 2.1.1'][f'A2'].font = font_name_table
            wb['Раздел 2.1.1']['A2'].alignment = Alignment(wrap_text=True)

            wb['Раздел 2.1.1'].column_dimensions['A'].width = 50
            wb['Раздел 2.1.1'].column_dimensions['C'].width = 15
            wb['Раздел 2.1.1'].column_dimensions['D'].width = 15

            wb['Раздел 2.1.1']['B2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['C2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['D2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['E2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['F2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['G2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['H2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['I2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['J2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['K2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['L2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['M2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.1']['Q2'].alignment = Alignment(wrap_text=True)

        # Считаем раздел 2.1.2
        df_po_2_1_2_base = df_po.copy()
        df_po_2_1_2_base['for_counting'] = 1
        # Отбираем слушателей обученных по программам переподготовки рабочих, служащих
        df_po_2_1_2_base = df_po_2_1_2_base[df_po_2_1_2_base[
                                                'Программа_профессионального_обучения_направление_подготовки'] == 'Программа переподготовки рабочих, служащих']

        # Проверяем размер датафрейма
        # Если он пустой то ничего не считаем а просто записываем в лист пустые строки
        if df_po_2_1_2_base.shape[0] == 0:
            wb['Раздел 2.1.2'][
                f'A1'] = 'Распределение слушателей,обученных по программам переподготовки рабочих,служащих'
            wb['Раздел 2.1.2'][f'A1'].font = font_name_table
            wb['Раздел 2.1.2'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.1.2'][f'A2'].font = font_name_table
            wb['Раздел 2.1.2']['A2'].alignment = Alignment(wrap_text=True)

            wb['Раздел 2.1.2'][f'A3'] = 'Код профессии при наличии'

            wb['Раздел 2.1.2'].column_dimensions['A'].width = 50
            wb['Раздел 2.1.2'].column_dimensions['C'].width = 15
            wb['Раздел 2.1.2'].column_dimensions['D'].width = 15
        else:
            # Создаем сводную таблицу для колонок 6,7,8,9
            df_po_2_1_2_6_9 = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                             columns=['Источник_финансирования_обучения'],
                                             values=['for_counting'],
                                             aggfunc='sum')
            df_po_2_1_2_6_9.columns = df_po_2_1_2_6_9.columns.droplevel()

            df_po_2_1_2_6_9['Всего по источникам финансирования'] = df_po_2_1_2_6_9.sum(axis=1)

            # Создаем сводную таблицу для колонок 13,14,15
            df_po_2_1_2_13_15 = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                               columns=['Форма_обучения'],
                                               values=['for_counting'],
                                               aggfunc='sum')
            df_po_2_1_2_13_15.columns = df_po_2_1_2_13_15.columns.droplevel()

            df_po_2_1_2_13_15['Всего по форме обучения'] = df_po_2_1_2_13_15.sum(axis=1)

            df_po_2_1_2 = pd.concat([df_po_2_1_2_6_9, df_po_2_1_2_13_15], axis=1)

            # Считаем по категориям финансирования
            df_po_2_1_2_cat_finance = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                                     columns=[
                                                         'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                                     values=['for_counting'],
                                                     aggfunc='sum')
            df_po_2_1_2_cat_finance.columns = df_po_2_1_2_cat_finance.columns.droplevel()

            df_po_2_1_2 = pd.concat([df_po_2_1_2, df_po_2_1_2_cat_finance], axis=1)

            # Считаем женщин

            df_po_2_1_2_women = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                               columns=['Пол_получателя'],
                                               values=['for_counting'],
                                               aggfunc='sum')
            df_po_2_1_2_women.columns = df_po_2_1_2_women.columns.droplevel()

            df_po_2_1_2 = pd.concat([df_po_2_1_2, df_po_2_1_2_women], axis=1)

            # Считаем тех кто прошел по индвидуальным учебным планам

            df_po_2_1_2_ind_up = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                                columns=[
                                                    'прошли_ускоренное_обучение_по_индивидуаль-ным_учебным_планам'],
                                                values=['for_counting'],
                                                aggfunc='sum')
            df_po_2_1_2_ind_up.columns = df_po_2_1_2_ind_up.columns.droplevel()

            df_po_2_1_2 = pd.concat([df_po_2_1_2, df_po_2_1_2_ind_up], axis=1)

            # Создаем проверку
            df_po_2_1_2['temp'] = df_po_2_1_2['Всего по источникам финансирования'] == df_po_2_1_2[
                'Всего по форме обучения']

            df_po_2_1_2['Совпадение сумм столбцов 6-9 и 13-15'] = df_po_2_1_2['temp'].apply(
                lambda x: 'СОВПАДАЕТ' if x is True else 'НЕ СОВПАДАЕТ!!!')
            df_po_2_1_2.drop(columns='temp', inplace=True)
            df_po_2_1_2.index.name = 'Код профессии при наличии'

            # Записываем датафрейм на лист
            wb['Раздел 2.1.2'][
                f'A1'] = 'Распределение слушателей,обученных по программам переподготовки рабочих,служащих'
            wb['Раздел 2.1.2'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_po_2_1_2, index=True, header=True):
                if len(r) != 1:
                    wb['Раздел 2.1.2'].append(r)
            wb['Раздел 2.1.2'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.1.2'][f'A2'].font = font_name_table
            wb['Раздел 2.1.2']['A2'].alignment = Alignment(wrap_text=True)

            wb['Раздел 2.1.2'][f'A3'] = 'Код профессии при наличии'

            wb['Раздел 2.1.2'].column_dimensions['A'].width = 50
            wb['Раздел 2.1.2'].column_dimensions['C'].width = 15
            wb['Раздел 2.1.2'].column_dimensions['D'].width = 15
            wb['Раздел 2.1.2']['B2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['C2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['D2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['E2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['F2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['G2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['H2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['I2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['J2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['K2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['L2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['M2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.2']['Q2'].alignment = Alignment(wrap_text=True)

        # Считаем лист 2.1.3
        # Считаем раздел 2.1.2
        df_po_2_1_3_base = df_po.copy()
        df_po_2_1_3_base['for_counting'] = 1

        # Отбираем слушателей обученных по программам переподготовки рабочих, служащих
        df_po_2_1_3_base = df_po_2_1_3_base[df_po_2_1_3_base[
                                                'Программа_профессионального_обучения_направление_подготовки'] == 'Программа повышения квалификации рабочих, служащих']

        if df_po_2_1_3_base.shape[0] == 0:
            wb['Раздел 2.1.3'][
                f'A1'] = 'Распределение слушателей,обученных по программам повышения квалификации рабочих,служащих'
            wb['Раздел 2.1.3'][f'A1'].font = font_name_table
            wb['Раздел 2.1.3'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.1.3'][f'A2'].font = font_name_table
            wb['Раздел 2.1.3']['A2'].alignment = Alignment(wrap_text=True)

            wb['Раздел 2.1.3'][f'A3'] = 'Код профессии при наличии'
            wb['Раздел 2.1.3'].column_dimensions['A'].width = 50
            wb['Раздел 2.1.3'].column_dimensions['C'].width = 15
            wb['Раздел 2.1.3'].column_dimensions['D'].width = 15
        else:

            # Создаем сводную таблицу для колонок 6,7,8,9
            df_po_2_1_3_6_9 = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                             columns=['Источник_финансирования_обучения'],
                                             values=['for_counting'],
                                             aggfunc='sum')
            df_po_2_1_3_6_9.columns = df_po_2_1_3_6_9.columns.droplevel()

            df_po_2_1_3_6_9['Всего по источникам финансирования'] = df_po_2_1_3_6_9.sum(axis=1)

            # Создаем сводную таблицу для колонок 13,14,15
            df_po_2_1_3_13_15 = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                               columns=['Форма_обучения'],
                                               values=['for_counting'],
                                               aggfunc='sum')
            df_po_2_1_3_13_15.columns = df_po_2_1_3_13_15.columns.droplevel()

            df_po_2_1_3_13_15['Всего по форме обучения'] = df_po_2_1_3_13_15.sum(axis=1)

            df_po_2_1_3 = pd.concat([df_po_2_1_3_6_9, df_po_2_1_3_13_15], axis=1)

            # Считаем по категориям финансирования
            df_po_2_1_3_cat_finance = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                                     columns=[
                                                         'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                                     values=['for_counting'],
                                                     aggfunc='sum')
            df_po_2_1_3_cat_finance.columns = df_po_2_1_3_cat_finance.columns.droplevel()

            df_po_2_1_3 = pd.concat([df_po_2_1_3, df_po_2_1_3_cat_finance], axis=1)

            # Считаем женщин

            df_po_2_1_3_women = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                               columns=['Пол_получателя'],
                                               values=['for_counting'],
                                               aggfunc='sum')
            df_po_2_1_3_women.columns = df_po_2_1_3_women.columns.droplevel()

            df_po_2_1_3 = pd.concat([df_po_2_1_3, df_po_2_1_3_women], axis=1)

            # Считаем тех кто прошел по индвидуальным учебным планам

            df_po_2_1_3_ind_up = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                                columns=[
                                                    'прошли_ускоренное_обучение_по_индивидуаль-ным_учебным_планам'],
                                                values=['for_counting'],
                                                aggfunc='sum')
            df_po_2_1_3_ind_up.columns = df_po_2_1_3_ind_up.columns.droplevel()

            df_po_2_1_3 = pd.concat([df_po_2_1_3, df_po_2_1_3_ind_up], axis=1)

            # Создаем проверку
            df_po_2_1_3['temp'] = df_po_2_1_3['Всего по источникам финансирования'] == df_po_2_1_3[
                'Всего по форме обучения']

            df_po_2_1_3['Совпадение сумм столбцов 6-9 и 13-15'] = df_po_2_1_3['temp'].apply(
                lambda x: 'СОВПАДАЕТ' if x is True else 'НЕ СОВПАДАЕТ!!!')
            df_po_2_1_3.drop(columns='temp', inplace=True)
            df_po_2_1_3.index.name = 'Код профессии при наличии'

            # Записываем датафрейм на лист
            wb['Раздел 2.1.3'][
                f'A1'] = 'Распределение слушателей,обученных по программам повышения квалификации рабочих,служащих'
            wb['Раздел 2.1.3'][f'A1'].font = font_name_table

            for r in dataframe_to_rows(df_po_2_1_3, index=True, header=True):
                if len(r) != 1:
                    wb['Раздел 2.1.3'].append(r)
            wb['Раздел 2.1.3'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.1.3'][f'A2'].font = font_name_table
            wb['Раздел 2.1.3']['A2'].alignment = Alignment(wrap_text=True)

            wb['Раздел 2.1.3'][f'A3'] = 'Код профессии при наличии'
            wb['Раздел 2.1.3'].column_dimensions['A'].width = 50
            wb['Раздел 2.1.3'].column_dimensions['C'].width = 15
            wb['Раздел 2.1.3'].column_dimensions['D'].width = 15
            wb['Раздел 2.1.3']['B2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['C2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['D2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['E2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['F2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['G2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['H2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['I2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['J2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['K2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['L2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['M2'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.1.3']['Q2'].alignment = Alignment(wrap_text=True)

        # Считаем лист 2.2
        df_po_2_2_base = df_po.copy()
        df_po_2_2_base['for_counting'] = 1

        # Создаем сводную таблицу
        df_po_2_2 = pd.pivot_table(df_po_2_2_base, index=['Категория_слушателя'],
                                   columns=['Программа_профессионального_обучения_направление_подготовки'],
                                   values=['for_counting'],
                                   aggfunc='sum')

        df_po_2_2.columns = df_po_2_2.columns.droplevel()

        df_po_2_2['Всего обучено по программам'] = df_po_2_2.sum(axis=1)

        # Записываем на лист
        # Записываем датафрейм на лист
        wb['Раздел 2.2'][f'A1'] = 'Обучение отдельных категорий слушателей'
        wb['Раздел 2.2'][f'A1'].font = font_name_table
        wb['Раздел 2.2'][
            f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
        wb['Раздел 2.2'][f'A2'].font = font_name_table
        wb['Раздел 2.2']['A2'].alignment = Alignment(wrap_text=True)

        for r in dataframe_to_rows(df_po_2_2, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.2'].append(r)
        wb['Раздел 2.2'][f'A3'] = 'Наименование показателей'
        wb['Раздел 2.2'].column_dimensions['A'].width = 80
        wb['Раздел 2.2'].column_dimensions['B'].width = 30
        wb['Раздел 2.2']['B3'].alignment = Alignment(wrap_text=True)
        wb['Раздел 2.2'].column_dimensions['C'].width = 30
        wb['Раздел 2.2']['C3'].alignment = Alignment(wrap_text=True)
        wb['Раздел 2.2'].column_dimensions['D'].width = 30
        wb['Раздел 2.2']['D3'].alignment = Alignment(wrap_text=True)
        wb['Раздел 2.2'].column_dimensions['D'].width = 30
        wb['Раздел 2.2']['E3'].alignment = Alignment(wrap_text=True)

        # Считаем раздел 2.3 Инвалиды
        df_po_2_3_base = df_po.copy()
        df_po_2_3_base['for_counting'] = 1
        df_po_2_3_base.fillna('Не заполнено', inplace=True)
        # Отбираем нездоровых
        df_po_2_3_base = df_po_2_3_base[
            (df_po_2_3_base['Сведения_об_ограничении_возможностей_здоровья'] != 'нет ОВЗ') & (
                    df_po_2_3_base['Сведения_об_ограничении_возможностей_здоровья'] != 'Не заполнено')]

        if df_po_2_3_base.shape[0] == 0:
            wb['Раздел 2.3'].column_dimensions['A'].width = 80
            wb['Раздел 2.3'][f'A1'] = 'Обучение лиц с ограниченными возможностями здоровья и инвалидов'
            wb['Раздел 2.3'][f'A1'].font = font_name_table
            wb['Раздел 2.3'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.3'][f'A2'].font = font_name_table
            wb['Раздел 2.3']['A2'].alignment = Alignment(wrap_text=True)
        else:
            # создаем сводную таблицу
            df_po_2_3 = pd.pivot_table(df_po_2_3_base, index=['Сведения_об_ограничении_возможностей_здоровья'],
                                       columns=['Программа_профессионального_обучения_направление_подготовки'],
                                       values=['for_counting'],
                                       aggfunc='sum')
            df_po_2_3.columns = df_po_2_3.columns.droplevel()
            df_po_2_3['Всего обучено'] = df_po_2_3.sum(axis=1)
            # Записываем в лист
            wb['Раздел 2.3'][f'A1'] = 'Обучение лиц с ограниченными возможностями здоровья и инвалидов'
            wb['Раздел 2.3'][f'A1'].font = font_name_table
            wb['Раздел 2.3'][
                f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
            wb['Раздел 2.3'][f'A2'].font = font_name_table
            wb['Раздел 2.3']['A2'].alignment = Alignment(wrap_text=True)
            for r in dataframe_to_rows(df_po_2_3, index=True, header=True):
                if len(r) != 1:
                    wb['Раздел 2.3'].append(r)

            wb['Раздел 2.3'][f'A3'] = 'Наименование показателей'
            wb['Раздел 2.3'].column_dimensions['A'].width = 80
            wb['Раздел 2.3'].column_dimensions['B'].width = 30
            wb['Раздел 2.3']['B3'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.3'].column_dimensions['C'].width = 30
            wb['Раздел 2.3']['C3'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.3'].column_dimensions['D'].width = 30
            wb['Раздел 2.3']['D3'].alignment = Alignment(wrap_text=True)
            wb['Раздел 2.3'].column_dimensions['D'].width = 30
            wb['Раздел 2.3']['E3'].alignment = Alignment(wrap_text=True)

        # Считаем лист 2.4
        df_po_2_4 = df_po.copy()
        df_po_2_4['for_counting'] = 1

        # Считаем общее количество
        df_po_2_4_all = pd.pivot_table(df_po_2_4,
                                       columns=['Программа_профессионального_обучения_направление_подготовки'],
                                       values=['for_counting'],
                                       aggfunc='sum')
        df_po_2_4_all.index = ['Всего обучено']

        # Записываем в лист
        wb['Раздел 2.4'][f'A1'] = 'Распределение слушателей,обученных по программам ПО,по уровню образования'
        wb['Раздел 2.4'][f'A1'].font = font_name_table
        wb['Раздел 2.4'][
            f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
        wb['Раздел 2.4'][f'A2'].font = font_name_table
        wb['Раздел 2.4']['A2'].alignment = Alignment(wrap_text=True)

        # Считаем по основному общему образованию
        df_po_2_4_02 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'основное общее (9 классов)']

        # считаем сумму
        df_po_2_4_02_all = pd.pivot_table(df_po_2_4_02,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_02_all.index = ['Основное общее']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_02_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_02_current_year = df_po_2_4_02.copy()

        df_po_2_4_02_current_year['Образование в текущем году'] = df_po_2_4_02_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_02_current_year = df_po_2_4_02_current_year[
            df_po_2_4_02_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_02_current_year_finsih = pd.pivot_table(df_po_2_4_02_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_02_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_02_current_year_finsih])

        # Считаем по среднему общему образованию
        df_po_2_4_04 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'среднее общее (11 классов)']

        # считаем сумму
        df_po_2_4_04_all = pd.pivot_table(df_po_2_4_04,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_04_all.index = ['Среднее общее']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_04_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_04_current_year = df_po_2_4_04.copy()

        df_po_2_4_04_current_year['Образование в текущем году'] = df_po_2_4_04_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_04_current_year = df_po_2_4_04_current_year[
            df_po_2_4_04_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_04_current_year_finsih = pd.pivot_table(df_po_2_4_04_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_04_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_04_current_year_finsih])

        # Считаем по среднему общему образованию
        df_po_2_4_06 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'СПО/НПО']

        # считаем сумму
        df_po_2_4_06_all = pd.pivot_table(df_po_2_4_06,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_06_all.index = ['Среднее профессиональное']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_06_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_06_current_year = df_po_2_4_06.copy()

        df_po_2_4_06_current_year['Образование в текущем году'] = df_po_2_4_06_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_06_current_year = df_po_2_4_06_current_year[
            df_po_2_4_06_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_06_current_year_finsih = pd.pivot_table(df_po_2_4_06_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_06_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_06_current_year_finsih])

        # Считаем по неполному высшему общему образованию
        df_po_2_4_10 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'неполное высшее образование']

        # считаем сумму
        df_po_2_4_10_all = pd.pivot_table(df_po_2_4_10,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_10_all.index = ['Неполное высшее образование']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_10_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_10_current_year = df_po_2_4_10.copy()

        df_po_2_4_10_current_year['Образование в текущем году'] = df_po_2_4_10_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_10_current_year = df_po_2_4_10_current_year[
            df_po_2_4_10_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_10_current_year_finsih = pd.pivot_table(df_po_2_4_10_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_10_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_10_current_year_finsih])

        # Считаем по  высшему Бакалавры образованию
        df_po_2_4_11 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'ВО (бакалавр)']

        # считаем сумму
        df_po_2_4_11_all = pd.pivot_table(df_po_2_4_11,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_11_all.index = ['ВО (бакалавр)']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_11_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_11_current_year = df_po_2_4_11.copy()

        df_po_2_4_11_current_year['Образование в текущем году'] = df_po_2_4_11_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_11_current_year = df_po_2_4_11_current_year[
            df_po_2_4_11_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_11_current_year_finsih = pd.pivot_table(df_po_2_4_11_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_11_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_11_current_year_finsih])

        # Считаем по  высшему специалист образованию
        df_po_2_4_13 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'ВО (специалист)']

        # считаем сумму
        df_po_2_4_13_all = pd.pivot_table(df_po_2_4_13,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_13_all.index = ['ВО (специалист)']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_13_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_13_current_year = df_po_2_4_13.copy()

        df_po_2_4_13_current_year['Образование в текущем году'] = df_po_2_4_13_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_13_current_year = df_po_2_4_13_current_year[
            df_po_2_4_13_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_13_current_year_finsih = pd.pivot_table(df_po_2_4_13_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_13_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_13_current_year_finsih])

        # Считаем по  высшему магистр образованию
        df_po_2_4_15 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'ВО (магистр)']

        # считаем сумму
        df_po_2_4_15_all = pd.pivot_table(df_po_2_4_15,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_15_all.index = ['ВО (магистр)']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_15_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_15_current_year = df_po_2_4_15.copy()

        df_po_2_4_15_current_year['Образование в текущем году'] = df_po_2_4_15_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_15_current_year = df_po_2_4_15_current_year[
            df_po_2_4_15_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_15_current_year_finsih = pd.pivot_table(df_po_2_4_15_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_15_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_15_current_year_finsih])

        # Считаем по  высшему магистр образованию
        df_po_2_4_17 = df_po_2_4[
            df_po_2_4['Уровень_образования'] == 'без образования (не имеют основного общего образования)']

        # считаем сумму
        df_po_2_4_17_all = pd.pivot_table(df_po_2_4_17,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_17_all.index = ['без образования (не имеют основного общего образования)']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_17_all], axis=0)

        # Считаем получивших образование в отчетном году
        df_po_2_4_17_current_year = df_po_2_4_17.copy()

        df_po_2_4_17_current_year['Образование в текущем году'] = df_po_2_4_17_current_year[
            'Год_получения_образования'].apply(
            lambda x: 'Да' if x == date.today().year else 'Нет')

        df_po_2_4_17_current_year = df_po_2_4_17_current_year[
            df_po_2_4_17_current_year['Образование в текущем году'] == 'Да']

        df_po_2_4_17_current_year_finsih = pd.pivot_table(df_po_2_4_17_current_year, columns=[
            'Программа_профессионального_обучения_направление_подготовки'],
                                                          values=['for_counting'],
                                                          aggfunc='sum')
        df_po_2_4_17_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_17_current_year_finsih])

        # Считаем инвалидов из людей без образования
        df_po_2_4_20 = df_po_2_4_17[df_po_2_4_17['Сведения_об_ограничении_возможностей_здоровья'] != 'нет ОВЗ']

        # считаем сумму
        df_po_2_4_20_all = pd.pivot_table(df_po_2_4_20,
                                          columns=['Программа_профессионального_обучения_направление_подготовки'],
                                          values=['for_counting'],
                                          aggfunc='sum')
        df_po_2_4_20_all.index = [
            'лица с ограниченными возможностиями здоровья(инвалиды,дети-сироты,лицо с ОВЗ) не имеющие ОО']

        # Добавляем к основному датафрейму
        df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_20_all], axis=0)

        for r in dataframe_to_rows(df_po_2_4_all, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.4'].append(r)
        # Устанавливаем ширину колонок
        wb['Раздел 2.4'].column_dimensions['A'].width = 80
        wb['Раздел 2.4'].column_dimensions['B'].width = 30
        wb['Раздел 2.4']['B3'].alignment = Alignment(wrap_text=True)
        wb['Раздел 2.4'].column_dimensions['C'].width = 30
        wb['Раздел 2.4']['C3'].alignment = Alignment(wrap_text=True)
        wb['Раздел 2.4'].column_dimensions['D'].width = 30
        wb['Раздел 2.4']['D3'].alignment = Alignment(wrap_text=True)

        # Считаем лист 2.5
        # Создаем дополнительную числовую колонку где каждое значение это 1, для удобства агрегирования
        df_po_2_5 = df_po.copy()
        # Добавляем колонку с 1
        df_po_2_5['for_counting'] = 1

        # Считаем строку 01 Всего
        df_2_5_all = pd.pivot_table(df_po_2_5,
                                    columns=['Программа_профессионального_обучения_направление_подготовки',
                                             'Источник_финансирования_обучения'],
                                    values=['for_counting'],
                                    aggfunc='sum')
        df_2_5_all.index = ['Всего']
        df_2_5_all.index.name = 'Код государства по ОКСМ'

        # Считаем данные по странам
        svod_df_po_2_5 = pd.pivot_table(df_po_2_5,
                                        index=['Гражданство_получателя_код_страны_по_ОКСМ'],
                                        columns=['Программа_профессионального_обучения_направление_подготовки',
                                                 'Источник_финансирования_обучения'],
                                        values=['for_counting'],
                                        aggfunc='sum')

        # Удаляем лишний мультииндекс
        svod_df_po_2_5.columns = svod_df_po_2_5.columns.droplevel()
        # заполняем нулями для корректного суммирования
        svod_df_po_2_5.fillna(0.0, inplace=True)

        # Соединяем датафреймы
        df_po_2_5_out = pd.concat([df_2_5_all, svod_df_po_2_5])
        # заменяем нули на нан, чтобы в итоговой таблице нули не отвлекали
        df_po_2_5_out.replace(0.0, np.NaN, inplace=True)

        wb['Раздел 2.5'][f'A1'] = 'Распределение слушателей, обученных по программам ПО, по гражданству'
        wb['Раздел 2.5'][f'A1'].font = font_name_table
        wb['Раздел 2.5'][
            f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!'
        wb['Раздел 2.5'][f'A2'].font = font_name_table

        for r in dataframe_to_rows(df_po_2_5_out, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.5'].append(r)
        wb['Раздел 2.5'].column_dimensions['A'].width = 50
        wb['Раздел 2.5'].column_dimensions['B'].width = 50
        wb['Раздел 2.5'].column_dimensions['F'].width = 50

        # Создаем раздел 2.6
        df_po_2_6 = df_po.copy()
        df_po_2_6['for_counting'] = 1

        # Считаем в общем сколько обучено
        df_po_2_6_all = pd.pivot_table(df_po_2_6,
                                       columns=['Программа_профессионального_обучения_направление_подготовки',
                                                'Пол_получателя'],
                                       values=['for_counting'],
                                       aggfunc='sum')
        df_po_2_6_all.index = ['Всего обучено']

        # Считаем распределение по возрастам
        df_po_2_6_by_age = pd.pivot_table(df_po_2_6, index=['Возрастная_категория_1ПО'],
                                          columns=['Программа_профессионального_обучения_направление_подготовки',
                                                   'Пол_получателя'],
                                          values=['for_counting'],
                                          aggfunc='sum')

        # Удаляем лишний мультииндекс
        df_po_2_6_by_age.columns = df_po_2_6_by_age.columns.droplevel()

        # Соединяем 2 датафрейма
        df_po_2_6_out = pd.concat([df_po_2_6_all, df_po_2_6_by_age])

        wb['Раздел 2.6'][f'A1'] = 'Распределение слушателей, обученных по программам ПО, по возрасту и полу'
        wb['Раздел 2.6'][f'A1'].font = font_name_table
        wb['Раздел 2.6'][
            f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!'
        wb['Раздел 2.6'][f'A2'].font = font_name_table
        wb['Раздел 2.6'][f'A3'] = 'Чтобы вычислить показатель ВСЕГО ОБУЧЕНО сложите значения в колонке Муж+Жен'
        wb['Раздел 2.6'][f'A3'].font = font_name_table

        for r in dataframe_to_rows(df_po_2_6_out, index=True, header=True):
            if len(r) != 1:
                wb['Раздел 2.6'].append(r)
        wb['Раздел 2.6'].column_dimensions['A'].width = 50
        wb['Раздел 2.6'].column_dimensions['B'].width = 50
        wb['Раздел 2.6']['B4'].alignment = Alignment(wrap_text=True)
        wb['Раздел 2.6'].column_dimensions['D'].width = 50
        wb['Раздел 2.6']['D4'].alignment = Alignment(wrap_text=True)
        wb['Раздел 2.6'].column_dimensions['F'].width = 50
        wb['Раздел 2.6']['F4'].alignment = Alignment(wrap_text=True)

        # Получаем текущее время для того чтобы использовать в названии
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder_report}/Часть отчета 1-ПО.xlsx {current_time}.xlsx')
    except NameError:
        messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    else:
        messagebox.showinfo('ЦОПП Бурятия', 'Создание части отчета 1-ПО\nЗавершено!')


def create_report_svod():
    """
    Функция для создания отчета по сводным показателям ЦОПП
    :return:
    """
    # Загружаем датафреймы
    try:
        dpo_df = pd.read_excel(name_file_data_report, sheet_name='ДПО',
                               dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str})
        po_df = pd.read_excel(name_file_data_report, sheet_name='ПО',
                              dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str})
        """
        Проверяем заполнена ли колонка Возрастная категория.Если заполнена, то значит таблица прошла через процедуру create_general_table
        Но нужно обработать случай когда нужно сделать отчет по одной таблице
        """
        if 'Текущий_возраст' not in dpo_df.columns or 'Текущий_возраст' not in po_df.columns:
            dpo_df['Текущий_возраст'] = dpo_df['Дата_рождения_получателя'].apply(calculate_age)
            dpo_df['Возрастная_категория_1ПК'] = pd.cut(dpo_df['Текущий_возраст'],
                                                        [0, 24, 29, 34, 39, 44, 49, 54, 59, 64, 101, 10000],
                                                        labels=['моложе 25 лет', '25-29', '30-34', '35-39',
                                                                '40-44', '45-49', '50-54', '55-59', '60-64',
                                                                '65 и более', 'Возраст  больше 101'])
            #
            po_df['Текущий_возраст'] = po_df['Дата_рождения_получателя'].apply(calculate_age)
            po_df['Возрастная_категория_1ПО'] = pd.cut(po_df['Текущий_возраст'],
                                                       [0, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27,
                                                        28,
                                                        29, 34, 39, 44, 49, 54, 59, 64, 101],
                                                       labels=['В возрасте моложе 14 лет', '14 лет', '15 лет', '16 лет',
                                                               '17 лет', '18 лет', '19 лет', '20 лет', '21 год',
                                                               '22 года',
                                                               '23 года', '24 года', '25 лет',
                                                               '26 лет', '27 лет', '28 лет', '29 лет', '30-34 лет',
                                                               '35-39 лет', '40-44 лет', '45-49 лет', '50-54 лет',
                                                               '55-59 лет',
                                                               '60-64 лет',
                                                               '65 лет и старше'])
            # Приводим Возрастную категорию к текстовому типу, иначе при fillna возникает ошибка, он не может заполнить категориальные данные
            dpo_df['Возрастная_категория_1ПК'] = dpo_df['Возрастная_категория_1ПК'].astype(str)
            po_df['Возрастная_категория_1ПО'] = po_df['Возрастная_категория_1ПО'].astype(str)

        # Заполняем пустые поля для удобства группировки
        dpo_df = dpo_df.fillna('Не заполнено!!!')
        po_df = po_df.fillna('Не заполнено!!!')

        # Получение общего количества прошедших обучение,количества прошедших по ДПО,по ПО
        total_students, total_students_dpo, total_students_po = counting_total_student(dpo_df, po_df)

        # Количество обучившихся ДПО и ПО

        # Получение количества обучившихся по видам
        df_counting_type_and_name_trainning = counting_type_of_training(dpo_df, po_df)

        # Создаем новый excel файл
        wb = openpyxl.Workbook()

        # Получаем активный лист
        sheet = wb.active
        sheet.title = 'Сводные данные'

        # Начинаем заполнение листа
        # Заполняем количество обучившихся, общее и по типам
        sheet['A1'] = 'Наименование показателя'
        sheet['A2'] = 'Количество прошедших обучение ДПО'
        sheet['A3'] = 'Количество прошедших обучение ПО'
        sheet['A4'] = 'Общее количество прошедших обучение в ЦОПП'

        sheet['B1'] = 'Количество обучившихся'
        sheet['B2'] = total_students_dpo
        sheet['B3'] = total_students_po
        sheet['B4'] = total_students

        # Добавляем круговую диаграмму
        pie_main = PieChart()
        labels = Reference(sheet, min_col=1, min_row=2, max_row=3)
        data = Reference(sheet, min_col=2, min_row=2, max_row=3)

        # Для отображения данных на диаграмме
        series = Series(data, title='Series 1')
        pie_main.append(series)

        s1 = pie_main.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showVal = True

        pie_main.add_data(data, titles_from_data=True)
        pie_main.set_categories(labels)
        pie_main.title = 'Распределение обучившихся'
        sheet.add_chart(pie_main, 'F1')
        # # Добавляем таблицу с по направлениям

        sheet['A7'] = 'Вид обучения'
        sheet['B7'] = 'Название программы'
        sheet['C7'] = 'Количество обучившихся'

        for row in df_counting_type_and_name_trainning.values.tolist():
            sheet.append(row)
        # Получаем последние активные ячейки чтобы записывалось по порядку и не налазило друг на друга
        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row

        sheet[f'A{max_row + 2}'] = 'Общее распределение прошедших обучение по полу'
        total_sex = counting_total_sex(dpo_df, po_df)
        # Добавляем в файл таблицу с распределением по полам
        for row in total_sex.values.tolist():
            sheet.append(row)

        # Получаем последние активные ячейки чтобы записывалось по порядку и не налазило друг на друга
        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row

        # Добавляем таблицу с разбиением по возрастам 1-ПК
        sheet[f'A{max_row + 2}'] = 'Распределение обучившихся по возрастным категориям 1-ПК'
        age_distribution_dpo = counting_age_distribution_dpo(dpo_df)
        for row in age_distribution_dpo.values.tolist():
            sheet.append(row)

        # Добавляем круговую диаграмму
        pie_age_dpo = PieChart()
        # Для того чтобы не зависело от количества строк в предыдущих таблицах
        labels = Reference(sheet, min_col=1, min_row=max_row + 3, max_row=max_row + 2 + len(age_distribution_dpo))
        data = Reference(sheet, min_col=2, min_row=max_row + 3, max_row=max_row + 2 + len(age_distribution_dpo))
        # Для отображения данных на диаграмме
        series = Series(data, title='Series 1')
        pie_age_dpo.append(series)

        s1 = pie_age_dpo.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showVal = True

        pie_age_dpo.add_data(data, titles_from_data=True)
        pie_age_dpo.set_categories(labels)
        pie_age_dpo.title = 'Распределение обучившихся по возрастным категориям 1-ПК'

        sheet.add_chart(pie_age_dpo, f'F{max_row + 2}')

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row

        # Добавляем таблицу с разбиением по возрастам 1-ПО
        sheet[f'A{max_row + 4}'] = 'Распределение обучившихся по возрастным категориям 1-ПО'
        age_distribution_po = counting_age_distribution_po(po_df)
        for row in age_distribution_po.values.tolist():
            sheet.append(row)

        # Добавляем круговую диаграмму
        pie_age_po = PieChart()
        # Для того чтобы не зависело от количества строк в предыдущих таблицах
        labels = Reference(sheet, min_col=1, min_row=max_row + 5, max_row=max_row + 4 + len(age_distribution_po))
        data = Reference(sheet, min_col=2, min_row=max_row + 5, max_row=max_row + 4 + len(age_distribution_po))
        # Для отображения данных на диаграмме
        series = Series(data, title='Series 1')
        pie_age_po.append(series)

        s1 = pie_age_po.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showVal = True

        pie_age_po.add_data(data, titles_from_data=True)
        pie_age_po.set_categories(labels)
        pie_age_po.title = 'Распределение обучившихся по возрастным категориям 1-ПО'

        sheet.add_chart(pie_age_po, f'F{max_row + 5}')

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row

        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 30

        # Сохраняем файл
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        wb.save(f'{path_to_end_folder_report}/Сводный отчет {current_time}.xlsx')

    except NameError:
        messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    except ValueError:
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте названия листов! Должно быть ДПО и ПО')
    # except KeyError:
    #     messagebox.showerror('ЦОПП Бурятия', 'Названия колонок не совпадают')
    # except:
    #     messagebox.showerror('ЦОПП Бурятия',
    #                          'Возникла ошибка')
    else:
        messagebox.showinfo('ЦОПП Бурятия', 'Сводный отчет успешно создан!')


def create_general_table():
    """
    Функция для создания общей таблицы с данными всех групп из множества отдельных таблицы на каждую группу
    :return:
    """
    pattern = re.compile(
        '^[А-ЯЁ]+_.+_(?:январь|февраль|март|апрель|май|июнь|июль|август|сентябрь|октябрь|ноябрь|декабрь)\.xlsx$')
    try:
        # Получаем базовые датафреймы
        df_dpo = pd.read_excel(name_file_template_table, sheet_name='ДПО',
                               dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                      'Серия_паспорта_в_формате_1111': str,
                                      'Номер_паспорта_в_формате_111111': str})
        df_po = pd.read_excel(name_file_template_table, sheet_name='ПО',
                              dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                     'Серия_паспорта_совершеннолетнего_или_родителя_законного_представителя_в_формате_1111': str,
                                     'Номер_паспорта_в_формате_111111': str
                                     })
        # Очищаем базовые датафреймы на случай  если там есть какие то строки. Необходимо чтобы шаблон был полностью пуст
        df_dpo = df_dpo.iloc[0:0]
        df_po = df_po.iloc[0:0]

        # Добавляем 2 колонки с возрастом и категорией для каждого базового датафрейма.Чтобы конкатенация прошла успешно
        df_dpo['Текущий_возраст'] = np.nan
        df_dpo['Возрастная_категория_1ПК'] = np.nan
        df_dpo['Дата_начала_курса'] = np.nan
        df_dpo['Дата_окончания_курса'] = np.nan
        df_dpo['Месяц_начала_курса'] = np.nan
        df_dpo['Месяц_окончания_курса'] = np.nan

        df_po['Текущий_возраст'] = np.nan
        df_po['Возрастная_категория_1ПО'] = np.nan
        df_po['Дата_начала_курса'] = np.nan
        df_po['Дата_окончания_курса'] = np.nan
        df_po['Месяц_начала_курса'] = np.nan
        df_po['Месяц_окончания_курса'] = np.nan

        # Получаем множество из навзваний колонок в шаблоне для каждого листа
        dpo_template_cols = set(df_dpo.columns)
        po_template_cols = set(df_po.columns)

        # Перебираем файлы собирая данные в промежуточные датафреймы и добавляя их в базовые
        for dirpath, dirnames, filenames in os.walk(path_to_files_groups):
            for filename in filenames:

                if re.search(pattern, filename):
                    print("Файл:", os.path.join(dirpath, filename))
                    # Создаем промежуточный датафрейм с данными с листа ДПО
                    temp_dpo = pd.read_excel(os.path.join(dirpath, filename), sheet_name='ДПО',
                                             dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                                    'Серия_паспорта_в_формате_1111': str,
                                                    'Номер_паспорта_в_формате_111111': str,
                                                    'Серия_документа_о_ВО_СПО': str,
                                                    'Номер_документа_о_ВО_СПО': str})
                    # Создаем промежуточный датафрейм с данными с листа ДПО
                    temp_po = pd.read_excel(os.path.join(dirpath, filename), sheet_name='ПО',
                                            dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                                   'Серия_паспорта_совершеннолетнего_или_родителя_законного_представителя_в_формате_1111': str,
                                                   'Номер_паспорта_в_формате_111111': str})

                    temp_dpo['Дата_рождения_получателя'] = pd.to_datetime(temp_dpo['Дата_рождения_получателя'],
                                                                          dayfirst=True, errors='coerce')
                    temp_dpo['Дата_выдачи_документа'] = pd.to_datetime(temp_dpo['Дата_выдачи_документа'],
                                                                       dayfirst=True, errors='coerce')
                    temp_dpo['Дата_выдачи_паспорта'] = pd.to_datetime(temp_dpo['Дата_выдачи_паспорта'],
                                                                      dayfirst=True, errors='coerce')

                    temp_po['Дата_рождения_получателя'] = pd.to_datetime(temp_po['Дата_рождения_получателя'],
                                                                         dayfirst=True, errors='coerce')
                    temp_po['Дата_выдачи_документа'] = pd.to_datetime(temp_po['Дата_выдачи_документа'],
                                                                      dayfirst=True, errors='coerce')
                    temp_po['Дата_выдачи_паспорта'] = pd.to_datetime(temp_po['Дата_выдачи_паспорта'],
                                                                     dayfirst=True, errors='coerce')

                    # если на листе не ноль строк то, обрабатываем
                    if temp_dpo.shape[0] > 0:
                        # Добавляем 2 колонки с характеристиками возраста
                        temp_dpo['Текущий_возраст'] = temp_dpo['Дата_рождения_получателя'].apply(calculate_age)
                        temp_dpo['Возрастная_категория_1ПК'] = pd.cut(temp_dpo['Текущий_возраст'],
                                                                      [0, 24, 29, 34, 39, 44, 49, 54, 59, 64, 101,
                                                                       10000],
                                                                      labels=['моложе 25 лет', '25-29', '30-34',
                                                                              '35-39',
                                                                              '40-44', '45-49', '50-54', '55-59',
                                                                              '60-64',
                                                                              '65 и более',
                                                                              'Возраст  больше 101'])
                        # Добавляем 4 колонки с характеристиками дат курсов
                        temp_dpo['Дата_начала_курса'] = temp_dpo[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(
                            extract_date_begin_course)
                        temp_dpo['Дата_окончания_курса'] = temp_dpo[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_end_course)
                        temp_dpo['Месяц_начала_курса'] = temp_dpo[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(
                            extract_month_begin_course)
                        temp_dpo['Месяц_окончания_курса'] = temp_dpo[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_month_end_course)
                    else:
                        # Создаем пустые колонки,чтобы не сбивалась структура таблицы
                        temp_dpo['Текущий_возраст'] = np.nan
                        temp_dpo['Возрастная_категория_1ПК'] = np.nan
                        temp_dpo['Дата_начала_курса'] = np.nan
                        temp_dpo['Дата_окончания_курса'] = np.nan
                        temp_dpo['Месяц_начала_курса'] = np.nan
                        temp_dpo['Месяц_окончания_курса'] = np.nan

                    if temp_po.shape[0] > 0:
                        # Обрабатываем датафрейм с ПО
                        temp_po['Текущий_возраст'] = temp_po['Дата_рождения_получателя'].apply(calculate_age)
                        temp_po['Возрастная_категория_1ПО'] = pd.cut(temp_po['Текущий_возраст'],
                                                                     [0, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24,
                                                                      25,
                                                                      26, 27, 28,
                                                                      29, 34, 39, 44, 49, 54, 59, 64, 101],
                                                                     labels=['В возрасте моложе 14 лет', '14 лет',
                                                                             '15 лет',
                                                                             '16 лет',
                                                                             '17 лет', '18 лет', '19 лет', '20 лет',
                                                                             '21 год', '22 года',
                                                                             '23 года', '24 года', '25 лет',
                                                                             '26 лет', '27 лет', '28 лет', '29 лет',
                                                                             '30-34 лет',
                                                                             '35-39 лет', '40-44 лет', '45-49 лет',
                                                                             '50-54 лет',
                                                                             '55-59 лет',
                                                                             '60-64 лет',
                                                                             '65 лет и старше'])
                        # Добавляем 4 колонки с характеристиками дат курсов
                        temp_po['Дата_начала_курса'] = temp_po[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(
                            extract_date_begin_course)
                        temp_po['Дата_окончания_курса'] = temp_po[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_end_course)
                        temp_po['Месяц_начала_курса'] = temp_po[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(
                            extract_month_begin_course)
                        temp_po['Месяц_окончания_курса'] = temp_po[
                            'Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_month_end_course)
                    else:
                        temp_po['Текущий_возраст'] = np.nan
                        temp_po['Возрастная_категория_1ПО'] = np.nan
                        temp_po['Дата_начала_курса'] = np.nan
                        temp_po['Дата_окончания_курса'] = np.nan
                        temp_po['Месяц_начала_курса'] = np.nan
                        temp_po['Месяц_окончания_курса'] = np.nan

                    # Конвертируем  столбцы с датами в краткий формат
                    temp_dpo['Дата_выдачи_документа'] = temp_dpo['Дата_выдачи_документа'].apply(convert_date)
                    temp_dpo['Дата_рождения_получателя'] = temp_dpo['Дата_рождения_получателя'].apply(convert_date)
                    temp_dpo['Дата_выдачи_паспорта'] = temp_dpo['Дата_выдачи_паспорта'].apply(convert_date)

                    temp_po['Дата_выдачи_документа'] = temp_po['Дата_выдачи_документа'].apply(convert_date)
                    temp_po['Дата_рождения_получателя'] = temp_po['Дата_рождения_получателя'].apply(convert_date)
                    temp_po['Дата_выдачи_паспорта'] = temp_po['Дата_выдачи_паспорта'].apply(convert_date)

                    # Проверяем состав колонок
                    temp_dpo_columns = set(temp_dpo.columns)
                    temp_po_columns = set(temp_po.columns)
                    # Если есть разница то выдаем сообщение предупреждение
                    diff_cols_dpo = dpo_template_cols - temp_dpo_columns
                    diff_cols_po = po_template_cols - temp_po_columns

                    if len(diff_cols_dpo) > 0:
                        messagebox.showerror('ЦОПП Бурятия',
                                             f'В файле {filename} на листе ДПО отличается состав колонок по сравнению с шаблоном {name_file_template_table}\n Проверьте наличие указанных колонок в обоих файлах: {diff_cols_dpo}\nдля корректной обработки')

                    if len(diff_cols_po) > 0:
                        messagebox.showerror('ЦОПП Бурятия',
                                             f'В файле {filename} на листе ПО отличается состав колонок по сравнению с шаблоном {name_file_template_table}\n Проверьте наличие указанных колонок в обоих файлах: {diff_cols_po}\nдля корректной обработки')

                    # Добавляем промежуточные датафреймы в исходные
                    #

                    df_dpo = pd.concat([df_dpo, temp_dpo], ignore_index=True)
                    df_po = pd.concat([df_po, temp_po], ignore_index=True)

        # Код сохранения датафрейма в разные листы и сохранением форматирования  взят отсюда https://azzrael.ru/python-pandas-openpyxl-excel
        wb = openpyxl.load_workbook(name_file_template_table)

        # Записываем лист ДПО

        for ir in range(0, len(df_dpo)):
            for ic in range(0, len(df_dpo.iloc[ir])):
                wb['ДПО'].cell(2 + ir, 1 + ic).value = df_dpo.iloc[ir][ic]

        wb['ДПО']['BO1'] = 'Текущий_возраст'
        wb['ДПО']['BP1'] = 'Возрастная_категория_1ПК'
        wb['ДПО']['BQ1'] = 'Дата_начала_курса'
        wb['ДПО']['BR1'] = 'Дата_окончания_курса'
        wb['ДПО']['BS1'] = 'Месяц_начала_курса'
        wb['ДПО']['BT1'] = 'Месяц_окончания_курса'

        # Записываем лист ПО
        for ir in range(0, len(df_po)):
            for ic in range(0, len(df_po.iloc[ir])):
                wb['ПО'].cell(2 + ir, 1 + ic).value = df_po.iloc[ir][ic]
        wb['ПО']['BJ1'] = 'Текущий_возраст'
        wb['ПО']['BK1'] = 'Возрастная_категория_1ПО'
        wb['ПО']['BL1'] = 'Дата_начала_курса'
        wb['ПО']['BM1'] = 'Дата_окончания_курса'
        wb['ПО']['BN1'] = 'Месяц_начала_курса'
        wb['ПО']['BO1'] = 'Месяц_окончания_курса'

        # Получаем текущее время для того чтобы использовать в названии

        t = time.localtime()
        current_time = time.strftime('%d_%m_%y', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder_doc}/Общая таблица слушателей ЦОПП от {current_time}.xlsx')


    except NameError as e:
        messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    # except:
    #     messagebox.showerror('ЦОПП Бурятия',
    #                          'Возникла ошибка,проверьте шаблон таблицы\nДобавляемы файлы должны иметь одинаковую структуру с шаблоном таблицы')
    else:
        messagebox.showinfo('ЦОПП Бурятия', 'Общая таблица успешно создана!')


# Функции для создания сводной таблицы
def counting_total_student(dpo_df, po_df):
    """
    Функция для подсчета общего количества студентов обучающихся в цопп, и количества обучающихся по ДПО И ПО
    :param dpo_df: датафрейм ДПО
    :param po_df: датафрейм ПО
    :return: кортеж вида: общее количество обучающихся,количество обучающихся ДПО,количество обучающихся ПО
    """
    # количество по типам
    total_dpo = dpo_df.shape[0]
    total_po = po_df.shape[0]
    # общее количество
    total = total_dpo + total_po

    return total, total_dpo, total_po


def counting_type_of_training(dpo, po):
    """
    Функция для создания сводной таблицы по категориям направление подготовки, название программы,количество обучающихся
    :param dpo: датафрейм ДПО
    :param po: датафрейм ПО
    :return: датафрейм сводной таблицы
    """
    # Создаем сводные таблицы проверяее перед этим не пустые ли таблицы
    if dpo.shape[0] > 0:
        dpo_svod_category_and_name = pd.pivot_table(dpo, index=[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка',
            'Наименование_дополнительной_профессиональной_программы'],
                                                    values=['ФИО_именительный'],
                                                    aggfunc='count')
        dpo_svod_category_and_name = dpo_svod_category_and_name.reset_index()
    else:
        dpo_svod_category_and_name = pd.DataFrame(
            columns=['Направление подготовки', 'Название программы', 'Количество обученных'])

    if po.shape[0] > 0:
        po_svod_category_and_name = pd.pivot_table(po,
                                                   index=['Программа_профессионального_обучения_направление_подготовки',
                                                          'Наименование_программы_профессионального_обучения'],
                                                   values=['ФИО_именительный'],
                                                   aggfunc='count')
        po_svod_category_and_name = po_svod_category_and_name.reset_index()

    else:
        po_svod_category_and_name = pd.DataFrame(
            columns=['Направление подготовки', 'Название программы', 'Количество обученных'])
    # Изменяем названия колонок, чтобы без проблем соединить 2 датафрейма
    dpo_svod_category_and_name.columns = ['Направление подготовки', 'Название программы', 'Количество обученных']
    po_svod_category_and_name.columns = ['Направление подготовки', 'Название программы', 'Количество обученных']
    # Создаем единую сводную таблицу
    general_svod_category_and_name = pd.concat([dpo_svod_category_and_name, po_svod_category_and_name],
                                               ignore_index=True)
    return general_svod_category_and_name


def counting_total_sex(dpo, po):
    """
    Функция для подсчета количества мужчин и женщин
    :param dpo: датафрейм ДПО
    :param po: датафрейм ПО
    :return: датафрейм сводной таблицы
    """
    # Создаем сводные таблицы Проверяем на пустой лист ДПО или ПО
    if dpo.shape[0] > 0:

        dpo_total_sex = pd.pivot_table(dpo, index=['Пол_получателя'],
                                       values=['ФИО_именительный'],
                                       aggfunc='count')
        dpo_total_sex = dpo_total_sex.reset_index()
    else:
        dpo_total_sex = pd.DataFrame(columns=['Пол', 'Количество'])

    if po.shape[0] > 0:
        po_total_sex = pd.pivot_table(po, index=['Пол_получателя'],
                                      values=['ФИО_именительный'],
                                      aggfunc='count')
        po_total_sex = po_total_sex.reset_index()
    else:
        po_total_sex = pd.DataFrame(columns=['Пол', 'Количество'])
    # Переименовываем колонки
    dpo_total_sex.columns = ['Пол', 'Количество']
    po_total_sex.columns = ['Пол', 'Количество']

    # Соединяем в единую таблицу
    general_total_sex = pd.concat([dpo_total_sex, po_total_sex], ignore_index=True)
    # Группируем по полю Пол чтобы суммировать значения
    sum_general_total_sex = general_total_sex.groupby(['Пол']).sum().reset_index()
    return sum_general_total_sex


def counting_age_distribution_dpo(dpo):
    """
    Функция для подсчета количества обучающихся по возрастным категориям
    :param dpo: датафрейм ДПО
    :return: датафрейм сводной таблицы
    """
    # Создаем сводные таблицы
    if dpo.shape[0] > 0:
        dpo_age_distribution = pd.pivot_table(dpo, index=['Возрастная_категория_1ПК'],
                                              values=['ФИО_именительный'],
                                              aggfunc='count')
        dpo_age_distribution = dpo_age_distribution.reset_index()
    else:
        dpo_age_distribution = pd.DataFrame(columns=['Возрастная_категория_1ПК', 'Количество'])

    return dpo_age_distribution


def counting_age_distribution_po(po):
    """
    Функция для подсчета количества обучающихся по возрастным категориям
    :param dpo: датафрейм ПО
    :return: датафрейм сводной таблицы
    """
    if po.shape[0] > 0:
        po_age_distribution = pd.pivot_table(po, index=['Возрастная_категория_1ПО'],
                                             values=['ФИО_именительный'],
                                             aggfunc='count')
        po_age_distribution = po_age_distribution.reset_index()
    else:
        po_age_distribution = pd.DataFrame(columns=['Возрастная_категория_1ПО', 'Количество'])
    return po_age_distribution


# Функции обработки данных для вкладки Обработка данных
def calculate_data():
    """
    Функция для подсчета данных из файлов
    :return:
    """
    count = 0
    count_errors = 0
    quantity_files = len(names_files_calculate_data)
    current_time = time.strftime('%H_%M_%S')
    # Состояние чекбокса
    mode_text = mode_text_value.get()

    # Получаем название обрабатываемого листа
    name_list_df = pd.read_excel(name_file_params_calculate_data, nrows=2)
    name_list = name_list_df['Значение'].loc[0]

    # Получаем количество листов в файле, на случай если название листа не совпадает с правильным
    quantity_list_in_file = name_list_df['Значение'].loc[1]

    # Получаем шаблон с данными, первую строку пропускаем, поскольку название обрабатываемого листа мы уже получили
    df = pd.read_excel(name_file_params_calculate_data, skiprows=2)

    # Создаем словарь параметров
    param_dict = dict()

    for row in df.itertuples():
        param_dict[row[1]] = row[2]
    # Создаем словарь для подсчета данных, копируя ключи из словаря параметров, значения в зависимости от способа обработки

    if mode_text == 'Yes':
        result_dct = {key: '' for key, value in param_dict.items()}
    else:
        result_dct = {key: 0 for key, value in param_dict.items()}

        # Создаем датафрейм для контроля процесса подсчета и заполняем словарь на основе которого будем делать итоговую таблицу

    check_df = pd.DataFrame(columns=param_dict.keys())
    # Вставляем колонку для названия файла
    check_df.insert(0, 'Название файла', '')
    for file in names_files_calculate_data:
        # Проверяем чтобы файл не был резервной копией.
        if '~$' in file:
            continue
        # Создаем словарь для создания строки которую мы будем добавлять в проверочный датафрейм
        new_row = dict()
        # Получаем  отбрасываем расширение файла
        full_name_file = file.split('.')[0]
        # Получаем имя файла  без пути
        name_file = full_name_file.split('/')[-1]
        try:

            new_row['Название файла'] = name_file

            wb = openpyxl.load_workbook(file)
            # Проверяем наличие листа
            if name_list in wb.sheetnames:
                sheet = wb[name_list]
            # проверяем количество листов в файле.Если значение равно 1 то просто берем первый лист, иначе вызываем ошибку
            elif quantity_list_in_file == 1:
                temp_name = wb.sheetnames[0]
                sheet = wb[temp_name]
            else:
                raise Exception

            for key, cell in param_dict.items():
                result_dct[key] += check_data(sheet[cell].value, mode_text)
                new_row[key] = sheet[cell].value

            temp_df = pd.DataFrame(new_row, index=['temp_index'])
            check_df = pd.concat([check_df, temp_df], ignore_index=True)

            # check_df = check_df.append(new_row, ignore_index=True)

            count += 1
        # Ловим исключения
        except NameError:
            messagebox.showerror('ЦОПП Бурятия',
                                 'Выберите файл с параметрами,обрабатываемые данные, конечную папку')
        except Exception as err:
            count_errors += 1
            with open(f'{path_to_end_folder_calculate_data}/ERRORS {current_time}.txt', 'a', encoding='utf-8') as f:
                f.write(f'Файл {name_file} не обработан!!!\n')

    check_df.to_excel(f'{path_to_end_folder_calculate_data}/Проверка вычисления.xlsx', index=False)

    # Создание итоговой таблицы результатов подсчета

    finish_result = pd.DataFrame()

    finish_result['Наименование показателя'] = result_dct.keys()
    finish_result['Значение показателя'] = result_dct.values()
    # Проводим обработку в зависимости от значения переключателя

    if mode_text == 'Yes':
        # Обрабатываем датафрейм считая текстовые данные
        count_text_df = count_text_value(finish_result)
        count_text_df.to_excel(f'{path_to_end_folder_calculate_data}/Подсчет текстовых значений.xlsx')
    else:
        finish_result.to_excel(f'{path_to_end_folder_calculate_data}/Итоговые значения.xlsx', index=False)

    if count_errors != 0:
        messagebox.showinfo('ЦОПП Бурятия',
                            f'Обработка файлов завершена!\nОбработано файлов:  {count} из {quantity_files}\n Необработанные файлы указаны в файле {path_to_end_folder_calculate_data}/ERRORS {current_time}.txt ')
    else:
        messagebox.showinfo('ЦОПП Бурятия',
                            f'Обработка файлов успешно завершена!\nОбработано файлов:  {count} из {quantity_files}')


def count_text_value(df):
    """
    Функция для подсчета количества вариантов того или иного показателя
    :param df: датафрейм с сырыми данными. Название показателя значение показателя(строка разделенная ;)
    :return: обработанный датафрейм с мультиндексом, где (Название показателя это индекс верхнего уровня, вариант показателя это индекс второго уровня а значение это сколько раз встречался
    этот вариант в обрабатываемых файлах)
    """
    data = dict()

    #
    for row in df.itertuples():
        value = row[2]
        if type(value) == float or type(value) == int:
            continue
        # Создаем список, разделяя строку по ;
        lst_value = row[2].split(';')[:-1]
        #     # Отрезаем последний элемент, поскольку это пустое значение
        temp_df = pd.DataFrame({'Value': lst_value})
        counts_series = temp_df['Value'].value_counts()
        # Делаем индекс колонкой и превращаем в обычную таблицу
        index_count_values = counts_series.reset_index()
        # Итерируемся по таблице.Это делается чтобы заполнить словарь на основе которого будет создаваться итоговая таблица
        for count_row in index_count_values.itertuples():
            # print(count_row)
            # Заполняем словарь
            data[(row[1], count_row[1])] = count_row[2]
    # Создаем на основе получившегося словаря таблицу
    out_df = pd.Series(data).to_frame().reset_index()
    out_df = out_df.set_index(['level_0', 'level_1'])
    out_df.index.names = ['Название показателя', 'Вариант показателя']
    out_df.rename(columns={0: 'Количество'}, inplace=True)
    return out_df


def check_data(cell, text_mode):
    """
    Функция для проверки значения ячейки. Для обработки пустых значений, строковых значений, дат
    :param cell: значение ячейки
    :return: 0 если значение ячейки не число
            число если значение ячейки число(ха звучит глуповато)
    думаю функция должна работать с дополнительным параметром, от которого будет зависеть подсчет значений навроде галочек или плюсов в анкетах или опросах.
    """
    # Проверяем режим работы. если текстовый, то просто складываем строки
    if text_mode == 'Yes':
        if cell is None:
            return ''
        else:
            temp_str = str(cell)
            return f'{temp_str};'
    # Если режим работы стандартный. Убрал подсчет строк и символов в числовом режиме, чтобы не запутывать.
    else:
        if cell is None:
            return 0
        if type(cell) == int:
            return cell
        elif type(cell) == float:
            return cell
        else:
            return 0

def create_report_indicator():
    """
    Функция для создания отчетов по индикаторам и госзаданию
    """
    try:
        dpo_df = pd.read_excel(name_file_data_report, sheet_name='ДПО')
        po_df = pd.read_excel(name_file_data_report, sheet_name='ПО')

        dpo_df['for_counting'] = 1
        po_df['for_counting'] = 1

        dpo_df = dpo_df[dpo_df[
                            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'] == 'Повышение квалификации']

        # Создаем файл excel
        wb = openpyxl.Workbook()
        # Создаем листы
        wb.create_sheet(title='Индикаторы', index=0)
        wb.create_sheet(title='Госзадание', index=1)
        # Удаляем пустой лист
        del wb['Sheet']
        # Создаем цвет заливки
        yellowFill = PatternFill(start_color='ffff00',
                                 end_color='ffff00',
                                 fill_type='solid')

        # Считаем индикаторы
        # Создаем базовый датафрейм
        base_df = pd.DataFrame(index=['программы профессиональных модулей для среднего профессионального образования',
                                      'программам для обучающихся общеобразовательных организаций',
                                      'программы под заказ работодателей', 'отраслевые программы',
                                      'программы для граждан предпенсионного возраста',
                                      'программы по компетенциям будущего, включая компетенции цифровой экономики'])

        # Отбираем нужные колонки
        dpo_choice = dpo_df[['for_counting', 'Индикатор', 'Месяц_окончания_курса']]
        po_choice = po_df[['for_counting', 'Индикатор', 'Месяц_окончания_курса']]

        # Объединяем в единый датафрейм
        indicator_df = pd.concat([dpo_choice, po_choice], ignore_index=True)

        for i in range(1, 13):
            temp_df = indicator_df[indicator_df['Месяц_окончания_курса'] == i]
            temp_group = temp_df.groupby(['Индикатор']).agg({'for_counting': sum})
            base_df = pd.concat([base_df, temp_group], axis=1)

        # Меняем названия колонок
        base_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь',
                           'ноябрь', 'декабрь']

        # Суммируем столбцы, превращаем в датафрейм и транспонируем столбец в строку
        sum_month = base_df.sum(axis=0).to_frame().T
        # Меняем индекс
        sum_month.index = ['Всего']

        base_df = pd.concat([base_df, sum_month])

        # Сохраняем на лист
        for r in dataframe_to_rows(base_df, index=True, header=True):
            if len(r) != 1:
                wb['Индикаторы'].append(r)

        wb['Индикаторы'].column_dimensions['A'].width = 90

        # Считаем госздание
        # ДПО ПК (ВО)
        base_vo_df = pd.DataFrame(index=['ДПО ПК(ВО)'])
        vo_df = dpo_df[dpo_df['Уровень_образования_ВО_СПО'] == 'Высшее образование']
        vo_df_choice = vo_df[['for_counting', 'Месяц_окончания_курса']]

        for i in range(1, 13):
            temp_df = vo_df_choice[vo_df_choice['Месяц_окончания_курса'] == i]
            temp_group = temp_df.groupby(['Месяц_окончания_курса']).agg({'for_counting': sum})
            if temp_group.shape[0] == 1:
                temp_group.index = ['ДПО ПК(ВО)']
                base_vo_df = pd.concat([base_vo_df, temp_group], axis=1)
            elif temp_group.shape[0] == 0:
                empty_group = pd.DataFrame(index=['ДПО ПК(ВО)'], columns=['for_counting'], data=np.NaN)
                base_vo_df = pd.concat([base_vo_df, empty_group], axis=1)

        # Меняем названия колонок
        base_vo_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь',
                              'ноябрь', 'декабрь']

        # Считаем СПО
        base_spo_df = pd.DataFrame(index=['ДПО ПК(СПО)'])
        spo_df = dpo_df[dpo_df['Уровень_образования_ВО_СПО'] == 'Среднее профессиональное образование']
        spo_df_choice = spo_df[['for_counting', 'Месяц_окончания_курса']]

        for i in range(1, 13):
            temp_df = spo_df_choice[spo_df_choice['Месяц_окончания_курса'] == i]
            temp_group = temp_df.groupby(['Месяц_окончания_курса']).agg({'for_counting': sum})
            if temp_group.shape[0] == 1:
                temp_group.index = ['ДПО ПК(СПО)']
                base_spo_df = pd.concat([base_spo_df, temp_group], axis=1)
            elif temp_group.shape[0] == 0:
                empty_group = pd.DataFrame(index=['ДПО ПК(СПО)'], columns=['for_counting'], data=np.NaN)
                base_spo_df = pd.concat([base_spo_df, empty_group], axis=1)

        base_spo_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь',
                               'октябрь',
                               'ноябрь', 'декабрь']

        # Считаем ПО
        base_po_df = pd.DataFrame(index=['ПО'])
        po_df_choice = po_df[['for_counting', 'Месяц_окончания_курса']]

        for i in range(1, 13):
            temp_df = po_df_choice[po_df_choice['Месяц_окончания_курса'] == i]
            temp_group = temp_df.groupby(['Месяц_окончания_курса']).agg({'for_counting': sum})
            if temp_group.shape[0] == 1:
                temp_group.index = ['ПО']
                base_po_df = pd.concat([base_po_df, temp_group], axis=1)
            elif temp_group.shape[0] == 0:
                empty_group = pd.DataFrame(index=['ПО'], columns=['for_counting'], data=np.NaN)
                base_po_df = pd.concat([base_po_df, empty_group], axis=1)

        base_po_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь',
                              'ноябрь', 'декабрь']

        out_df = pd.concat([base_vo_df, base_spo_df, base_po_df])

        # Суммируем столбцы, превращаем в датафрейм и транспонируем столбец в строку
        sum_month_out_df = out_df.sum(axis=0).to_frame().T
        # Меняем индекс
        sum_month_out_df.index = ['Всего']

        out_df = pd.concat([out_df, sum_month_out_df])

        # Сохраняем на лист
        for r in dataframe_to_rows(out_df, index=True, header=True):
            if len(r) != 1:
                wb['Госзадание'].append(r)

        wb['Госзадание'].column_dimensions['A'].width = 50

        # заливка строки

        # заливаем строку
        for cell in wb['Индикаторы'][8]:
            cell.fill = yellowFill

        for cell in wb['Госзадание'][5]:
            cell.fill = yellowFill

        # Получаем текущее время для того чтобы использовать в названии
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder_report}/Отчет по индикаторам и госзаданию {current_time}.xlsx')
    except NameError:
        messagebox.showerror('ЦОПП Бурятия',
                             'Выберите файл с параметрами,обрабатываемые данные, конечную папку')
    else:
        messagebox.showinfo('ЦОПП Бурятия','Создание отчета по индикаторам и госзаданию завершено')


if __name__ == '__main__':
    window = Tk()
    window.title('ЦОПП Бурятия')
    window.geometry('700x970')
    window.resizable(False, False)



    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку создания документов по шаблону
    tab_create_doc = ttk.Frame(tab_control)
    tab_control.add(tab_create_doc, text='Создание документов')
    tab_control.pack(expand=1, fill='both')

    # Создаем вкладку для создания сводной таблицы
    tab_create_general_table = ttk.Frame(tab_control)
    tab_control.add(tab_create_general_table, text='Создание сводной таблицы')
    tab_control.pack(expand=1, fill='both')

    # Создаем вкладку создания отчетов
    tab_create_report = ttk.Frame(tab_control)
    tab_control.add(tab_create_report, text='Создание отчетов')
    tab_control.pack(expand=1, fill='both')

    # Создаем вкладку для обработки таблиц excel  с одинаковой структурой
    tab_calculate_data = ttk.Frame(tab_control)
    tab_control.add(tab_calculate_data, text='Обработка данных')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Создание документов
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_create_doc,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nГенерация документов по шаблону')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')
    img = PhotoImage(file=path_to_img)
    Label(tab_create_doc,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Переключатель:индивидуальный или списочный приказл
    # Создаем переменную хранящую тип документа, в зависимости от значения будет использоваться та или иная функция
    group_rb_type_doc = IntVar()
    # Создаем фрейм для размещения переключателей(pack и грид не используются в одном контейнере)
    frame_rb_type_doc = LabelFrame(tab_create_doc, text='Выберите тип создаваемого документа')
    frame_rb_type_doc.grid(column=0, row=1, padx=10)
    #
    Radiobutton(frame_rb_type_doc, text='Индивидуальные документы', variable=group_rb_type_doc, value=0).pack()
    Radiobutton(frame_rb_type_doc, text='Списочный документ', variable=group_rb_type_doc, value=1).pack()

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_doc = LabelFrame(tab_create_doc, text='Подготовка')
    frame_data_for_doc.grid(column=0, row=2, padx=10)

    # Создаем кнопку Выбрать шаблон
    btn_template_doc = Button(frame_data_for_doc, text='1) Выберите шаблон документа', font=('Arial Bold', 20),
                              command=select_file_template_doc
                              )
    btn_template_doc.grid(column=0, row=3, padx=10, pady=10)
    #
    # Создаем кнопку Выбрать файл с данными
    btn_data_doc = Button(frame_data_for_doc, text='2) Выберите файл с данными', font=('Arial Bold', 20),
                          command=select_file_data_doc
                          )
    btn_data_doc.grid(column=0, row=4, padx=10, pady=10)

    # Поле для ввода названия генериуемых документов
    # Определяем текстовую переменную
    entry_name_file = StringVar()
    # Описание поля
    label_name_column_name_file = Label(frame_data_for_doc, text='3) Введите название создаваемых документов\n'
                                                                 'например Договор,Справка,Ведомость  и т.п.')
    label_name_column_name_file.grid(column=0, row=5, padx=10, pady=10)
    # поле ввода
    type_file_column_entry = Entry(frame_data_for_doc, textvariable=entry_name_file, width=30)
    type_file_column_entry.grid(column=0, row=6, padx=5, pady=5, ipadx=30, ipady=15)

    #
    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_doc = Button(frame_data_for_doc, text='4) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder_doc
                                       )
    btn_choose_end_folder_doc.grid(column=0, row=7, padx=10, pady=10)

    # Создаем переменную для хранения результа переключения чекбокса
    mode_combine_value = StringVar()

    # Устанавливаем значение по умолчанию для этой переменной. По умолчанию будет вестись подсчет числовых данных
    mode_combine_value.set('No')
    # Создаем чекбокс для выбора режима подсчета

    chbox_mode_calculate = Checkbutton(frame_data_for_doc,
                                       text='Поставьте галочку, если вам нужно чтобы все документы\n были объединены в один файл',
                                       variable=mode_combine_value,
                                       offvalue='No',
                                       onvalue='Yes')
    chbox_mode_calculate.grid(column=0, row=8, padx=10, pady=10)

    #
    # Создаем кнопку для запуска функции генерации файлов ДПО

    btn_create_files_dpo = Button(tab_create_doc, text='Создать документы ДПО', font=('Arial Bold', 20),
                                  command=generate_docs_dpo
                                  )
    btn_create_files_dpo.grid(column=0, row=9, padx=10, pady=10)

    # Создаем кнопку для запуска функции генерации файлов ПО
    btn_create_files_po = Button(tab_create_doc, text='Создать документы ПО', font=('Arial Bold', 20),
                                 command=generate_docs_po
                                 )
    btn_create_files_po.grid(column=0, row=10, padx=10, pady=10)

    # Создаем кнопку для создания документов из таблиц с произвольной структурой
    btn_create_files_other = Button(tab_create_doc, text='Создать документы\n из произвольной таблицы',
                                    font=('Arial Bold', 20),
                                    command=generate_docs_other
                                    )
    btn_create_files_other.grid(column=0, row=11, padx=10, pady=10)




    # Добавляем виджеты на вкладку создания отчетов
    lbl_hello = Label(tab_create_report,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nСоздание отчетов')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка . Пришлось переименовывать переменную, иначе картинка не отображалась
    path_to_img_report = resource_path('logo.png')
    img_report = PhotoImage(file=path_to_img_report)
    Label(tab_create_report,
          image=img_report
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_report = LabelFrame(tab_create_report, text='Подготовка')
    frame_data_for_report.grid(column=0, row=2, padx=10)

    # Создаем кнопку Выбрать файл с данными
    btn_data_report = Button(frame_data_for_report, text='1) Выберите файл с данными', font=('Arial Bold', 20),
                             command=select_file_data_report
                             )
    btn_data_report.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_report = Button(frame_data_for_report, text='2) Выберите конечную папку',
                                          font=('Arial Bold', 20),
                                          command=select_end_folder_report
                                          )
    btn_choose_end_folder_report.grid(column=0, row=5, padx=10, pady=10)

    # Создаем облать для размещения кнопок создания отчетов
    frame_create_report = LabelFrame(tab_create_report, text='Создание отчетов')
    frame_create_report.grid(column=0, row=6, padx=10)

    # Создание сводного отчета по показателям ЦОПП

    btn_report_svod = Button(frame_create_report, text='Создать сводный отчет', font=('Arial Bold', 20),
                             command=create_report_svod
                             )
    btn_report_svod.grid(column=0, row=7, padx=10, pady=10)

    btn_report_one_pk = Button(frame_create_report, text='Подсчитать часть данных\nдля отчета 1-ПК',
                               font=('Arial Bold', 15),
                               command=create_report_one_pk
                               )
    btn_report_one_pk.grid(column=0, row=8, padx=10, pady=10)

    btn_report_one_po = Button(frame_create_report, text='Подсчитать часть данных\nдля отчета 1-ПО',
                               font=('Arial Bold', 15),
                               command=create_report_one_po
                               )
    btn_report_one_po.grid(column=0, row=9, padx=10, pady=10)


    # Создание отчета по индикаторам и госзаданию
    btn_report_indicator = Button(frame_create_report, text='Отчет по индикаторам\nи госзаданию',
                               font=('Arial Bold', 15),
                               command=create_report_indicator
                               )
    btn_report_indicator.grid(column=0, row=11, padx=10, pady=10)


    # размещаем виджеты на вкладке Прочее
    lbl_hello = Label(tab_create_general_table,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nСоздание сводной таблицы по всем курсам')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка . Пришлось переименовывать переменную, иначе картинка не отображалась
    path_to_img_other = resource_path('logo.png')
    img_other = PhotoImage(file=path_to_img_report)
    Label(tab_create_general_table,
          image=img_other
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_other = LabelFrame(tab_create_general_table, text='Подготовка')
    frame_data_for_other.grid(column=0, row=2, padx=10)

    # Создаем кнопку для выбора шаблона таблицы
    btn_table_other_template = Button(frame_data_for_other, text='Выберите шаблон таблицы', font=('Arial Bold', 20),
                                      command=select_file_template_table
                                      )
    btn_table_other_template.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку Выбрать файлы с данными
    btn_data_other = Button(frame_data_for_other, text='Выберите папку\n с данными всех курсов',
                            font=('Arial Bold', 20),
                            command=select_files_data_groups
                            )
    btn_data_other.grid(column=0, row=4, padx=10, pady=10)
    #
    btn_choose_end_folder_doc = Button(frame_data_for_other, text='Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder_doc
                                       )
    btn_choose_end_folder_doc.grid(column=0, row=5, padx=10, pady=10)

    # Кнопка создать общую таблицу

    btn_create_general_table = Button(tab_create_general_table, text='Создать общую таблицу', font=('Arial Bold', 20),
                                      command=create_general_table
                                      )
    btn_create_general_table.grid(column=0, row=6, padx=10, pady=10)



    # Добавляем виджеты на вклдаку Обработки данных
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_calculate_data,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nПодсчет данных из файлов Excel\nс одинаковой структурой')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_calculate = resource_path('logo.png')
    img_calculate = PhotoImage(file=path_to_img)
    Label(tab_calculate_data,
          image=img_calculate
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с параметрами
    btn_select_file_params = Button(tab_calculate_data, text='1) Выбрать файл с параметрами', font=('Arial Bold', 20),
                                    command=select_file_params_calculate_data
                                    )
    btn_select_file_params.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку Выбрать файл с данными
    btn_select_files_data = Button(tab_calculate_data, text='2) Выбрать файлы с данными', font=('Arial Bold', 20),
                                   command=select_files_data_calculate_data
                                   )
    btn_select_files_data.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_calculate_data, text='3) Выбрать конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder_calculate_data
                                   )
    btn_choose_end_folder.grid(column=0, row=4, padx=10, pady=10)

    # Создаем переменную для хранения результа переключения чекбокса
    mode_text_value = tkinter.StringVar()

    # Устанавливаем значение по умолчанию для этой переменной. По умолчанию будет вестись подсчет числовых данных
    mode_text_value.set('No')
    # Создаем чекбокс для выбора режима подсчета

    chbox_mode_calculate = tkinter.Checkbutton(tab_calculate_data,
                                               text='Поставьте галочку, если вам нужно посчитать текстовые данные ',
                                               variable=mode_text_value,
                                               offvalue='No',
                                               onvalue='Yes')
    chbox_mode_calculate.grid(column=0, row=5, padx=10, pady=10)

    # Создаем кнопку для запуска подсчета файлов

    btn_calculate = Button(tab_calculate_data, text='4) Подсчитать', font=('Arial Bold', 20),
                           command=calculate_data
                           )
    btn_calculate.grid(column=0, row=6, padx=10, pady=10)

    window.mainloop()
