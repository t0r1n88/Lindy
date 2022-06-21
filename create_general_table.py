import tkinter
import numpy as np
import pandas as pd
import os
from dateutil.parser import ParserError
from docxtpl import DocxTemplate
from docxcompose.composer import Composer
from docx import Document
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import time
import datetime
from datetime import date
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, Reference, PieChart, PieChart3D, Series
pd.options.display.max_colwidth = 100
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import re
import tempfile


def calculate_age(born):
    """
    Функция для расчета текущего возраста взято с https://stackoverflow.com/questions/2217488/age-from-birthdate-in-python/9754466#9754466
    :param born: дата рождения
    :return: возраст
    """

    try:
        today = date.today()
        return today.year - born.year - ((today.month, today.day) < (born.month, born.day))
    except TypeError:
        print(born)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()
    except ValueError:
        print(f' Ошибка при подсчете текущего возраста ячейки {born}')
        messagebox.showerror('ЦОПП Бурятия', 'Пустая ячейка с датой или некорректная запись!!!')
        quit()
    except:
        print(f' Ошибка при подсчете текущего возраста ячейки {born}')
        messagebox.showerror('ЦОПП Бурятия', 'Отсутствует или некорректная дата \nПроверьте файл!')
        quit()

def check_date_columns(i, value):
    """
    Функция для проверки типа колонки. Необходимо найти колонки с датой
    :param i:
    :param value:
    :return:
    """
    #  Да да это просто
    if '00:00:00' in str(value):
        try:
            itog = pd.to_datetime(str(value),infer_datetime_format=True)

        except ParserError:
            pass
        except ValueError:
            pass
        except TypeError:
            pass
        else:
            return i

def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return ''
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()

def convert_date(cell):
    """
    Функция для конвертации даты в формате 1957-05-10 в формат 10.05.1957(строковый)
    """

    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except TypeError:
        print(f' Ошибка при конвертации ячейки {cell}')
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()
    except ValueError:
        print(f' Ошибка при конвертации ячейки {cell}')
        messagebox.showerror('ЦОПП Бурятия', 'Пустая ячейка с датой или некорректная запись!!!')
        quit()

def extract_date_begin_course(cell:str):
    """
    Функция для извлечения даты начала курса
    """

    try:
        # Находим обе даты
        match = re.findall(r'\d\d.\d\d.\d\d\d\d', cell)
        # date_course = datetime.datetime.strptime(match[0], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return match[0]
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()



def extract_date_end_course(cell:str):
    """
    Функция для извлечения даты окончания курса
    """
    try:
        # Находим обе даты
        match = re.findall(r'\d\d.\d\d.\d\d\d\d', cell)
        # Конвертируем строку
        # date_course = datetime.datetime.strptime(match[1], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return match[1]
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()

def extract_month_begin_course(cell:str):
    """
    Функция для извлечения месяца начала курса в формате от 1 до 12
    """
    try:
        # Находим оба месяца выделив месяц круглыми скобками
        match = re.findall(r'\d\d.(\d\d).\d\d\d\d', cell)
        # Конвертируем строку
        # date_course = datetime.datetime.strptime(match[1], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return int(match[0])
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()


def extract_month_end_course(cell:str):
    """
    Функция для извлечения месяца окончания курса в формате от 1 до 12
    """
    try:
        # Находим обе даты
        match = re.findall(r'\d\d.(\d\d).\d\d\d\d', cell)
        # Конвертируем строку
        # date_course = datetime.datetime.strptime(match[1], '%d.%m.%Y')
        # string_date = datetime.datetime.strftime(date_course, '%d.%m.%Y')
        return int(match[1])
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()
    except IndexError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячейки \nПериод_обучения_в_формате_с_дата_начала_по_дата_окончания!!!'
                                             '\nГод должен состоять из 4 цифр(Например 2022)!!!')
        quit()

name_file_template_table ='Шаблон Базы Данных от 14.03.2022.xlsx'


path_to_files_groups = 'data\Тест'

path_to_end_folder_doc = 'data'
pattern = re.compile(
        '^[А-ЯЁ]+_.+_(?:январь|февраль|март|апрель|май|июнь|июль|август|сентябрь|октябрь|ноябрь|декабрь)\.xlsx$')
try:
    # Получаем базовые датафреймы
    df_dpo = pd.read_excel(name_file_template_table, sheet_name='ДПО',
                           dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                  'Серия_паспорта_в_формате_1111':str,
                                  'Номер_паспорта_в_формате_111111':str})
    df_po = pd.read_excel(name_file_template_table, sheet_name='ПО',
                          dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                 'Серия_паспорта_совершеннолетнего_или_родителя_законного_представителя_в_формате_1111': str,
                                 'Номер_паспорта_в_формате_111111': str
                                 })
    # Очищаем базовые датафреймы на случай  если там есть какие то строки. Необходимо чтобы шаблон был полностью пуст
    df_dpo = df_dpo.iloc[0:0]
    df_po = df_po.iloc[0:0]

    # Добавляем 2 колонки с возрастом и категорией для каждого базового датафрейма.Чтобы конкатенация прошла успешно
    df_dpo['Текущий_возраст'] = np.nan
    df_dpo['Возрастная_категория_1ПК'] = np.nan
    df_dpo['Дата_начала_курса'] = np.nan
    df_dpo['Дата_окончания_курса'] = np.nan
    df_dpo['Месяц_начала_курса'] = np.nan
    df_dpo['Месяц_окончания_курса'] = np.nan


    df_po['Текущий_возраст'] = np.nan
    df_po['Возрастная_категория_1ПО'] = np.nan
    df_po['Дата_начала_курса'] = np.nan
    df_po['Дата_окончания_курса'] = np.nan
    df_po['Месяц_начала_курса'] = np.nan
    df_po['Месяц_окончания_курса'] = np.nan

    # Получаем множество из навзваний колонок в шаблоне для каждого листа
    dpo_template_cols = set(df_dpo.columns)
    po_template_cols = set(df_po.columns)

    # Перебираем файлы собирая данные в промежуточные датафреймы и добавляя их в базовые
    for dirpath, dirnames, filenames in os.walk(path_to_files_groups):
        for filename in filenames:

            if re.search(pattern, filename):
                print("Файл:", os.path.join(dirpath, filename))
                # Создаем промежуточный датафрейм с данными с листа ДПО
                temp_dpo = pd.read_excel(os.path.join(dirpath, filename), sheet_name='ДПО',
                                         dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                                'Серия_паспорта_в_формате_1111': str,
                                                'Номер_паспорта_в_формате_111111': str,
                                                'Серия_документа_о_ВО_СПО':str,
                                                'Номер_документа_о_ВО_СПО':str})
                # Создаем промежуточный датафрейм с данными с листа ДПО
                temp_po = pd.read_excel(os.path.join(dirpath, filename), sheet_name='ПО',
                                        dtype={'Гражданство_получателя_код_страны_по_ОКСМ': str,
                                               'Серия_паспорта_совершеннолетнего_или_родителя_законного_представителя_в_формате_1111': str,
                                               'Номер_паспорта_в_формате_111111': str})
                # Добавляем 2 колонки с характеристиками возраста
                temp_dpo['Текущий_возраст'] = temp_dpo['Дата_рождения_получателя'].apply(calculate_age)
                temp_dpo['Возрастная_категория_1ПК'] = pd.cut(temp_dpo['Текущий_возраст'],
                                                              [0, 24, 29, 34, 39, 44, 49, 54, 59, 64, 101, 10000],
                                                              labels=['моложе 25 лет', '25-29', '30-34', '35-39',
                                                                      '40-44', '45-49', '50-54', '55-59', '60-64',
                                                                      '65 и более',
                                                                      'Возраст  больше 101'])
                # Добавляем 4 колонки с характеристиками дат курсов
                temp_dpo['Дата_начала_курса'] = temp_dpo['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_begin_course)
                temp_dpo['Дата_окончания_курса'] = temp_dpo['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_end_course)
                temp_dpo['Месяц_начала_курса'] = temp_dpo['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_month_begin_course)
                temp_dpo['Месяц_окончания_курса'] = temp_dpo['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_month_end_course)
                # Обрабатываем датафрейм с ПО
                temp_po['Текущий_возраст'] = temp_po['Дата_рождения_получателя'].apply(calculate_age)
                temp_po['Возрастная_категория_1ПО'] = pd.cut(temp_po['Текущий_возраст'],
                                                             [0, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25,
                                                              26, 27, 28,
                                                              29, 34, 39, 44, 49, 54, 59, 64, 101],
                                                             labels=['В возрасте моложе 14 лет', '14 лет', '15 лет',
                                                                     '16 лет',
                                                                     '17 лет', '18 лет', '19 лет', '20 лет',
                                                                     '21 год', '22 года',
                                                                     '23 года', '24 года', '25 лет',
                                                                     '26 лет', '27 лет', '28 лет', '29 лет',
                                                                     '30-34 лет',
                                                                     '35-39 лет', '40-44 лет', '45-49 лет',
                                                                     '50-54 лет',
                                                                     '55-59 лет',
                                                                     '60-64 лет',
                                                                     '65 лет и старше'])
                # Добавляем 4 колонки с характеристиками дат курсов
                temp_po['Дата_начала_курса'] = temp_po['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_begin_course)
                temp_po['Дата_окончания_курса'] = temp_po['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_date_end_course)
                temp_po['Месяц_начала_курса'] = temp_po['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_month_begin_course)
                temp_po['Месяц_окончания_курса'] = temp_po['Период_обучения_в_формате_с_дата_начала_по_дата_окончания'].apply(extract_month_end_course)


                # Конвертируем  столбцы с датами в краткий формат
                temp_dpo['Дата_выдачи_документа'] = temp_dpo['Дата_выдачи_документа'].apply(convert_date)
                temp_dpo['Дата_рождения_получателя'] = temp_dpo['Дата_рождения_получателя'].apply(convert_date)
                temp_dpo['Дата_выдачи_паспорта'] = temp_dpo['Дата_выдачи_паспорта'].apply(convert_date)

                temp_po['Дата_выдачи_документа'] = temp_po['Дата_выдачи_документа'].apply(convert_date)
                temp_po['Дата_рождения_получателя'] = temp_po['Дата_рождения_получателя'].apply(convert_date)
                temp_po['Дата_выдачи_паспорта'] = temp_po['Дата_выдачи_паспорта'].apply(convert_date)

                # Проверяем состав колонок
                temp_dpo_columns = set(temp_dpo.columns)
                temp_po_columns = set(temp_po.columns)
                # Если есть разница то выдаем сообщение предупреждение
                diff_cols_dpo = dpo_template_cols-temp_dpo_columns
                diff_cols_po = po_template_cols-temp_po_columns

                if len(diff_cols_dpo) > 0:
                    messagebox.showerror('ЦОПП Бурятия',f'В файле {filename} на листе ДПО отличается состав колонок по сравнению с шаблоном {name_file_template_table}\n Проверьте наличие указанных колонок в обоих файлах: {diff_cols_dpo}\nдля корректной обработки')

                if len(diff_cols_po) > 0:
                    messagebox.showerror('ЦОПП Бурятия',f'В файле {filename} на листе ПО отличается состав колонок по сравнению с шаблоном {name_file_template_table}\n Проверьте наличие указанных колонок в обоих файлах: {diff_cols_po}\nдля корректной обработки')





                # Добавляем промежуточные датафреймы в исходные
                #

                df_dpo = pd.concat([df_dpo, temp_dpo], ignore_index=True)
                df_po = pd.concat([df_po, temp_po], ignore_index=True)




    # Код сохранения датафрейма в разные листы и сохранением форматирования  взят отсюда https://azzrael.ru/python-pandas-openpyxl-excel
    wb = openpyxl.load_workbook(name_file_template_table)

    # Записываем лист ДПО

    for ir in range(0, len(df_dpo)):
        for ic in range(0, len(df_dpo.iloc[ir])):
            wb['ДПО'].cell(2 + ir, 1 + ic).value = df_dpo.iloc[ir][ic]

    wb['ДПО']['BO1'] = 'Текущий_возраст'
    wb['ДПО']['BP1'] = 'Возрастная_категория_1ПК'
    wb['ДПО']['BQ1'] = 'Дата_начала_курса'
    wb['ДПО']['BR1'] = 'Дата_окончания_курса'
    wb['ДПО']['BS1'] = 'Месяц_начала_курса'
    wb['ДПО']['BT1'] = 'Месяц_окончания_курса'


    # Записываем лист ПО
    for ir in range(0, len(df_po)):
        for ic in range(0, len(df_po.iloc[ir])):
            wb['ПО'].cell(2 + ir, 1 + ic).value = df_po.iloc[ir][ic]
    wb['ПО']['BJ1'] = 'Текущий_возраст'
    wb['ПО']['BK1'] = 'Возрастная_категория_1ПО'
    wb['ПО']['BL1'] = 'Дата_начала_курса'
    wb['ПО']['BM1'] = 'Дата_окончания_курса'
    wb['ПО']['BN1'] = 'Месяц_начала_курса'
    wb['ПО']['BO1'] = 'Месяц_окончания_курса'


    # Получаем текущее время для того чтобы использовать в названии

    t = time.localtime()
    current_time = time.strftime('%d_%m_%y', t)
    # Сохраняем итоговый файл
    wb.save(f'{path_to_end_folder_doc}/Общая таблица слушателей ЦОПП от {current_time}.xlsx')
except NameError as e:
    messagebox.showinfo('ЦОПП Бурятия', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
# except:
#     messagebox.showerror('ЦОПП Бурятия',
#                          'Возникла ошибка,проверьте шаблон таблицы\nДобавляемы файлы должны иметь одинаковую структуру с шаблоном таблицы')
else:
    messagebox.showinfo('ЦОПП Бурятия', 'Общая таблица успешно создана!')



