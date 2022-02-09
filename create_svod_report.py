import pandas as pd
import openpyxl
import time
from openpyxl.styles import Font
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, Reference,PieChart,PieChart3D,Series
from copy import deepcopy


# Функции для создания сводной таблицы
def counting_total_student(dpo_df, po_df):
    """
    Функция для подсчета общего количества студентов обучающихся в цопп, и количества обучающихся по ДПО И ПО
    :param dpo_df: датафрейм ДПО
    :param po_df: датафрейм ПО
    :return: кортеж вида: общее количество обучающихся,количество обучающихся ДПО,количество обучающихся ПО
    """
    # количество по типам
    total_dpo = dpo_df.shape[0]
    total_po = po_df.shape[0]
    # общее количество
    total = total_dpo + total_po

    return total, total_dpo, total_po


def counting_type_of_training(dpo, po):
    """
    Функция для создания сводной таблицы по категориям направление подготовки, название программы,количество обучающихся
    :param dpo: датафрейм ДПО
    :param po: датафрейм ПО
    :return: датафрейм сводной таблицы
    """
    # Создаем сводные таблицы проверяее перед этим не пустые ли таблицы
    if dpo.shape[0] > 0:
        dpo_svod_category_and_name = pd.pivot_table(dpo, index=[
            'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка',
            'Наименование_дополнительной_профессиональной_программы'],
                                                    values=['ФИО_именительный'],
                                                    aggfunc='count')
        dpo_svod_category_and_name = dpo_svod_category_and_name.reset_index()
    else:
        dpo_svod_category_and_name = pd.DataFrame(columns=['Направление подготовки', 'Название программы', 'Количество обученных'])

    if po.shape[0] > 0:
        po_svod_category_and_name = pd.pivot_table(po,
                                                   index=['Программа_профессионального_обучения_направление_подготовки',
                                                          'Наименование_программы_профессионального_обучения'],
                                                   values=['ФИО_именительный'],
                                                   aggfunc='count')
        po_svod_category_and_name = po_svod_category_and_name.reset_index()

    else:
        po_svod_category_and_name = pd.DataFrame(
            columns=['Направление подготовки', 'Название программы', 'Количество обученных'])

    # Изменяем названия колонок, чтобы без проблем соединить 2 датафрейма
    dpo_svod_category_and_name.columns = ['Направление подготовки', 'Название программы', 'Количество обученных']
    po_svod_category_and_name.columns = ['Направление подготовки', 'Название программы', 'Количество обученных']
    # Создаем единую сводную таблицу
    general_svod_category_and_name = pd.concat([dpo_svod_category_and_name, po_svod_category_and_name],
                                               ignore_index=True)
    return general_svod_category_and_name


def counting_total_sex(dpo, po):
    """
    Функция для подсчета количества мужчин и женщин
    :param dpo: датафрейм ДПО
    :param po: датафрейм ПО
    :return: датафрейм сводной таблицы
    """
    # Создаем сводные таблицы Проверяем на пустой лист ДПО или ПО
    if dpo.shape[0] > 0:

        dpo_total_sex = pd.pivot_table(dpo, index=['Пол_получателя'],
                                       values=['ФИО_именительный'],
                                       aggfunc='count')
        dpo_total_sex = dpo_total_sex.reset_index()
    else:
        dpo_total_sex = pd.DataFrame(columns=['Пол', 'Количество'])

    if po.shape[0] > 0:
        po_total_sex = pd.pivot_table(po, index=['Пол_получателя'],
                                  values=['ФИО_именительный'],
                                  aggfunc='count')
        po_total_sex = po_total_sex.reset_index()
    else:
        po_total_sex = pd.DataFrame(columns=['Пол', 'Количество'])

    # Переименовываем колонки
    dpo_total_sex.columns = ['Пол', 'Количество']
    po_total_sex.columns = ['Пол', 'Количество']

    # Соединяем в единую таблицу
    general_total_sex = pd.concat([dpo_total_sex, po_total_sex], ignore_index=True)
    # Группируем по полю Пол чтобы суммировать значения
    sum_general_total_sex = general_total_sex.groupby(['Пол']).sum().reset_index()
    return sum_general_total_sex


def counting_age_distribution(dpo, po):
    """
    Функция для подсчета количества обучающихся по возрастным категориям
    :param dpo: датафрейм ДПО
    :param po: датафрейм ПО
    :return: датафрейм сводной таблицы
    """
    # Создаем сводные таблицы
    if dpo.shape[0] > 0:
        dpo_age_distribution = pd.pivot_table(dpo, index=['Возрастная_категория'],
                                              values=['ФИО_именительный'],
                                              aggfunc='count')
        dpo_age_distribution = dpo_age_distribution.reset_index()
    else:
        dpo_age_distribution = pd.DataFrame(columns=['Возрастная_категория', 'Количество'])
    if po.shape[0] > 0:
        po_age_distribution = pd.pivot_table(po, index=['Возрастная_категория'],
                                             values=['ФИО_именительный'],
                                             aggfunc='count')
        po_age_distribution = po_age_distribution.reset_index()
    else:
        po_age_distribution = pd.DataFrame(columns=['Возрастная_категория', 'Количество'])


    # Меняем колонки
    dpo_age_distribution.columns = ['Возрастная_категория', 'Количество']
    po_age_distribution.columns = ['Возрастная_категория', 'Количество']

    # Создаем единую сводную таблицу
    general_age_distribution = pd.concat([dpo_age_distribution, po_age_distribution], ignore_index=True)
    # Повторно группируем чтобы соединить категории из обеих таблиц
    general_age_distribution = general_age_distribution.groupby(['Возрастная_категория']).sum().reset_index()

    return general_age_distribution

name_file_data_report = 'единичная ДПО 2 файла.xlsx'
path_to_end_folder_report ='data/'

dpo_df = pd.read_excel(name_file_data_report, sheet_name='ДПО')
po_df = pd.read_excel(name_file_data_report, sheet_name='ПО')

print(dpo_df.head())
# Заполняем пустые поля для удобства группировки
dpo_df = dpo_df.fillna('Не заполнено!!!')
po_df = po_df.fillna('Не заполнено!!!')
# Создаем переменную для хранения строки на которой заканчивается предыдущий показатель
border_row = 2
border_column = 2

# Получение общего количества прошедших обучение,количества прошедших по ДПО,по ПО
total_students, total_students_dpo, total_students_po = counting_total_student(dpo_df, po_df)

# Количество обучившихся ДПО и ПО

# Получение количества обучившихся по видам
df_counting_type_and_name_trainning = counting_type_of_training(dpo_df, po_df)

# Создаем новый excel файл
wb = openpyxl.Workbook()

# Получаем активный лист
sheet = wb.active
sheet.title = 'Сводные данные'

# Начинаем заполнение листа
# Заполняем количество обучившихся, общее и по типам
sheet['A1'] = 'Наименование показателя'
sheet['A2'] = 'Количество прошедших обучение ДПО'
sheet['A3'] = 'Количество прошедших обучение ПО'
sheet['A4'] = 'Общее количество прошедших обучение в ЦОПП'

sheet['B1'] = 'Количество обучающихся'
sheet['B2'] = total_students_dpo
sheet['B3'] = total_students_po
sheet['B4'] = total_students

# Добавляем круговую диаграмму
pie_main = PieChart()
labels = Reference(sheet, min_col=1, min_row=2, max_row=3)
data = Reference(sheet, min_col=2, min_row=2, max_row=3)

# Для отображения данных на диаграмме
series = Series(data, title='Series 1')
pie_main.append(series)

s1 = pie_main.series[0]
s1.dLbls = DataLabelList()
s1.dLbls.showVal = True

pie_main.add_data(data, titles_from_data=True)
pie_main.set_categories(labels)
pie_main.title = 'Распределение обучившихся'
sheet.add_chart(pie_main, 'F2')
# # Добавляем таблицу с по направлениям

sheet['A7'] = 'Вид обучения'
sheet['B7'] = 'Название программы'
sheet['C7'] = 'Количество обучающихся'

for row in df_counting_type_and_name_trainning.values.tolist():
    sheet.append(row)
# Получаем последние активные ячейки чтобы записывалось по порядку и не налазило друг на друга
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

sheet[f'A{max_row + 2}'] = 'Общее распределение прошедших обучение по полу'
total_sex = counting_total_sex(dpo_df, po_df)
# Добавляем в файл таблицу с распределением по полам
for row in total_sex.values.tolist():
    sheet.append(row)

# Получаем последние активные ячейки чтобы записывалось по порядку и не налазило друг на друга
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Добавляем таблицу с разбиением по возрастам
sheet[f'A{max_row + 2}'] = 'Общее распределение обучающихся по возрасту'
age_distribution = counting_age_distribution(dpo_df, po_df)
for row in age_distribution.values.tolist():
    sheet.append(row)

# Добавляем круговую диаграмму
pie_age = PieChart()
# Для того чтобы не зависело от количества строк в предыдущих таблицах
labels = Reference(sheet, min_col=1, min_row=max_row + 3, max_row=max_row + 2 + len(age_distribution))
data = Reference(sheet, min_col=2, min_row=max_row + 3, max_row=max_row + 2 + len(age_distribution))
# Для отображения данных на диаграмме
series = Series(data, title='Series 1')
pie_age.append(series)

s1 = pie_age.series[0]
s1.dLbls = DataLabelList()
s1.dLbls.showVal = True

pie_age.add_data(data, titles_from_data=True)
pie_age.set_categories(labels)
pie_age.title = 'Распределение обучившихся по возрастным категориям'

sheet.add_chart(pie_age, f'F{max_row + 2}')

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

sheet.column_dimensions['A'].width = 50
sheet.column_dimensions['B'].width = 30
# Сохраняем файл
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
wb.save(f'{path_to_end_folder_report}/Сводный отчет {current_time}.xlsx')
