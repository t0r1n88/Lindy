import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import time
import datetime
from datetime import date

name_file_data_report = 'Общая таблица слушателей ЦОПП от 18_03_22.xlsx'
path_to_end_folder_report = 'data'

dpo_df = pd.read_excel(name_file_data_report, sheet_name='ДПО')
po_df = pd.read_excel(name_file_data_report, sheet_name='ПО')

dpo_df['for_counting'] = 1
po_df['for_counting'] = 1

dpo_df = dpo_df[dpo_df[
                    'Дополнительная_профессиональная_программа_повышение_квалификации_профессиональная_переподготовка'] == 'Повышение квалификации']

# Создаем файл excel
wb = openpyxl.Workbook()
# Создаем листы
wb.create_sheet(title='Индикаторы', index=0)
wb.create_sheet(title='Госзадание', index=1)
# Создаем цвет заливки
yellowFill = PatternFill(start_color='ffff00',
                         end_color='ffff00',
                         fill_type='solid')

# Считаем индикаторы
# Создаем базовый датафрейм
base_df = pd.DataFrame(index=['программы профессиональных модулей для среднего профессионального образования',
                              'программам для обучающихся общеобразовательных организаций',
                              'программы под заказ работодателей', 'отраслевые программы',
                              'программы для граждан предпенсионного возраста',
                              'программы по компетенциям будущего, включая компетенции цифровой экономики'])

# Отбираем нужные колонки
dpo_choice = dpo_df[['for_counting', 'Индикатор', 'Месяц_окончания_курса']]
po_choice = po_df[['for_counting', 'Индикатор', 'Месяц_окончания_курса']]

# Объединяем в единый датафрейм
indicator_df = pd.concat([dpo_choice, po_choice], ignore_index=True)

for i in range(1, 13):
    temp_df = indicator_df[indicator_df['Месяц_окончания_курса'] == i]
    temp_group = temp_df.groupby(['Индикатор']).agg({'for_counting': sum})
    base_df = pd.concat([base_df, temp_group], axis=1)

# Меняем названия колонок
base_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь',
                   'ноябрь', 'декабрь']



# Суммируем столбцы, превращаем в датафрейм и транспонируем столбец в строку
sum_month = base_df.sum(axis=0).to_frame().T
# Меняем индекс
sum_month.index = ['Всего']

base_df = pd.concat([base_df, sum_month])


# Сохраняем на лист
for r in dataframe_to_rows(base_df, index=True, header=True):
    if len(r) != 1:
        wb['Индикаторы'].append(r)

wb['Индикаторы'].column_dimensions['A'].width = 90

# Считаем госздание
# ДПО ПК (ВО)
base_vo_df = pd.DataFrame(index=['ДПО ПК(ВО)'])
vo_df = dpo_df[dpo_df['Уровень_образования_ВО_СПО'] == 'Высшее образование']
vo_df_choice = vo_df[['for_counting', 'Месяц_окончания_курса']]

for i in range(1, 13):
    temp_df = vo_df_choice[vo_df_choice['Месяц_окончания_курса'] == i]
    temp_group = temp_df.groupby(['Месяц_окончания_курса']).agg({'for_counting': sum})
    if temp_group.shape[0] == 1:
        temp_group.index = ['ДПО ПК(ВО)']
        base_vo_df = pd.concat([base_vo_df, temp_group], axis=1)
    elif temp_group.shape[0] == 0:
        empty_group = pd.DataFrame(index=['ДПО ПК(ВО)'], columns=['for_counting'], data=np.NaN)
        base_vo_df = pd.concat([base_vo_df, empty_group], axis=1)

# Меняем названия колонок
base_vo_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь',
                      'ноябрь', 'декабрь']

# Считаем СПО
base_spo_df = pd.DataFrame(index=['ДПО ПК(СПО)'])
spo_df = dpo_df[dpo_df['Уровень_образования_ВО_СПО'] == 'Среднее профессиональное образование']
spo_df_choice = spo_df[['for_counting', 'Месяц_окончания_курса']]

for i in range(1, 13):
    temp_df = spo_df_choice[spo_df_choice['Месяц_окончания_курса'] == i]
    temp_group = temp_df.groupby(['Месяц_окончания_курса']).agg({'for_counting': sum})
    if temp_group.shape[0] == 1:
        temp_group.index = ['ДПО ПК(СПО)']
        base_spo_df = pd.concat([base_spo_df, temp_group], axis=1)
    elif temp_group.shape[0] == 0:
        empty_group = pd.DataFrame(index=['ДПО ПК(СПО)'], columns=['for_counting'], data=np.NaN)
        base_spo_df = pd.concat([base_spo_df, empty_group], axis=1)



base_spo_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь',
                       'ноябрь', 'декабрь']



# Считаем ПО
base_po_df = pd.DataFrame(index=['ПО'])
po_df_choice = po_df[['for_counting', 'Месяц_окончания_курса']]

for i in range(1, 13):
    temp_df = po_df_choice[po_df_choice['Месяц_окончания_курса'] == i]
    temp_group = temp_df.groupby(['Месяц_окончания_курса']).agg({'for_counting': sum})
    if temp_group.shape[0] == 1:
        temp_group.index = ['ПО']
        base_po_df = pd.concat([base_po_df, temp_group], axis=1)
    elif temp_group.shape[0] == 0:
        empty_group = pd.DataFrame(index=['ПО'], columns=['for_counting'], data=np.NaN)
        base_po_df = pd.concat([base_po_df, empty_group], axis=1)

base_po_df.columns = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь',
                      'ноябрь', 'декабрь']



out_df = pd.concat([base_vo_df, base_spo_df, base_po_df])

# Суммируем столбцы, превращаем в датафрейм и транспонируем столбец в строку
sum_month_out_df = out_df.sum(axis=0).to_frame().T
# Меняем индекс
sum_month_out_df.index = ['Всего']

out_df = pd.concat([out_df, sum_month_out_df])



# Сохраняем на лист
for r in dataframe_to_rows(out_df, index=True, header=True):
    if len(r) != 1:
        wb['Госзадание'].append(r)

wb['Госзадание'].column_dimensions['A'].width = 50

# заливка строки

# заливаем строку
for cell in wb['Индикаторы'][8]:
    cell.fill = yellowFill

for cell in wb['Госзадание'][5]:
    cell.fill = yellowFill

# Получаем текущее время для того чтобы использовать в названии
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл
wb.save(f'{path_to_end_folder_report}/Отчет по индикаторам и госзданию {current_time}.xlsx')