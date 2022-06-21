import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time
import datetime
from datetime import date

name_file_data_report = 'Общая таблица слушателей ЦОПП 1-ПО.xlsx'
path_to_end_folder_report = 'data'

df_po = pd.read_excel(name_file_data_report,sheet_name='ПО',dtype={'Гражданство_получателя_код_страны_по_ОКСМ':str})

#Создаем шрифт которым будем выделять названия таблиц
font_name_table = Font(name='Arial Black',size=15,italic=True)


# Создаем файл excel
wb = openpyxl.Workbook()
# Создаем листы
wb.create_sheet(title='Раздел 1.3',index=0)
wb.create_sheet(title='Раздел 2.1.1',index=1)
wb.create_sheet(title='Раздел 2.1.2',index=2)
wb.create_sheet(title='Раздел 2.1.3',index=3)
wb.create_sheet(title='Раздел 2.2',index=4)
wb.create_sheet(title='Раздел 2.3',index=5)
wb.create_sheet(title='Раздел 2.4',index=6)
wb.create_sheet(title='Раздел 2.5',index=7)
wb.create_sheet(title='Раздел 2.6',index=8)
# Удаляем пустой лист
del wb['Sheet']



#1.3
df_po_1_3_base = df_po.copy()
df_po_1_3_base['for_counting'] = 1
df_po_1_3_base.fillna('Не заполнено', inplace=True)

# Считаем лист 1.3
# Считаем количество реализованных программ
# группируем. Так как нам нужны текстовые данные то применяем создаем строку с помощью join
po_quantity_program_on_type_provisional = df_po_1_3_base.groupby(
    'Наименование_программы_профессионального_обучения').agg(
    {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})
# Применяем к полученной серии функцию разделения по запятой. Предполо
po_quantity_program_on_type_provisional['Программа_профессионального_обучения_направление_подготовки'] = \
    po_quantity_program_on_type_provisional[
        'Программа_профессионального_обучения_направление_подготовки'].apply(
        lambda x: x.split(';')[0])
df_po_1_3 = po_quantity_program_on_type_provisional[
    'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

# переименовываем колонку с
df_po_1_3.rename(
    columns={'Программа_профессионального_обучения_направление_подготовки': 'Число реализованных программ'},
    inplace=True)

# Сортирируем индекс чтобы было легче добавлять столбцы
df_po_1_3.sort_index(ascending=False, inplace=True)

# Считаем количество обученных по каждой программе
df_po_1_3_quantity_students = df_po_1_3_base.groupby(
    ['Программа_профессионального_обучения_направление_подготовки']).agg({'for_counting': 'sum'})

df_po_1_3_quantity_students.sort_index(ascending=False, inplace=True)

# Соединяем 2 датафрейма
df_po_1_3['Всего обучено'] = df_po_1_3_quantity_students

# Считаем количество программ с сетевой формой
# Создаем датафрейм с теми данными о сетевой форме
df_po_1_3_network_base = df_po_1_3_base[
    df_po_1_3_base['Использование_сетевой_формы_обучения'] == 'Сетевая форма']

# Считаем количество реализованных программ
# группируем. Так как нам нужны текстовые данные то применяем создаем строку с помощью join
po_quantity_program_on_type_network = df_po_1_3_network_base.groupby(
    'Наименование_программы_профессионального_обучения').agg(
    {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})
# Применяем к полученной серии функцию разделения по запятой. Предполо
po_quantity_program_on_type_network['Программа_профессионального_обучения_направление_подготовки'] = \
    po_quantity_program_on_type_network['Программа_профессионального_обучения_направление_подготовки'].apply(
        lambda x: x.split(';')[0])
df_po_1_3_network = po_quantity_program_on_type_network[
    'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

df_po_1_3_network.rename(
    columns={'Программа_профессионального_обучения_направление_подготовки': 'Число реализованных программ'},
    inplace=True)

df_po_1_3_network.sort_index(ascending=False, inplace=True)

df_po_1_3['Число программ с сетевой формой'] = df_po_1_3_network

# Считаем число слушателей на сетевых программах
df_po_1_3['Численность слушателей сетевых программ'] = df_po_1_3_network_base.groupby(
    'Программа_профессионального_обучения_направление_подготовки').agg({'for_counting': 'sum'})

# Считаем электронное обучение
df_po_1_3_distant_ao = df_po_1_3_base[(df_po_1_3_base['Использование_ЭО'] != 'Без применения ЭО') & ( df_po_1_3_base['Использование_ЭО'] !='Не заполнено') ]

po_group_quantity_distant_ao = df_po_1_3_distant_ao.groupby(
    'Наименование_программы_профессионального_обучения').agg(
    {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})

# Применяем к полученной серии функцию разделения по запятой.
po_group_quantity_distant_ao['Программа_профессионального_обучения_направление_подготовки'] = \
    po_group_quantity_distant_ao['Программа_профессионального_обучения_направление_подготовки'].apply(
        lambda x: x.split(';')[0])
df_po_1_3['Число программ реализуемых с помощью ЭО'] = po_group_quantity_distant_ao[
    'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

# Считаем количество слушателей
df_po_1_3['Численность слушателей обученных с применением ЭО'] = df_po_1_3_distant_ao.groupby(
    'Программа_профессионального_обучения_направление_подготовки').agg({'for_counting': 'sum'})

# Считаем пользователей с исключительно ЭО
df_po_1_3_distant_only_ao = df_po_1_3_base[(df_po_1_3_base['Использование_ЭО'] == 'Исключительно с ЭО')]

df_po_1_3['Численность слушателей обученных только с ЭО '] = df_po_1_3_distant_only_ao.groupby(
    'Программа_профессионального_обучения_направление_подготовки').agg({'for_counting': 'sum'})

df_po_1_3_distant_dot = df_po_1_3_base[(df_po_1_3_base['Использование_ДОТ'] != 'Без применения ДОТ') & (df_po_1_3_base['Использование_ДОТ'] !='Не заполнено')]

po_group_quantity_distant_dot = df_po_1_3_distant_dot.groupby(
    'Наименование_программы_профессионального_обучения').agg(
    {'Программа_профессионального_обучения_направление_подготовки': lambda x: ';'.join(x)})

# Применяем к полученной серии функцию разделения по запятой.
po_group_quantity_distant_dot['Программа_профессионального_обучения_направление_подготовки'] = \
    po_group_quantity_distant_dot['Программа_профессионального_обучения_направление_подготовки'].apply(
        lambda x: x.split(';')[0])
df_po_1_3['Число программ реализуемых с помощью ДОТ'] = po_group_quantity_distant_dot[
    'Программа_профессионального_обучения_направление_подготовки'].value_counts().to_frame()

df_po_1_3['Численность слушателей обученных с применением ДОТ'] = df_po_1_3_distant_dot.groupby(
    ['Программа_профессионального_обучения_направление_подготовки']).agg({'for_counting': 'sum'})

# Считаем слушателей обученных только с ДОТ
df_po_1_3_distant_dot_only = df_po_1_3_base[df_po_1_3_base['Использование_ДОТ'] == 'Исключительно с ДОТ']

df_po_1_3['Численность слушателей обученных только с  ДОТ'] = df_po_1_3_distant_dot_only.groupby(
    ['Программа_профессионального_обучения_направление_подготовки']).agg({'for_counting': 'sum'})

df_po_1_3.index.name = 'Вид образовательных программ'

# Записываем датафрейм на лист
wb['Раздел 1.3'][f'A1'] = 'Сведения о образовательных программах,реализуемых организацией'
wb['Раздел 1.3'][f'A1'].font = font_name_table

for r in dataframe_to_rows(df_po_1_3, index=True, header=True):
    if len(r) != 1:
        wb['Раздел 1.3'].append(r)
wb['Раздел 1.3'][f'A2'] = 'Наименование образовательных программ'
wb['Раздел 1.3'].column_dimensions['A'].width = 60
wb['Раздел 1.3'].column_dimensions['B'].width = 20
wb['Раздел 1.3'].column_dimensions['D'].width = 20
wb['Раздел 1.3'].column_dimensions['E'].width = 20
wb['Раздел 1.3'].column_dimensions['F'].width = 20
wb['Раздел 1.3'].column_dimensions['G'].width = 20
wb['Раздел 1.3'].column_dimensions['H'].width = 20
wb['Раздел 1.3'].column_dimensions['I'].width = 20
wb['Раздел 1.3'].column_dimensions['J'].width = 20
wb['Раздел 1.3'].column_dimensions['K'].width = 20

wb['Раздел 1.3']['B2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['C2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['D2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['E2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['F2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['G2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['H2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['I2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['J2'].alignment = Alignment(wrap_text=True)
wb['Раздел 1.3']['K2'].alignment = Alignment(wrap_text=True)

#2.1.1
# Создаем раздел 2.1.1
df_po_2_1_1_base = df_po.copy()
df_po_2_1_1_base['for_counting'] = 1
# Отбираем слушателей обученных по программам проф подготовки по профессиям рабочих,должностям служащих
df_po_2_1_1_base = df_po_2_1_1_base[df_po_2_1_1_base[
                                        'Программа_профессионального_обучения_направление_подготовки'] == 'Программа профессиональной подготовки по профессии рабочего, должности служащего']

# Проверяем есть ли подходяшие данные
if df_po_2_1_1_base.shape[0] == 0:
    wb['Раздел 2.1.1'][
        f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
    wb['Раздел 2.1.1'][f'A2'].font = font_name_table
    wb['Раздел 2.1.1']['A2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1'][f'A3'] = 'Код профессии при наличии'
    wb['Раздел 2.1.1'].column_dimensions['A'].width = 50
    wb['Раздел 2.1.1'].column_dimensions['C'].width = 15
    wb['Раздел 2.1.1'].column_dimensions['D'].width = 15

else:
    # Считаем лист  2.3.1
    # Создаем сводную таблицу для колонок 6,7,8,9
    df_po_2_1_1_6_9 = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                     columns=['Источник_финансирования_обучения'],
                                     values=['for_counting'],
                                     aggfunc='sum')
    df_po_2_1_1_6_9.columns = df_po_2_1_1_6_9.columns.droplevel()
    # Суммируем по столбцам
    df_po_2_1_1_6_9['Всего по источникам финансирования'] = df_po_2_1_1_6_9.sum(axis=1)
    # Создаем сводную таблицу для колонок 13,14,15
    df_po_2_1_1_13_15 = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                       columns=['Форма_обучения'],
                                       values=['for_counting'],
                                       aggfunc='sum')
    df_po_2_1_1_13_15.columns = df_po_2_1_1_13_15.columns.droplevel()

    df_po_2_1_1_13_15['Всего по форме обучения'] = df_po_2_1_1_13_15.sum(axis=1)

    df_po_2_1_1 = pd.concat([df_po_2_1_1_6_9, df_po_2_1_1_13_15], axis=1)
    # Считаем по категориям финансирования
    df_po_2_1_1_cat_finance = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                             columns=[
                                                 'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                             values=['for_counting'],
                                             aggfunc='sum')
    df_po_2_1_1_cat_finance.columns = df_po_2_1_1_cat_finance.columns.droplevel()

    df_po_2_1_1 = pd.concat([df_po_2_1_1, df_po_2_1_1_cat_finance], axis=1)

    # Считаем женщин

    df_po_2_1_1_women = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                       columns=['Пол_получателя'],
                                       values=['for_counting'],
                                       aggfunc='sum')
    df_po_2_1_1_women.columns = df_po_2_1_1_women.columns.droplevel()

    df_po_2_1_1 = pd.concat([df_po_2_1_1, df_po_2_1_1_women], axis=1)

    # Считаем тех кто прошел по индвидуальным учебным планам

    df_po_2_1_1_ind_up = pd.pivot_table(df_po_2_1_1_base, index=['Код_профессии_при_наличии'],
                                        columns=[
                                            'прошли_ускоренное_обучение_по_индивидуаль-ным_учебным_планам'],
                                        values=['for_counting'],
                                        aggfunc='sum')
    df_po_2_1_1_ind_up.columns = df_po_2_1_1_ind_up.columns.droplevel()

    df_po_2_1_1 = pd.concat([df_po_2_1_1, df_po_2_1_1_ind_up], axis=1)

    # Создаем проверку
    df_po_2_1_1['temp'] = df_po_2_1_1['Всего по источникам финансирования'] == df_po_2_1_1[
        'Всего по форме обучения']

    df_po_2_1_1['Совпадение сумм столбцов 6-9 и 13-15'] = df_po_2_1_1['temp'].apply(
        lambda x: 'СОВПАДАЕТ' if x is True else 'НЕ СОВПАДАЕТ!!!')
    df_po_2_1_1.drop(columns='temp', inplace=True)
    df_po_2_1_1.index.name = 'Код профессии при наличии'

    # Записываем датафрейм на лист
    wb['Раздел 2.1.1'][
        f'A1'] = 'Распределение слушателей,обученных по программам профессиональной подготовки по професииям рабочих,должностям служащих'
    wb['Раздел 2.1.1'][f'A1'].font = font_name_table

    for r in dataframe_to_rows(df_po_2_1_1, index=True, header=True):
        if len(r) != 1:
            wb['Раздел 2.1.1'].append(r)
    wb['Раздел 2.1.1'][
        f'A2'] = 'Код професии'
    wb['Раздел 2.1.1'][f'A2'].font = font_name_table
    wb['Раздел 2.1.1']['A2'].alignment = Alignment(wrap_text=True)

    wb['Раздел 2.1.1'].column_dimensions['A'].width = 50
    wb['Раздел 2.1.1'].column_dimensions['C'].width = 15
    wb['Раздел 2.1.1'].column_dimensions['D'].width = 15

    wb['Раздел 2.1.1']['B2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['C2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['D2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['E2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['F2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['G2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['H2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['I2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['J2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['K2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['L2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['M2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.1']['Q2'].alignment = Alignment(wrap_text=True)


# Считаем раздел 2.1.2
df_po_2_1_2_base = df_po.copy()
df_po_2_1_2_base['for_counting'] = 1
# Отбираем слушателей обученных по программам переподготовки рабочих, служащих
df_po_2_1_2_base = df_po_2_1_2_base[df_po_2_1_2_base[
                                        'Программа_профессионального_обучения_направление_подготовки'] == 'Программа переподготовки рабочих, служащих']

# Проверяем размер датафрейма
# Если он пустой то ничего не считаем а просто записываем в лист пустые строки
if df_po_2_1_2_base.shape[0] == 0:
    wb['Раздел 2.1.2'][
        f'A1'] = 'Распределение слушателей,обученных по программам переподготовки рабочих,служащих'
    wb['Раздел 2.1.2'][f'A1'].font = font_name_table
    wb['Раздел 2.1.2'][
        f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
    wb['Раздел 2.1.2'][f'A2'].font = font_name_table
    wb['Раздел 2.1.2']['A2'].alignment = Alignment(wrap_text=True)

    wb['Раздел 2.1.2'][f'A3'] = 'Код профессии при наличии'

    wb['Раздел 2.1.2'].column_dimensions['A'].width = 50
    wb['Раздел 2.1.2'].column_dimensions['C'].width = 15
    wb['Раздел 2.1.2'].column_dimensions['D'].width = 15
else:
    # Создаем сводную таблицу для колонок 6,7,8,9
    df_po_2_1_2_6_9 = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                     columns=['Источник_финансирования_обучения'],
                                     values=['for_counting'],
                                     aggfunc='sum')
    df_po_2_1_2_6_9.columns = df_po_2_1_2_6_9.columns.droplevel()

    df_po_2_1_2_6_9['Всего по источникам финансирования'] = df_po_2_1_2_6_9.sum(axis=1)

    # Создаем сводную таблицу для колонок 13,14,15
    df_po_2_1_2_13_15 = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                       columns=['Форма_обучения'],
                                       values=['for_counting'],
                                       aggfunc='sum')
    df_po_2_1_2_13_15.columns = df_po_2_1_2_13_15.columns.droplevel()

    df_po_2_1_2_13_15['Всего по форме обучения'] = df_po_2_1_2_13_15.sum(axis=1)

    df_po_2_1_2 = pd.concat([df_po_2_1_2_6_9, df_po_2_1_2_13_15], axis=1)

    # Считаем по категориям финансирования
    df_po_2_1_2_cat_finance = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                             columns=[
                                                 'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                             values=['for_counting'],
                                             aggfunc='sum')
    df_po_2_1_2_cat_finance.columns = df_po_2_1_2_cat_finance.columns.droplevel()

    df_po_2_1_2 = pd.concat([df_po_2_1_2, df_po_2_1_2_cat_finance], axis=1)

    # Считаем женщин

    df_po_2_1_2_women = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                       columns=['Пол_получателя'],
                                       values=['for_counting'],
                                       aggfunc='sum')
    df_po_2_1_2_women.columns = df_po_2_1_2_women.columns.droplevel()

    df_po_2_1_2 = pd.concat([df_po_2_1_2, df_po_2_1_2_women], axis=1)

    # Считаем тех кто прошел по индвидуальным учебным планам

    df_po_2_1_2_ind_up = pd.pivot_table(df_po_2_1_2_base, index=['Код_профессии_при_наличии'],
                                        columns=[
                                            'прошли_ускоренное_обучение_по_индивидуаль-ным_учебным_планам'],
                                        values=['for_counting'],
                                        aggfunc='sum')
    df_po_2_1_2_ind_up.columns = df_po_2_1_2_ind_up.columns.droplevel()

    df_po_2_1_2 = pd.concat([df_po_2_1_2, df_po_2_1_2_ind_up], axis=1)

    # Создаем проверку
    df_po_2_1_2['temp'] = df_po_2_1_2['Всего по источникам финансирования'] == df_po_2_1_2[
        'Всего по форме обучения']

    df_po_2_1_2['Совпадение сумм столбцов 6-9 и 13-15'] = df_po_2_1_2['temp'].apply(
        lambda x: 'СОВПАДАЕТ' if x is True else 'НЕ СОВПАДАЕТ!!!')
    df_po_2_1_2.drop(columns='temp', inplace=True)
    df_po_2_1_2.index.name = 'Код профессии при наличии'

    # Записываем датафрейм на лист
    wb['Раздел 2.1.2'][
        f'A1'] = 'Распределение слушателей,обученных по программам переподготовки рабочих,служащих'
    wb['Раздел 2.1.2'][f'A1'].font = font_name_table

    for r in dataframe_to_rows(df_po_2_1_2, index=True, header=True):
        if len(r) != 1:
            wb['Раздел 2.1.2'].append(r)
    wb['Раздел 2.1.2'][
        f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
    wb['Раздел 2.1.2'][f'A2'].font = font_name_table
    wb['Раздел 2.1.2']['A2'].alignment = Alignment(wrap_text=True)

    wb['Раздел 2.1.2'][f'A3'] = 'Код профессии при наличии'

    wb['Раздел 2.1.2'].column_dimensions['A'].width = 50
    wb['Раздел 2.1.2'].column_dimensions['C'].width = 15
    wb['Раздел 2.1.2'].column_dimensions['D'].width = 15
    wb['Раздел 2.1.2']['B2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['C2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['D2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['E2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['F2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['G2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['H2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['I2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['J2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['K2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['L2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['M2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.2']['Q2'].alignment = Alignment(wrap_text=True)

# Считаем лист 2.1.3
# Считаем раздел 2.1.2
df_po_2_1_3_base = df_po.copy()
df_po_2_1_3_base['for_counting'] = 1

# Отбираем слушателей обученных по программам переподготовки рабочих, служащих
df_po_2_1_3_base = df_po_2_1_3_base[df_po_2_1_3_base[
                                        'Программа_профессионального_обучения_направление_подготовки'] == 'Программа повышения квалификации рабочих, служащих']

if df_po_2_1_3_base.shape[0] == 0:
    wb['Раздел 2.1.3'][
        f'A1'] = 'Распределение слушателей,обученных по программам повышения квалификации рабочих,служащих'
    wb['Раздел 2.1.3'][f'A1'].font = font_name_table
    wb['Раздел 2.1.3'][
        f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
    wb['Раздел 2.1.3'][f'A2'].font = font_name_table
    wb['Раздел 2.1.3']['A2'].alignment = Alignment(wrap_text=True)

    wb['Раздел 2.1.3'][f'A3'] = 'Код профессии при наличии'
    wb['Раздел 2.1.3'].column_dimensions['A'].width = 50
    wb['Раздел 2.1.3'].column_dimensions['C'].width = 15
    wb['Раздел 2.1.3'].column_dimensions['D'].width = 15
else:

    # Создаем сводную таблицу для колонок 6,7,8,9
    df_po_2_1_3_6_9 = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                     columns=['Источник_финансирования_обучения'],
                                     values=['for_counting'],
                                     aggfunc='sum')
    df_po_2_1_3_6_9.columns = df_po_2_1_3_6_9.columns.droplevel()

    df_po_2_1_3_6_9['Всего по источникам финансирования'] = df_po_2_1_3_6_9.sum(axis=1)

    # Создаем сводную таблицу для колонок 13,14,15
    df_po_2_1_3_13_15 = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                       columns=['Форма_обучения'],
                                       values=['for_counting'],
                                       aggfunc='sum')
    df_po_2_1_3_13_15.columns = df_po_2_1_3_13_15.columns.droplevel()

    df_po_2_1_3_13_15['Всего по форме обучения'] = df_po_2_1_3_13_15.sum(axis=1)

    df_po_2_1_3 = pd.concat([df_po_2_1_3_6_9, df_po_2_1_3_13_15], axis=1)

    # Считаем по категориям финансирования
    df_po_2_1_3_cat_finance = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                             columns=[
                                                 'Источник_финансирования_индикаторы_физ_лицо_юр_лицо_бюдж_ассигнования_собственные_средства_ЦОПП'],
                                             values=['for_counting'],
                                             aggfunc='sum')
    df_po_2_1_3_cat_finance.columns = df_po_2_1_3_cat_finance.columns.droplevel()

    df_po_2_1_3 = pd.concat([df_po_2_1_3, df_po_2_1_3_cat_finance], axis=1)

    # Считаем женщин

    df_po_2_1_3_women = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                       columns=['Пол_получателя'],
                                       values=['for_counting'],
                                       aggfunc='sum')
    df_po_2_1_3_women.columns = df_po_2_1_3_women.columns.droplevel()

    df_po_2_1_3 = pd.concat([df_po_2_1_3, df_po_2_1_3_women], axis=1)

    # Считаем тех кто прошел по индвидуальным учебным планам

    df_po_2_1_3_ind_up = pd.pivot_table(df_po_2_1_3_base, index=['Код_профессии_при_наличии'],
                                        columns=[
                                            'прошли_ускоренное_обучение_по_индивидуаль-ным_учебным_планам'],
                                        values=['for_counting'],
                                        aggfunc='sum')
    df_po_2_1_3_ind_up.columns = df_po_2_1_3_ind_up.columns.droplevel()

    df_po_2_1_3 = pd.concat([df_po_2_1_3, df_po_2_1_3_ind_up], axis=1)

    # Создаем проверку
    df_po_2_1_3['temp'] = df_po_2_1_3['Всего по источникам финансирования'] == df_po_2_1_3[
        'Всего по форме обучения']

    df_po_2_1_3['Совпадение сумм столбцов 6-9 и 13-15'] = df_po_2_1_3['temp'].apply(
        lambda x: 'СОВПАДАЕТ' if x is True else 'НЕ СОВПАДАЕТ!!!')
    df_po_2_1_3.drop(columns='temp', inplace=True)
    df_po_2_1_3.index.name = 'Код профессии при наличии'

    # Записываем датафрейм на лист
    wb['Раздел 2.1.3'][
        f'A1'] = 'Распределение слушателей,обученных по программам повышения квалификации рабочих,служащих'
    wb['Раздел 2.1.3'][f'A1'].font = font_name_table

    for r in dataframe_to_rows(df_po_2_1_3, index=True, header=True):
        if len(r) != 1:
            wb['Раздел 2.1.3'].append(r)
    wb['Раздел 2.1.3'][
        f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
    wb['Раздел 2.1.3'][f'A2'].font = font_name_table
    wb['Раздел 2.1.3']['A2'].alignment = Alignment(wrap_text=True)

    wb['Раздел 2.1.3'][f'A3'] = 'Код профессии при наличии'
    wb['Раздел 2.1.3'].column_dimensions['A'].width = 50
    wb['Раздел 2.1.3'].column_dimensions['C'].width = 15
    wb['Раздел 2.1.3'].column_dimensions['D'].width = 15
    wb['Раздел 2.1.3']['B2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['C2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['D2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['E2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['F2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['G2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['H2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['I2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['J2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['K2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['L2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['M2'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.1.3']['Q2'].alignment = Alignment(wrap_text=True)

# Считаем лист 2.2
df_po_2_2_base = df_po.copy()
df_po_2_2_base['for_counting'] = 1

# Создаем сводную таблицу
df_po_2_2 = pd.pivot_table(df_po_2_2_base, index=['Категория_слушателя'],
                           columns=['Программа_профессионального_обучения_направление_подготовки'],
                           values=['for_counting'],
                           aggfunc='sum')

df_po_2_2.columns = df_po_2_2.columns.droplevel()

df_po_2_2['Всего обучено по программам'] = df_po_2_2.sum(axis=1)

# Записываем на лист
# Записываем датафрейм на лист
wb['Раздел 2.2'][f'A1'] = 'Обучение отдельных категорий слушателей'
wb['Раздел 2.2'][f'A1'].font = font_name_table
wb['Раздел 2.2'][
    f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
wb['Раздел 2.2'][f'A2'].font = font_name_table
wb['Раздел 2.2']['A2'].alignment = Alignment(wrap_text=True)

for r in dataframe_to_rows(df_po_2_2, index=True, header=True):
    if len(r) != 1:
        wb['Раздел 2.2'].append(r)
wb['Раздел 2.2'][f'A3'] = 'Наименование показателей'
wb['Раздел 2.2'].column_dimensions['A'].width = 80
wb['Раздел 2.2'].column_dimensions['B'].width = 30
wb['Раздел 2.2']['B3'].alignment = Alignment(wrap_text=True)
wb['Раздел 2.2'].column_dimensions['C'].width = 30
wb['Раздел 2.2']['C3'].alignment = Alignment(wrap_text=True)
wb['Раздел 2.2'].column_dimensions['D'].width = 30
wb['Раздел 2.2']['D3'].alignment = Alignment(wrap_text=True)
wb['Раздел 2.2'].column_dimensions['D'].width = 30
wb['Раздел 2.2']['E3'].alignment = Alignment(wrap_text=True)

# Считаем раздел 2.3 Инвалиды
df_po_2_3_base = df_po.copy()
df_po_2_3_base['for_counting'] = 1
df_po_2_3_base.fillna('Не заполнено', inplace=True)
# Отбираем нездоровых
df_po_2_3_base = df_po_2_3_base[
    (df_po_2_3_base['Сведения_об_ограничении_возможностей_здоровья'] != 'нет ОВЗ') & (
            df_po_2_3_base['Сведения_об_ограничении_возможностей_здоровья'] != 'Не заполнено')]

if df_po_2_3_base.shape[0] == 0:
    wb['Раздел 2.3'].column_dimensions['A'].width = 80
    wb['Раздел 2.3'][f'A1'] = 'Обучение лиц с ограниченными возможностями здоровья и инвалидов'
    wb['Раздел 2.3'][f'A1'].font = font_name_table
    wb['Раздел 2.3'][
        f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
    wb['Раздел 2.3'][f'A2'].font = font_name_table
    wb['Раздел 2.3']['A2'].alignment = Alignment(wrap_text=True)
else:
    # создаем сводную таблицу
    df_po_2_3 = pd.pivot_table(df_po_2_3_base, index=['Сведения_об_ограничении_возможностей_здоровья'],
                               columns=['Программа_профессионального_обучения_направление_подготовки'],
                               values=['for_counting'],
                               aggfunc='sum')
    df_po_2_3.columns = df_po_2_3.columns.droplevel()
    df_po_2_3['Всего обучено'] = df_po_2_3.sum(axis=1)
    # Записываем в лист
    wb['Раздел 2.3'][f'A1'] = 'Обучение лиц с ограниченными возможностями здоровья и инвалидов'
    wb['Раздел 2.3'][f'A1'].font = font_name_table
    wb['Раздел 2.3'][
        f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
    wb['Раздел 2.3'][f'A2'].font = font_name_table
    wb['Раздел 2.3']['A2'].alignment = Alignment(wrap_text=True)
    for r in dataframe_to_rows(df_po_2_3, index=True, header=True):
        if len(r) != 1:
            wb['Раздел 2.3'].append(r)

    wb['Раздел 2.3'][f'A3'] = 'Наименование показателей'
    wb['Раздел 2.3'].column_dimensions['A'].width = 80
    wb['Раздел 2.3'].column_dimensions['B'].width = 30
    wb['Раздел 2.3']['B3'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.3'].column_dimensions['C'].width = 30
    wb['Раздел 2.3']['C3'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.3'].column_dimensions['D'].width = 30
    wb['Раздел 2.3']['D3'].alignment = Alignment(wrap_text=True)
    wb['Раздел 2.3'].column_dimensions['D'].width = 30
    wb['Раздел 2.3']['E3'].alignment = Alignment(wrap_text=True)

# Считаем лист 2.4
df_po_2_4 = df_po.copy()
df_po_2_4['for_counting'] = 1

# Считаем общее количество
df_po_2_4_all = pd.pivot_table(df_po_2_4,
                               columns=['Программа_профессионального_обучения_направление_подготовки'],
                               values=['for_counting'],
                               aggfunc='sum')
df_po_2_4_all.index = ['Всего обучено']

# Записываем в лист
wb['Раздел 2.4'][f'A1'] = 'Распределение слушателей,обученных по программам ПО,по уровню образования'
wb['Раздел 2.4'][f'A1'].font = font_name_table
wb['Раздел 2.4'][
    f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!!'
wb['Раздел 2.4'][f'A2'].font = font_name_table
wb['Раздел 2.4']['A2'].alignment = Alignment(wrap_text=True)

# Считаем по основному общему образованию
df_po_2_4_02 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'основное общее (9 классов)']

# считаем сумму
df_po_2_4_02_all = pd.pivot_table(df_po_2_4_02,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_02_all.index = ['Основное общее']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_02_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_02_current_year = df_po_2_4_02.copy()

df_po_2_4_02_current_year['Образование в текущем году'] = df_po_2_4_02_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_02_current_year = df_po_2_4_02_current_year[
    df_po_2_4_02_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_02_current_year_finsih = pd.pivot_table(df_po_2_4_02_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_02_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_02_current_year_finsih])

# Считаем по среднему общему образованию
df_po_2_4_04 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'среднее общее (11 классов)']

# считаем сумму
df_po_2_4_04_all = pd.pivot_table(df_po_2_4_04,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_04_all.index = ['Среднее общее']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_04_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_04_current_year = df_po_2_4_04.copy()

df_po_2_4_04_current_year['Образование в текущем году'] = df_po_2_4_04_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_04_current_year = df_po_2_4_04_current_year[
    df_po_2_4_04_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_04_current_year_finsih = pd.pivot_table(df_po_2_4_04_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_04_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_04_current_year_finsih])

# Считаем по среднему общему образованию
df_po_2_4_06 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'СПО/НПО']

# считаем сумму
df_po_2_4_06_all = pd.pivot_table(df_po_2_4_06,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_06_all.index = ['Среднее профессиональное']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_06_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_06_current_year = df_po_2_4_06.copy()

df_po_2_4_06_current_year['Образование в текущем году'] = df_po_2_4_06_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_06_current_year = df_po_2_4_06_current_year[
    df_po_2_4_06_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_06_current_year_finsih = pd.pivot_table(df_po_2_4_06_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_06_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_06_current_year_finsih])

# Считаем по неполному высшему общему образованию
df_po_2_4_10 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'неполное высшее образование']

# считаем сумму
df_po_2_4_10_all = pd.pivot_table(df_po_2_4_10,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_10_all.index = ['Неполное высшее образование']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_10_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_10_current_year = df_po_2_4_10.copy()

df_po_2_4_10_current_year['Образование в текущем году'] = df_po_2_4_10_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_10_current_year = df_po_2_4_10_current_year[
    df_po_2_4_10_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_10_current_year_finsih = pd.pivot_table(df_po_2_4_10_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_10_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_10_current_year_finsih])

# Считаем по  высшему Бакалавры образованию
df_po_2_4_11 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'ВО (бакалавр)']

# считаем сумму
df_po_2_4_11_all = pd.pivot_table(df_po_2_4_11,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_11_all.index = ['ВО (бакалавр)']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_11_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_11_current_year = df_po_2_4_11.copy()

df_po_2_4_11_current_year['Образование в текущем году'] = df_po_2_4_11_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_11_current_year = df_po_2_4_11_current_year[
    df_po_2_4_11_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_11_current_year_finsih = pd.pivot_table(df_po_2_4_11_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_11_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_11_current_year_finsih])

# Считаем по  высшему специалист образованию
df_po_2_4_13 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'ВО (специалист)']

# считаем сумму
df_po_2_4_13_all = pd.pivot_table(df_po_2_4_13,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_13_all.index = ['ВО (специалист)']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_13_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_13_current_year = df_po_2_4_13.copy()

df_po_2_4_13_current_year['Образование в текущем году'] = df_po_2_4_13_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_13_current_year = df_po_2_4_13_current_year[
    df_po_2_4_13_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_13_current_year_finsih = pd.pivot_table(df_po_2_4_13_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_13_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_13_current_year_finsih])

# Считаем по  высшему магистр образованию
df_po_2_4_15 = df_po_2_4[df_po_2_4['Уровень_образования'] == 'ВО (магистр)']

# считаем сумму
df_po_2_4_15_all = pd.pivot_table(df_po_2_4_15,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_15_all.index = ['ВО (магистр)']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_15_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_15_current_year = df_po_2_4_15.copy()

df_po_2_4_15_current_year['Образование в текущем году'] = df_po_2_4_15_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_15_current_year = df_po_2_4_15_current_year[
    df_po_2_4_15_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_15_current_year_finsih = pd.pivot_table(df_po_2_4_15_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_15_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_15_current_year_finsih])

# Считаем по  высшему магистр образованию
df_po_2_4_17 = df_po_2_4[
    df_po_2_4['Уровень_образования'] == 'без образования (не имеют основного общего образования)']

# считаем сумму
df_po_2_4_17_all = pd.pivot_table(df_po_2_4_17,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_17_all.index = ['без образования (не имеют основного общего образования)']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_17_all], axis=0)

# Считаем получивших образование в отчетном году
df_po_2_4_17_current_year = df_po_2_4_17.copy()

df_po_2_4_17_current_year['Образование в текущем году'] = df_po_2_4_17_current_year[
    'Год_получения_образования'].apply(
    lambda x: 'Да' if x == date.today().year else 'Нет')

df_po_2_4_17_current_year = df_po_2_4_17_current_year[
    df_po_2_4_17_current_year['Образование в текущем году'] == 'Да']

df_po_2_4_17_current_year_finsih = pd.pivot_table(df_po_2_4_17_current_year, columns=[
    'Программа_профессионального_обучения_направление_подготовки'],
                                                  values=['for_counting'],
                                                  aggfunc='sum')
df_po_2_4_17_current_year_finsih.index = ['из них получили указанное образование в отчетном году']

df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_17_current_year_finsih])

# Считаем инвалидов из людей без образования
df_po_2_4_20 = df_po_2_4_17[df_po_2_4_17['Сведения_об_ограничении_возможностей_здоровья'] != 'нет ОВЗ']

# считаем сумму
df_po_2_4_20_all = pd.pivot_table(df_po_2_4_20,
                                  columns=['Программа_профессионального_обучения_направление_подготовки'],
                                  values=['for_counting'],
                                  aggfunc='sum')
df_po_2_4_20_all.index = [
    'лица с ограниченными возможностиями здоровья(инвалиды,дети-сироты,лицо с ОВЗ) не имеющие ОО']

# Добавляем к основному датафрейму
df_po_2_4_all = pd.concat([df_po_2_4_all, df_po_2_4_20_all], axis=0)

for r in dataframe_to_rows(df_po_2_4_all, index=True, header=True):
    if len(r) != 1:
        wb['Раздел 2.4'].append(r)
# Устанавливаем ширину колонок
wb['Раздел 2.4'].column_dimensions['A'].width = 80
wb['Раздел 2.4'].column_dimensions['B'].width = 30
wb['Раздел 2.4']['B3'].alignment = Alignment(wrap_text=True)
wb['Раздел 2.4'].column_dimensions['C'].width = 30
wb['Раздел 2.4']['C3'].alignment = Alignment(wrap_text=True)
wb['Раздел 2.4'].column_dimensions['D'].width = 30
wb['Раздел 2.4']['D3'].alignment = Alignment(wrap_text=True)

# Считаем лист 2.5
# Создаем дополнительную числовую колонку где каждое значение это 1, для удобства агрегирования
df_po_2_5 = df_po.copy()
# Добавляем колонку с 1
df_po_2_5['for_counting'] = 1

# Считаем строку 01 Всего
df_2_5_all = pd.pivot_table(df_po_2_5,
                            columns=['Программа_профессионального_обучения_направление_подготовки',
                                     'Источник_финансирования_обучения'],
                            values=['for_counting'],
                            aggfunc='sum')
df_2_5_all.index = ['Всего']
df_2_5_all.index.name = 'Код государства по ОКСМ'

# Считаем данные по странам
svod_df_po_2_5 = pd.pivot_table(df_po_2_5,
                                index=['Гражданство_получателя_код_страны_по_ОКСМ'],
                                columns=['Программа_профессионального_обучения_направление_подготовки',
                                         'Источник_финансирования_обучения'],
                                values=['for_counting'],
                                aggfunc='sum')

# Удаляем лишний мультииндекс
svod_df_po_2_5.columns = svod_df_po_2_5.columns.droplevel()
# заполняем нулями для корректного суммирования
svod_df_po_2_5.fillna(0.0, inplace=True)

# Соединяем датафреймы
df_po_2_5_out = pd.concat([df_2_5_all, svod_df_po_2_5])
# заменяем нули на нан, чтобы в итоговой таблице нули не отвлекали
df_po_2_5_out.replace(0.0, np.NaN, inplace=True)

wb['Раздел 2.5'][f'A1'] = 'Распределение слушателей, обученных по программам ПО, по гражданству'
wb['Раздел 2.5'][f'A1'].font = font_name_table
wb['Раздел 2.5'][
    f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!'
wb['Раздел 2.5'][f'A2'].font = font_name_table

for r in dataframe_to_rows(df_po_2_5_out, index=True, header=True):
    if len(r) != 1:
        wb['Раздел 2.5'].append(r)
wb['Раздел 2.5'].column_dimensions['A'].width = 50
wb['Раздел 2.5'].column_dimensions['B'].width = 50
wb['Раздел 2.5'].column_dimensions['F'].width = 50

# Создаем раздел 2.6
df_po_2_6 = df_po.copy()
df_po_2_6['for_counting'] = 1

# Считаем в общем сколько обучено
df_po_2_6_all = pd.pivot_table(df_po_2_6,
                               columns=['Программа_профессионального_обучения_направление_подготовки',
                                        'Пол_получателя'],
                               values=['for_counting'],
                               aggfunc='sum')
df_po_2_6_all.index = ['Всего обучено']

# Считаем распределение по возрастам
df_po_2_6_by_age = pd.pivot_table(df_po_2_6, index=['Возрастная_категория_1ПО'],
                                  columns=['Программа_профессионального_обучения_направление_подготовки',
                                           'Пол_получателя'],
                                  values=['for_counting'],
                                  aggfunc='sum')

# Удаляем лишний мультииндекс
df_po_2_6_by_age.columns = df_po_2_6_by_age.columns.droplevel()

# Соединяем 2 датафрейма
df_po_2_6_out = pd.concat([df_po_2_6_all, df_po_2_6_by_age])

wb['Раздел 2.6'][f'A1'] = 'Распределение слушателей, обученных по программам ПО, по возрасту и полу'
wb['Раздел 2.6'][f'A1'].font = font_name_table
wb['Раздел 2.6'][
    f'A2'] = 'В таблице отображаются ТОЛЬКО ТЕ показатели которые присутствуют в ИСХОДНОЙ таблице!!!'
wb['Раздел 2.6'][f'A2'].font = font_name_table
wb['Раздел 2.6'][f'A3'] = 'Чтобы вычислить показатель ВСЕГО ОБУЧЕНО сложите значения в колонке Муж+Жен'
wb['Раздел 2.6'][f'A3'].font = font_name_table

for r in dataframe_to_rows(df_po_2_6_out, index=True, header=True):
    if len(r) != 1:
        wb['Раздел 2.6'].append(r)
wb['Раздел 2.6'].column_dimensions['A'].width = 50
wb['Раздел 2.6'].column_dimensions['B'].width = 50
wb['Раздел 2.6']['B4'].alignment = Alignment(wrap_text=True)
wb['Раздел 2.6'].column_dimensions['D'].width = 50
wb['Раздел 2.6']['D4'].alignment = Alignment(wrap_text=True)
wb['Раздел 2.6'].column_dimensions['F'].width = 50
wb['Раздел 2.6']['F4'].alignment = Alignment(wrap_text=True)

# Получаем текущее время для того чтобы использовать в названии
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл
wb.save(f'{path_to_end_folder_report}/Часть отчета 1-ПО.xlsx {current_time}.xlsx')


