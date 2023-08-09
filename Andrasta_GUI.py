import tkinter
import numpy as np
import sys
import pandas as pd
import os
from docxtpl import DocxTemplate
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
import time
import datetime
pd.options.mode.chained_assignment = None  # default='warn'
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class NotTotal(BaseException):
    pass

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
def select_file_template_doc():
    """
    Функция для выбора файла шаблона
    :return: Путь к файлу шаблона
    """
    global name_file_template_doc
    name_file_template_doc = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))


def select_file_data_obraz_pk():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global name_file_data_obraz_program_pk
    # Получаем путь к файлу
    name_file_data_obraz_program_pk = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_educ_obraz_pk():
    """
    Функция для выбора папки куда будут генерироваться файлы
    :return:
    """
    global path_to_end_folder_obraz_program_pk
    path_to_end_folder_obraz_program_pk = filedialog.askdirectory()

def select_file_template_educ_program_pk():
    """
    Функция для выбора файла шаблона
    :return: Путь к файлу шаблона
    """
    global name_file_template_educ_program_pk
    name_file_template_educ_program_pk = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))

# Вспомогательные для ПО
def select_file_data_obraz_po():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global name_file_data_obraz_program_po
    # Получаем путь к файлу
    name_file_data_obraz_program_po = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_end_folder_educ_obraz_po():
    """
    Функция для выбора папки куда будут генерироваться файлы
    :return:
    """
    global path_to_end_folder_obraz_program_po
    path_to_end_folder_obraz_program_po = filedialog.askdirectory()

def select_file_template_educ_program_po():
    """
    Функция для выбора файла шаблона
    :return: Путь к файлу шаблона
    """
    global name_file_template_educ_program_po
    name_file_template_educ_program_po = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))

def convert_to_int(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return 0
    if cell.isdigit():
        return int(cell)
    else:
        return 0

def convert_date(cell):
    """
    Функция для конвертации даты в формате 1957-05-10 в формат 10.05.1957(строковый)
    """

    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except TypeError:
        print(cell)
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()
    except ValueError:
        pass
        # print(cell)
        # # messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', 'Пустая ячейка с датой или некорректная запись!!!')
        # # quit()

def create_educ_program_pk():
    """
    Функция для генерации образовательных программ
    """
    try:
        # Открываем таблицу
        wb = openpyxl.load_workbook(name_file_data_obraz_program_pk)

        name_sheet_up = wb.sheetnames[0]  # получаем название листа с учебным планом
        name_sheet_data = wb.sheetnames[1]  # получаем название листа с данными программы

        """
        1) Ищем на какой строке находится ИТОГО
        2) Мы знаем что там должно быть 7 колонок
        """
        target_value = 'ИТОГО'

        # Поиск значения в выбранном столбце
        column_number = 1  # Номер столбца, в котором ищем значение (например, столбец A)
        target_row = None  # Номер строки с искомым значением

        for row in wb[name_sheet_up].iter_rows(min_row=1, min_col=column_number, max_col=column_number):
            cell_value = row[0].value
            if cell_value == target_value:
                target_row = row[0].row
                break

        if not target_row:
            # если не находим слово ИТОГО то выдаем исключение
            raise NotTotal
        # если значение найдено то считываем нужное количество строк и  7 колонок
        df_up = pd.read_excel(name_file_data_obraz_program_pk, sheet_name=name_sheet_up, nrows=target_row,
                              usecols='A:F', dtype=str)

        df_up.iloc[:, 1:5] = df_up.iloc[:, 1:5].applymap(convert_to_int)  # 1) Приводим к инту колонки 2-5

        # Заполняем возможные пустые строки
        df_up['Наименование_раздела'] = df_up['Наименование_раздела'].fillna('Не заполнено название раздела')
        # Очищаем от возможнных пробелов
        df_up['Наименование_раздела'] = df_up['Наименование_раздела'].apply(lambda x: x.strip())

        df_up['Форма_промежуточного_итогового_контроля'] = df_up['Форма_промежуточного_итогового_контроля'].fillna('')


        # Создаем датафрейм учебной программы без учета строки ИТОГО для таблиц краткой аннотации
        short_df_up = df_up[df_up['Наименование_раздела'] != 'ИТОГО']
        short_df_up = short_df_up[~short_df_up['Наименование_раздела'].str.contains('Итоговая аттестация')]

        # получаем единичные значения из листа с данными
        single_row_df = pd.read_excel(name_file_data_obraz_program_pk, sheet_name=name_sheet_data, nrows=1,
                                      usecols='A:N')
        single_row_df.iloc[:, 6] = single_row_df.iloc[:, 6].apply(convert_date)  # обрабатываем колонку с датой

        # Очищаем от лишнего поля которые заполняет пользователь
        # Заполняем возможные пустые строки
        single_row_df['Наименование_программы'] = single_row_df['Наименование_программы'].fillna('Не заполнено !!!')
        # Очищаем от возможнных пробелов
        single_row_df['Наименование_программы'] = single_row_df['Наименование_программы'].apply(lambda x: x.strip())

        # single_row_df['Профессиональный_стандарт'] = single_row_df['Профессиональный_стандарт'].fillna(
        #     'Не заполнено !!!')
        # # Очищаем от возможнных пробелов
        # single_row_df['Профессиональный_стандарт'] = single_row_df['Профессиональный_стандарт'].apply(
        #     lambda x: x.strip())

        # получаем датафрейм с технологиями обучения
        tech_df = pd.read_excel(name_file_data_obraz_program_pk, sheet_name=name_sheet_data, usecols='O:Q')

        tech_df.dropna(thresh=2, inplace=True)  # очищаем от строк в которых не заполнены 2 колонки

        tech_df['Разработчики_программы'] = tech_df['Разработчики_программы'].fillna('Не заполнено')
        # Очищаем от возможнных пробелов
        tech_df['Характеристика_технологии_обучения'] = tech_df['Характеристика_технологии_обучения'].apply(
            lambda x: x.strip())
        tech_df['Технологии_обучения'] = tech_df['Технологии_обучения'].apply(lambda x: x.strip())
        tech_df['Разработчики_программы'] = tech_df['Разработчики_программы'].apply(lambda x: x.strip())

        #

        # создаем список технологий
        educ_lst = tech_df['Технологии_обучения'].tolist()

        # Конвертируем датафрейм с описанием программы в список словарей
        data_program = single_row_df.to_dict('records')

        context = data_program[0]
        # текстовые составные переменные
        context['Технологии_обучения'] = ';\n'.join(educ_lst)

        # Добавляем датафреймы
        context['lst_tech'] = tech_df.to_dict('records')  # образовательные технологии
        context['up_lst'] = df_up.to_dict('records')  # учебный план
        context['short_up_lst'] = short_df_up.to_dict('records')  # учебный план

        lst_dev = [value for value in tech_df['Разработчики_программы'].tolist() if value != 'Не заполнено']
        context['lst_dev'] = lst_dev

        doc = DocxTemplate(name_file_template_educ_program_pk)
        # Создаем документ
        doc.render(context)
        # сохраняем документ
        # название программы
        name_pk = single_row_df['Наименование_программы'].tolist()[0]
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        doc.save(
            f'{path_to_end_folder_obraz_program_pk}/Программа ПК {name_pk} {current_time}.docx')





    except IndexError:
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', 'Заполните полностью строку 2 на листе 1.По программе!!!')
    except NameError:
        messagebox.showinfo('Андраста ver 1.85 Создание программ ПК и ПО', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    except FileNotFoundError:
        # сообщение на случай если путь до папки куда сохраняется файл слишком длинный
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', f'Слишком длинный путь до сохраняемого файла!\nВыберите другую папку')
    except KeyError as e:
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', f'Не найдено название колонки {e.args}')
    else:
        messagebox.showinfo('Андраста ver 1.85 Создание программ ПК и ПО', 'Создание образовательной программы\nЗавершено!')

def create_educ_program_po():
    """
    Функция для генерации программ ПО
    """
    try:
        # Открываем файл
        wb = openpyxl.load_workbook(name_file_data_obraz_program_po)

        name_sheet_up = wb.sheetnames[0]  # получаем название листа с учебным планом
        name_sheet_data = wb.sheetnames[1]  # получаем название листа с данными программы

        """
        1) Ищем на какой строке находится ИТОГО
        2) Мы знаем что там должно быть 7 колонок
        """
        target_value = 'ИТОГО'

        # Поиск значения в выбранном столбце
        column_number = 1  # Номер столбца, в котором ищем значение (например, столбец A)
        target_row = None  # Номер строки с искомым значением

        for row in wb[name_sheet_up].iter_rows(min_row=1, min_col=column_number, max_col=column_number):
            cell_value = row[0].value
            if cell_value == target_value:
                target_row = row[0].row
                break

        if not target_row:
            # если не находим слово ИТОГО то выдаем исключение
            raise NotTotal
        # если значение найдено то считываем нужное количество строк и  7 колонок
        df_up = pd.read_excel(name_file_data_obraz_program_po, sheet_name=name_sheet_up, nrows=target_row,
                              usecols='A:G', dtype=str)

        df_up.iloc[:, 1:6] = df_up.iloc[:, 1:6].applymap(convert_to_int)  # 1) Приводим к инту колонки 2-6

        # Заполняем возможные пустые строки
        df_up['Наименование_раздела'] = df_up['Наименование_раздела'].fillna('Не заполнено название раздела')
        # Очищаем от возможнных пробелов
        df_up['Наименование_раздела'] = df_up['Наименование_раздела'].apply(lambda x: x.strip())


        # Создаем датафрейм учебной программы без учета строки ИТОГО для таблиц краткой аннотации
        short_df_up = df_up[df_up['Наименование_раздела'] != 'ИТОГО']
        short_df_up = short_df_up[short_df_up['Наименование_раздела'] != 'Итоговая аттестация']



        # получаем единичные значения из листа с данными
        single_row_df = pd.read_excel(name_file_data_obraz_program_po, sheet_name=name_sheet_data, nrows=1,
                                      usecols='A:K')
        single_row_df.iloc[:, 8] = single_row_df.iloc[:, 8].apply(convert_date)  # обрабатываем колонку с датой


        # Очищаем от лишнего поля которые заполняет пользователь
        # Заполняем возможные пустые строки
        single_row_df['Наименование_профессии'] = single_row_df['Наименование_профессии'].fillna('Не заполнено !!!')
        # Очищаем от возможнных пробелов
        single_row_df['Наименование_профессии'] = single_row_df['Наименование_профессии'].apply(lambda x: x.strip())

        single_row_df['Профессиональный_стандарт'] = single_row_df['Профессиональный_стандарт'].fillna(
            'Не заполнено !!!')
        # Очищаем от возможнных пробелов
        single_row_df['Профессиональный_стандарт'] = single_row_df['Профессиональный_стандарт'].apply(
            lambda x: x.strip())


        # получаем датафрейм с технологиями обучения
        tech_df = pd.read_excel(name_file_data_obraz_program_po, sheet_name=name_sheet_data, usecols='L:O')

        tech_df.dropna(thresh=2, inplace=True)  # очищаем от строк в которых не заполнены 2 колонки

        tech_df['Разработчики_программы'] = tech_df['Разработчики_программы'].fillna('Не заполнено')
        # Очищаем от возможнных пробелов
        tech_df['Характеристика_технологии_обучения'] = tech_df['Характеристика_технологии_обучения'].apply(
            lambda x: x.strip())
        tech_df['Технологии_обучения'] = tech_df['Технологии_обучения'].apply(lambda x: x.strip())
        tech_df['Разработчики_программы'] = tech_df['Разработчики_программы'].apply(lambda x: x.strip())


        tech_df['Уровни_квалификации'] = tech_df['Уровни_квалификации'].fillna(0)
        tech_df['Уровни_квалификации'] = tech_df['Уровни_квалификации'].astype(int)

        # создаем переменную для уровней квалификации
        educ_lst = tech_df['Технологии_обучения'].tolist()


        levels_qual = tech_df['Уровни_квалификации'].to_list()
        levels_qual = list(filter(lambda x: x != 0, levels_qual))
        levels_qual = list(map(str, levels_qual))


        # Конвертируем датафрейм с описанием программы в список словарей
        data_program = single_row_df.to_dict('records')

        context = data_program[0]
        # текстовые составные переменные
        context['Уровни_квалификации'] = ','.join(levels_qual)
        context['Технологии_обучения'] = ';\n'.join(educ_lst)


        # Добавляем датафреймы
        context['lst_tech'] = tech_df.to_dict('records')  # образовательные технологии
        context['up_lst'] = df_up.to_dict('records')  # учебный план
        context['short_up_lst'] = short_df_up.to_dict('records')  # учебный план

        lst_dev = [value for value in tech_df['Разработчики_программы'].tolist() if value != 'Не заполнено']
        context['lst_dev'] = lst_dev

        doc = DocxTemplate(name_file_template_educ_program_po)
        # Создаем документ
        doc.render(context)
        # сохраняем документ
        # название программы
        name_prof = single_row_df['Наименование_профессии'].tolist()[0]
        razr = single_row_df['Разряд'].tolist()[0]
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        doc.save(
            f'{path_to_end_folder_obraz_program_po}/Программа ПО {name_prof} {razr} разряда {current_time}.docx')




    except NameError:
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')

    except NotTotal:
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО','На первом листе в первой колонке отсутствует слово ИТОГО')
    except FileNotFoundError:
        # сообщение на случай если путь до папки куда сохраняется файл слишком длинный
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', f'Слишком длинный путь до сохраняемого файла!\nВыберите другую папку')
    except KeyError as e:
        messagebox.showerror('Андраста ver 1.85 Создание программ ПК и ПО', f'Не найдено название колонки {e.args}')
    else:
        messagebox.showinfo('Андраста ver 1.85 Создание программ ПК и ПО', 'Создание образовательной программы\nЗавершено!')

if __name__ == '__main__':
    window = Tk()
    window.title('Андраста ver 1.85 Создание программ ПК и ПО и ПК')
    window.geometry('700x600')
    window.resizable(False, False)


    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # # Создаем вкладку создания программ повышения квалификации ПК по шаблону
    tab_create_educ_program_pk = ttk.Frame(tab_control)
    tab_control.add(tab_create_educ_program_pk, text='Создание программ ПК')
    tab_control.pack(expand=1, fill='both')
    #
    # Создаем вкладку создания программ профессионального обучения ПО по шаблону
    tab_create_educ_program_po = ttk.Frame(tab_control)
    tab_control.add(tab_create_educ_program_po, text='Создание программ ПО')
    tab_control.pack(expand=1, fill='both')
    #
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_create_educ_program_pk,
                      text='Центр опережающей профессиональной подготовки\n Республики Бурятия\nГенерация программ\nповышения квалификации',font=15)
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_pk = resource_path('logo.png')

    img_pk = PhotoImage(file=path_to_img_pk)
    Label(tab_create_educ_program_pk,
          image=img_pk
          ).grid(column=1, row=0, padx=10, pady=25)


    # Создаем кнопку Выбрать файл с данными
    btn_data_data_obraz_pk = Button(tab_create_educ_program_pk, text='1) Выберите файл с данными', font=('Arial Bold', 20),
                          command=select_file_data_obraz_pk
                          )
    btn_data_data_obraz_pk.grid(column=0, row=2, padx=10, pady=10)

    #Создаем кнопку выбора шаблона
    # Создаем кнопку Выбрать файл с данными
    btn_template_educ_program_pk = Button(tab_create_educ_program_pk, text='2) Выберите шаблон', font=('Arial Bold', 20),
                          command=select_file_template_educ_program_pk
                          )
    btn_template_educ_program_pk.grid(column=0, row=3, padx=10, pady=10)



    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_educ_program_pk = Button(tab_create_educ_program_pk, text='3) Выберите конечную папку', font=('Arial Bold', 20),
                                       command=select_end_folder_educ_obraz_pk
                                       )
    btn_choose_end_folder_educ_program_pk.grid(column=0, row=4, padx=10, pady=10)

    btn_create_educ_program_pk = Button(tab_create_educ_program_pk, text='4) Создать программу ПК', font=('Arial Bold', 20),
                                       command=create_educ_program_pk
                                       )
    btn_create_educ_program_pk.grid(column=0, row=5, padx=10, pady=10)


    # Добавляем виджеты на вклдаку создания программ ПО

    # Создаем метку для описания назначения программы
    lbl_hello_po = Label(tab_create_educ_program_po,
                      text='Центр опережающей профессиональной подготовки\n Республики Бурятия\nГенерация программ\nпрофессионального обучения',font=15)
    lbl_hello_po.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_po = resource_path('logo.png')

    img_po = PhotoImage(file=path_to_img_po)
    Label(tab_create_educ_program_po,
          image=img_po
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_data_data_obraz_po = Button(tab_create_educ_program_po, text='1) Выберите файл с данными', font=('Arial Bold', 20),
                                 command=select_file_data_obraz_po
                                 )
    btn_data_data_obraz_po.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку выбора шаблона
    # Создаем кнопку Выбрать файл с данными
    btn_template_educ_program_po = Button(tab_create_educ_program_po, text='2) Выберите шаблон', font=('Arial Bold', 20),
                                       command=select_file_template_educ_program_po
                                       )
    btn_template_educ_program_po.grid(column=0, row=3, padx=10, pady=10)

    #Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_educ_program_po = Button(tab_create_educ_program_po, text='3) Выберите конечную папку',
                                                font=('Arial Bold', 20),
                                                command=select_end_folder_educ_obraz_po
                                                )
    btn_choose_end_folder_educ_program_po.grid(column=0, row=4, padx=10, pady=10)

    btn_create_educ_program_po = Button(tab_create_educ_program_po, text='4) Создать программу ПО', font=('Arial Bold', 20),
                                     command=create_educ_program_po
                                     )
    btn_create_educ_program_po.grid(column=0, row=5, padx=10, pady=10)

    window.mainloop()
