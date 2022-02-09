import tkinter as tk
import openpyxl
import pandas as pd
import os

from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time

# Отображать все колонки в пандас
pd.set_option('display.max_columns', None)


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller
     Для того чтобы упаковать картинку в exe"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_file_params():
    """
    Функция для выбора файла c ячейками которые нужно подсчитать
    :return: Путь к файлу
    """
    global name_file_params
    name_file_params = filedialog.askopenfilename(
        filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_files_data():
    """
    Функция для выбора файлов с данными параметры из которых нужно подсчитать
    :return: Путь к файлам с данными
    """
    global names_files_data
    # Получаем путь к файлу
    names_files_data = filedialog.askopenfilenames(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder():
    """
    Функция для выбора папки куда будут генерироваться файл  с результатом подсчета и файл с проверочной инфомрацией
    :return:
    """
    global path_to_end_folder
    path_to_end_folder = filedialog.askdirectory()


def calculate_data():
    """
    Функция для подсчета данных из файлов
    :return:
    """
    count = 0
    count_errors = 0
    quantity_files = len(names_files_data)
    current_time = time.strftime('%H_%M_%S')
    # Состояние чекбокса
    mode_text = mode_text_value.get()

    try:
        # Получаем название обрабатываемого листа
        name_list_df = pd.read_excel(name_file_params, nrows=2)
        name_list = name_list_df['Значение'].loc[0]

        # Получаем количество листов в файле, на случай если название листа не совпадает с правильным
        quantity_list_in_file = name_list_df['Значение'].loc[1]

        # Получаем шаблон с данными, первую строку пропускаем, поскольку название обрабатываемого листа мы уже получили
        df = pd.read_excel(name_file_params, skiprows=2)

        # Создаем словарь параметров
        param_dict = dict()

        for row in df.itertuples():
            param_dict[row[1]] = row[2]
        # Создаем словарь для подсчета данных, копируя ключи из словаря параметров, значения в зависимости от способа обработки

        if mode_text == 'Yes':
            result_dct = {key: '' for key, value in param_dict.items()}
        else:
            result_dct = {key: 0 for key, value in param_dict.items()}

            # Создаем датафрейм для контроля процесса подсчета и заполняем словарь на основе которого будем делать итоговую таблицу

        check_df = pd.DataFrame(columns=param_dict.keys())
        # Вставляем колонку для названия файла
        check_df.insert(0, 'Название файла', '')
        for file in names_files_data:
            # Проверяем чтобы файл не был резервной копией.
            if '~$' in file:
                continue
            # Создаем словарь для создания строки которую мы будем добавлять в проверочный датафрейм
            new_row = dict()
            # Получаем  отбрасываем расширение файла
            full_name_file = file.split('.')[0]
            # Получаем имя файла  без пути
            name_file = full_name_file.split('/')[-1]
            try:

                new_row['Название файла'] = name_file

                wb = openpyxl.load_workbook(file)
                # Проверяем наличие листа
                if name_list in wb.sheetnames:
                    sheet = wb[name_list]
                # проверяем количество листов в файле.Если значение равно 1 то просто берем первый лист, иначе вызываем ошибку
                elif quantity_list_in_file == 1:
                    temp_name = wb.sheetnames[0]
                    sheet = wb[temp_name]
                else:
                    raise Exception

                for key, cell in param_dict.items():
                    result_dct[key] += check_data(sheet[cell].value, mode_text)
                    new_row[key] = sheet[cell].value
                check_df = check_df.append(new_row, ignore_index=True)
                count += 1
            # Ловим исключения
            except Exception as err:
                count_errors += 1
                with open(f'{path_to_end_folder}/ERRORS {current_time}.txt', 'a', encoding='utf-8') as f:
                    f.write(f'Файл {name_file} не обработан!!!\n')

        check_df.to_excel(f'{path_to_end_folder}/Проверка вычисления.xlsx', index=False)

        # Создание итоговой таблицы результатов подсчета

        finish_result = pd.DataFrame()

        finish_result['Наименование показателя'] = result_dct.keys()
        finish_result['Значение показателя'] = result_dct.values()
        # Проводим обработку в зависимости от значения переключателя

        if mode_text == 'Yes':
            # Обрабатываем датафрейм считая текстовые данные
            count_text_df = count_text_value(finish_result)
            count_text_df.to_excel(f'{path_to_end_folder}/Подсчет текстовых значений.xlsx')
        else:
            finish_result.to_excel(f'{path_to_end_folder}/Итоговые значения.xlsx', index=False)

        if count_errors != 0:
            messagebox.showinfo('Cassandra',
                                f'Обработка файлов завершена!\nОбработано файлов:  {count} из {quantity_files}\n Необработанные файлы указаны в файле {path_to_end_folder}/ERRORS {current_time}.txt ')
        else:
            messagebox.showinfo('Cassandra',
                                f'Обработка файлов успешно завершена!\nОбработано файлов:  {count} из {quantity_files}')

    except NameError:
        messagebox.showerror('Cassandra', 'Выберите файл с параметрами,обрабатываемые данные, конечную папку')


def count_text_value(df):
    """
    Функция для подсчета количества вариантов того или иного показателя
    :param df: датафрейм с сырыми данными. Название показателя значение показателя(строка разделенная ;)
    :return: обработанный датафрейм с мультиндексом, где (Название показателя это индекс верхнего уровня, вариант показателя это индекс второго уровня а значение это сколько раз встречался
    этот вариант в обрабатываемых файлах)
    """
    data = dict()

    #
    for row in df.itertuples():
        value = row[2]
        if type(value) == float or type(value) == int:
            continue
        # Создаем список, разделяя строку по ;
        lst_value = row[2].split(';')[:-1]
        #     # Отрезаем последний элемент, поскольку это пустое значение
        temp_df = pd.DataFrame({'Value': lst_value})
        counts_series = temp_df['Value'].value_counts()
        # Делаем индекс колонкой и превращаем в обычную таблицу
        index_count_values = counts_series.reset_index()
        # Итерируемся по таблице.Это делается чтобы заполнить словарь на основе которого будет создаваться итоговая таблица
        for count_row in index_count_values.itertuples():
            # print(count_row)
            # Заполняем словарь
            data[(row[1], count_row[1])] = count_row[2]
    # Создаем на основе получившегося словаря таблицу
    out_df = pd.Series(data).to_frame().reset_index()
    out_df = out_df.set_index(['level_0', 'level_1'])
    out_df.index.names = ['Название показателя', 'Вариант показателя']
    out_df.rename(columns={0: 'Количество'}, inplace=True)
    return out_df


def check_data(cell, text_mode):
    """
    Функция для проверки значения ячейки. Для обработки пустых значений, строковых значений, дат
    :param cell: значение ячейки
    :return: 0 если значение ячейки не число
            число если значение ячейки число(ха звучит глуповато)
    думаю функция должна работать с дополнительным параметром, от которого будет зависеть подсчет значений навроде галочек или плюсов в анкетах или опросах.
    """
    # Проверяем режим работы. если текстовый, то просто складываем строки
    if text_mode == 'Yes':
        if cell is None:
            return ''
        else:
            temp_str = str(cell)
            return f'{temp_str};'
    # Если режим работы стандартный. Убрал подсчет строк и символов в числовом режиме, чтобы не запутывать.
    else:
        if cell is None:
            return 0
        if type(cell) == int:
            return cell
        elif type(cell) == float:
            return cell
        else:
            return 0


if __name__ == '__main__':
    window = Tk()
    window.title('Cassandra')
    window.geometry('600x800')
    window.resizable(False, False)

    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку создания документов по шаблону
    tab_calculate_data = ttk.Frame(tab_control)
    tab_control.add(tab_calculate_data, text='Обработка данных')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_calculate_data,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nПодсчет значений заданных ячеек из нескольких файлов Excel',
                      font=25)
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')
    img = PhotoImage(file=path_to_img)
    Label(tab_calculate_data,
          image=img
          ).grid(column=0, row=1, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с параметрами
    btn_select_file_params = Button(tab_calculate_data, text='1) Выбрать файл с параметрами', font=('Arial Bold', 20),
                                    command=select_file_params
                                    )
    btn_select_file_params.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку Выбрать файл с данными
    btn_select_files_data = Button(tab_calculate_data, text='2) Выбрать файлы с данными', font=('Arial Bold', 20),
                                   command=select_files_data
                                   )
    btn_select_files_data.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_calculate_data, text='3) Выбрать конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder
                                   )
    btn_choose_end_folder.grid(column=0, row=4, padx=10, pady=10)

    # Создаем переменную для хранения результа переключения чекбокса
    mode_text_value = tk.StringVar()
    # Устанавливаем значение по умолчанию для этой переменной. По умолчанию будет вестись подсчет числовых данных
    mode_text_value.set('No')
    # Создаем чекбокс для выбора режима подсчета

    chbox_mode_calculate = tk.Checkbutton(tab_calculate_data,
                                          text='Поставьте галочку, если вам нужно посчитать текстовые данные ',
                                          variable=mode_text_value,
                                          offvalue='No',
                                          onvalue='Yes')
    chbox_mode_calculate.grid(column=0, row=5, padx=10, pady=10)

    # Создаем кнопку для запуска подсчета файлов

    btn_calculate = Button(tab_calculate_data, text='4) Подсчитать', font=('Arial Bold', 20),
                           command=calculate_data
                           )
    btn_calculate.grid(column=0, row=6, padx=10, pady=10)

    window.mainloop()
