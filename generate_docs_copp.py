"""
Скрипт для генерации документов ДПО и ПО из шаблонов
"""
import pandas as pd
import numpy as np
import os
from dateutil.parser import ParserError
from docxtpl import DocxTemplate
from docxcompose.composer import Composer
from docx import Document
from docx2pdf import convert
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from jinja2 import exceptions
import time
import datetime
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import sys
import locale
import tempfile
import re
from pytrovich.detector import PetrovichGenderDetector
from pytrovich.enums import NamePart, Gender, Case
from pytrovich.maker import PetrovichDeclinationMaker
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)

def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'
    except TypeError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'


def processing_date_column(df, lst_columns):
    """
    Функция для обработки столбцов с датами. конвертация в строку формата ДД.ММ.ГГГГ
    """
    # получаем первую строку
    first_row = df.iloc[0, lst_columns]

    lst_first_row = list(first_row)  # Превращаем строку в список
    lst_date_columns = []  # Создаем список куда будем сохранять колонки в которых находятся даты
    tupl_row = list(zip(lst_columns,
                        lst_first_row))  # Создаем список кортежей формата (номер колонки,значение строки в этой колонке)

    for idx, value in tupl_row:  # Перебираем кортеж
        result = check_date_columns(idx, value)  # проверяем является ли значение датой
        if result:  # если да то добавляем список порядковый номер колонки
            lst_date_columns.append(result)
        else:  # иначе проверяем следующее значение
            continue
    for i in lst_date_columns:  # Перебираем список с колонками дат, превращаем их в даты и конвертируем в нужный строковый формат
        df.iloc[:, i] = pd.to_datetime(df.iloc[:, i], errors='coerce', dayfirst=True)
        df.iloc[:, i] = df.iloc[:, i].apply(create_doc_convert_date)

def check_date_columns(i, value):
    """
    Функция для проверки типа колонки. Необходимо найти колонки с датой
    :param i:
    :param value:
    :return:
    """
    try:
        itog = pd.to_datetime(str(value), infer_datetime_format=True)
    except:
        pass
    else:
        return i

def clean_value(value):
    """
    Функция для обработки значений колонки от  пустых пробелов,нан
    :param value: значение ячейки
    :return: очищенное значение
    """
    if value is np.nan:
        return 'Не заполнено'
    str_value = str(value)
    if str_value == '':
        return 'Не заполнено'
    elif str_value ==' ':
        return 'Не заполнено'

    return str_value

"""
Функции для создания склонений по падежам и инициалов
"""
def capitalize_double_name(word):
    """
    Функция для того чтобы в двойных именах и фамилиях вторая часть была также с большой буквы
    """
    lst_word = word.split('-')  # сплитим по дефису
    if len(lst_word) == 1:  # если длина списка равна 1 то это не двойная фамилия и просто возвращаем слово

        return word
    elif len(lst_word) == 2:
        first_word = lst_word[0].capitalize()  # делаем первую букву слова заглавной а остальные строчными
        second_word = lst_word[1].capitalize()
        return f'{first_word}-{second_word}'
    else:
        return 'Не удалось просклонять'


def case_lastname(maker, lastname, gender, case: Case):
    """
    Функция для обработки и склонения фамилии. Это нужно для обработки случаев двойной фамилии
    """

    lst_lastname = lastname.split('-')  # сплитим по дефису

    if len(lst_lastname) == 1:  # если длина списка равна 1 то это не двойная фамилия и просто обрабатываем слово
        case_result_lastname = maker.make(NamePart.LASTNAME, gender, case, lastname)
        return case_result_lastname
    elif len(lst_lastname) == 2:
        first_lastname = lst_lastname[0].capitalize()  # делаем первую букву слова заглавной а остальные строчными
        second_lastname = lst_lastname[1].capitalize()
        # Склоняем по отдельности
        first_lastname = maker.make(NamePart.LASTNAME, gender, case, first_lastname)
        second_lastname = maker.make(NamePart.LASTNAME, gender, case, second_lastname)

        return f'{first_lastname}-{second_lastname}'


def detect_gender(detector, lastname, firstname, middlename):
    """
    Функция для определения гендера слова
    """
    #     detector = PetrovichGenderDetector() # создаем объект детектора
    try:
        gender_result = detector.detect(lastname=lastname, firstname=firstname, middlename=middlename)
        return gender_result
    except StopIteration:  # если не удалось определить то считаем что гендер андрогинный
        return Gender.ANDROGYNOUS


def decl_on_case(fio: str, case: Case) -> str:
    """
    Функция для склонения ФИО по падежам
    """
    fio = fio.strip()  # очищаем строку от пробельных символов с начала и конца
    part_fio = fio.split()  # разбиваем по пробелам создавая список где [0] это Фамилия,[1]-Имя,[2]-Отчество

    if len(part_fio) == 3:  # проверяем на длину и обрабатываем только те что имеют длину 3 во всех остальных случаях просим просклонять самостоятельно
        maker = PetrovichDeclinationMaker()  # создаем объект класса
        lastname = part_fio[0].capitalize()  # Фамилия
        firstname = part_fio[1].capitalize()  # Имя
        middlename = part_fio[2].capitalize()  # Отчество

        # Определяем гендер для корректного склонения
        detector = PetrovichGenderDetector()  # создаем объект детектора
        gender = detect_gender(detector, lastname, firstname, middlename)
        # Склоняем

        case_result_lastname = case_lastname(maker, lastname, gender, case)  # обрабатываем фамилию
        case_result_firstname = maker.make(NamePart.FIRSTNAME, gender, case, firstname)
        case_result_firstname = capitalize_double_name(case_result_firstname)  # обрабатываем случаи двойного имени
        case_result_middlename = maker.make(NamePart.MIDDLENAME, gender, case, middlename)
        # Возвращаем результат
        result_fio = f'{case_result_lastname} {case_result_firstname} {case_result_middlename}'
        return result_fio

    else:
        return 'Проверьте количество слов, должно быть 3 разделенных пробелами слова'


def create_initials(cell, checkbox, space):
    """
    Функция для создания инициалов
    """
    lst_fio = cell.split(' ')  # сплитим по пробелу
    if len(lst_fio) == 3:  # проверяем на стандартный размер в 3 слова иначе ничего не меняем
        if checkbox == 'ФИ':
            if space == 'без пробела':
                # возвращаем строку вида Иванов И.И.
                return f'{lst_fio[0]} {lst_fio[1][0].upper()}.{lst_fio[2][0].upper()}.'
            else:
                # возвращаем строку с пробелом после имени Иванов И. И.
                return f'{lst_fio[0]} {lst_fio[1][0].upper()}. {lst_fio[2][0].upper()}.'

        else:
            if space == 'без пробела':
                # И.И. Иванов
                return f'{lst_fio[1][0].upper()}.{lst_fio[2][0].upper()}. {lst_fio[0]}'
            else:
                # И. И. Иванов
                return f'{lst_fio[1][0].upper()}. {lst_fio[2][0].upper()}. {lst_fio[0]}'
    else:
        return cell

def declension_fio_by_case(df,fio_column):
    """
    Функция для склонения ФИО по падежам и создания инициалов
    :param fio_column: название колонки с ФИО
    :param data_decl_case: Путь к файлу
    :param path_to_end_folder_decl_case: Путь  к итоговой папке
    :return: файл Excel в котором после колонки fio_column добавляется 29 колонок с падежами
    """

    temp_df = pd.DataFrame()  # временный датафрейм для хранения колонок просклоненных по падежам

    # Получаем номер колонки с фио которые нужно обработать
    lst_columns = list(df.columns)  # Превращаем в список
    index_fio_column = lst_columns.index(fio_column)  # получаем индекс

    # Обрабатываем nan значения и те которые обозначены пробелом
    df[fio_column].fillna('Не заполнено', inplace=True)
    df[fio_column] = df[fio_column].apply(lambda x: x.strip())
    df[fio_column] = df[fio_column].apply(
        lambda x: x if x else 'Не заполнено')  # Если пустая строка то заменяем на значение Не заполнено

    temp_df['Родительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.GENITIVE))
    temp_df['Дательный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.DATIVE))
    temp_df['Винительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.ACCUSATIVE))
    temp_df['Творительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.INSTRUMENTAL))
    temp_df['Предложный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.PREPOSITIONAL))
    temp_df['Фамилия_инициалы'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Инициалы_фамилия'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Фамилия_инициалы_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами родительный падеж
    temp_df['Фамилия_инициалы_род_падеж'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_род_падеж_пробел'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_род_падеж'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_род_падеж_пробел'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами дательный падеж
    temp_df['Фамилия_инициалы_дат_падеж'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_дат_падеж_пробел'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_дат_падеж'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_дат_падеж_пробел'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами винительный падеж
    temp_df['Фамилия_инициалы_вин_падеж'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_вин_падеж_пробел'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_вин_падеж'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_вин_падеж_пробел'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами творительный падеж
    temp_df['Фамилия_инициалы_твор_падеж'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_твор_падеж_пробел'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_твор_падеж'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_твор_падеж_пробел'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))
    # Создаем колонки для склонения фамилий с иницалами предложный падеж
    temp_df['Фамилия_инициалы_пред_падеж'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_пред_падеж_пробел'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_пред_падеж'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_пред_падеж_пробел'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Вставляем получившиеся колонки после базовой колонки с фио
    df.insert(index_fio_column + 1, 'Родительный_падеж', temp_df['Родительный_падеж'])
    df.insert(index_fio_column + 2, 'Дательный_падеж', temp_df['Дательный_падеж'])
    df.insert(index_fio_column + 3, 'Винительный_падеж', temp_df['Винительный_падеж'])
    df.insert(index_fio_column + 4, 'Творительный_падеж', temp_df['Творительный_падеж'])
    df.insert(index_fio_column + 5, 'Предложный_падеж', temp_df['Предложный_падеж'])
    df.insert(index_fio_column + 6, 'Фамилия_инициалы', temp_df['Фамилия_инициалы'])
    df.insert(index_fio_column + 7, 'Инициалы_фамилия', temp_df['Инициалы_фамилия'])
    df.insert(index_fio_column + 8, 'Фамилия_инициалы_пробел', temp_df['Фамилия_инициалы_пробел'])
    df.insert(index_fio_column + 9, 'Инициалы_фамилия_пробел', temp_df['Инициалы_фамилия_пробел'])
    # Добавляем колонки с склонениями инициалов родительный падеж
    df.insert(index_fio_column + 10, 'Фамилия_инициалы_род_падеж', temp_df['Фамилия_инициалы_род_падеж'])
    df.insert(index_fio_column + 11, 'Фамилия_инициалы_род_падеж_пробел',
              temp_df['Фамилия_инициалы_род_падеж_пробел'])
    df.insert(index_fio_column + 12, 'Инициалы_фамилия_род_падеж', temp_df['Инициалы_фамилия_род_падеж'])
    df.insert(index_fio_column + 13, 'Инициалы_фамилия_род_падеж_пробел',
              temp_df['Инициалы_фамилия_род_падеж_пробел'])
    # Добавляем колонки с склонениями инициалов дательный падеж
    df.insert(index_fio_column + 14, 'Фамилия_инициалы_дат_падеж', temp_df['Фамилия_инициалы_дат_падеж'])
    df.insert(index_fio_column + 15, 'Фамилия_инициалы_дат_падеж_пробел',
              temp_df['Фамилия_инициалы_дат_падеж_пробел'])
    df.insert(index_fio_column + 16, 'Инициалы_фамилия_дат_падеж', temp_df['Инициалы_фамилия_дат_падеж'])
    df.insert(index_fio_column + 17, 'Инициалы_фамилия_дат_падеж_пробел',
              temp_df['Инициалы_фамилия_дат_падеж_пробел'])
    # Добавляем колонки с склонениями инициалов винительный падеж
    df.insert(index_fio_column + 18, 'Фамилия_инициалы_вин_падеж', temp_df['Фамилия_инициалы_вин_падеж'])
    df.insert(index_fio_column + 19, 'Фамилия_инициалы_вин_падеж_пробел',
              temp_df['Фамилия_инициалы_вин_падеж_пробел'])
    df.insert(index_fio_column + 20, 'Инициалы_фамилия_вин_падеж', temp_df['Инициалы_фамилия_вин_падеж'])
    df.insert(index_fio_column + 21, 'Инициалы_фамилия_вин_падеж_пробел',
              temp_df['Инициалы_фамилия_вин_падеж_пробел'])
    # Добавляем колонки с склонениями инициалов творительный падеж
    df.insert(index_fio_column + 22, 'Фамилия_инициалы_твор_падеж', temp_df['Фамилия_инициалы_твор_падеж'])
    df.insert(index_fio_column + 23, 'Фамилия_инициалы_твор_падеж_пробел',
              temp_df['Фамилия_инициалы_твор_падеж_пробел'])
    df.insert(index_fio_column + 24, 'Инициалы_фамилия_твор_падеж', temp_df['Инициалы_фамилия_твор_падеж'])
    df.insert(index_fio_column + 25, 'Инициалы_фамилия_твор_падеж_пробел',
              temp_df['Инициалы_фамилия_твор_падеж_пробел'])
    # Добавляем колонки с склонениями инициалов предложный падеж
    df.insert(index_fio_column + 26, 'Фамилия_инициалы_пред_падеж', temp_df['Фамилия_инициалы_пред_падеж'])
    df.insert(index_fio_column + 27, 'Фамилия_инициалы_пред_падеж_пробел',
              temp_df['Фамилия_инициалы_пред_падеж_пробел'])
    df.insert(index_fio_column + 28, 'Инициалы_фамилия_пред_падеж', temp_df['Инициалы_фамилия_пред_падеж'])
    df.insert(index_fio_column + 29, 'Инициалы_фамилия_пред_падеж_пробел',
              temp_df['Инициалы_фамилия_пред_падеж_пробел'])

    return df





"""
Вспомогательные функции
"""
def clean_columns(df:pd.DataFrame,lst_columns:list)->pd.DataFrame:
    """
    Функция для очистки определенных колонок от значений Nan, приведения к строковому виду и очистке от лишних пробелов
    с помощью strip
    """
    df[lst_columns] = df[lst_columns].fillna('Не заполнено !!!').astype(str)
    df[lst_columns] = df[lst_columns].apply(lambda x:x.strip)
    return df

def processing_snils(snils):
    """
    Функция для приведения строки СНИЛС к виду ХХХ-ХХХ-ХХХ ХХ, в противном случае возвращается значение Ошибка
    """
    snils = str(snils)
    result = re.findall(r'\d', snils)
    if len(result) == 11:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        third_group = ''.join(result[6:9])
        four_group = ''.join(result[9:11])

        out_snils = f'{first_group}-{second_group}-{third_group} {four_group}'
        return out_snils
    else:
        return f'Неправильное значение СНИЛС {snils}'



def generate_docs(path_to_folder_template:str,file_data:str,path_to_end_folder:str,type_course):
    """
    path_to_folder_template: путь к папке с шаблонами
    file_data: путь к таблице
    path_to_end_folder: путь к конечной папке
    type_course: тип курса ДПО или ПО
    """
    print(type_course)

    df = pd.read_excel(file_data,sheet_name=type_course,dtype=str)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)  # применяем strip, чтобы все данные корректно вставлялись
    # Обрабатываем колонки с датами
    df['Дата_выдачи_документа'] = pd.to_datetime(df['Дата_выдачи_документа'], errors='coerce',dayfirst=True)
    df['Дата_рождения_получателя'] = pd.to_datetime(df['Дата_рождения_получателя'], errors='coerce',dayfirst=True)

    # Создаем объединенную колонку ФИО
    part_fio_columns = ['Фамилия_получателя','Имя_получателя','Отчество_получателя'] # названия колонок составляющих ФИО
    df['ФИО'] = df.apply(lambda row:' '.join(row[part_fio_columns]),axis=1)
    df = declension_fio_by_case(df,'ФИО')



    # Обрабатываем колонку СНИЛС приводя данные там к виду ХХХ-ХХХ-ХХХ ХХ
    df['СНИЛС'] = df['СНИЛС'].apply(processing_snils)


    # TODO генерация файла ФИС -ФРДО
    # Словарь замены значений гражданства получателя. В ФИС ФРДО нужно использовать числовые значения
    dct_citizenship = {'Российская Федерация':'643','Азербайджан':'031','Армения':'051','Беларусь':'112','Казахстан':'398','Киргизия':'417','Молдова':'498',
                       'Таджикистан':'762','Туркменистан':'795','Узбекистан':'860','Украина':'804','Литва':'440','Латвия':'428','Эстония':'233',}

    df['Гражданство_получателя'] = df['Гражданство_получателя'].replace(dct_citizenship) # проводим замену.






    lst_frdo_columns = ['','','','','','','','','','','','','','','',]  # Список колонок используемых в ФРДО




    # lst_xlsx = [] # список для хранения шаблонов Excel
    # # Открывае шаблон ДПО ФИС ФРДО
    # lst_df = df.values.tolist() # превращаем в список списков
    # template_dpo_fis_frdo = openpyxl.load_workbook(f'{path_to_folder_template}/ДПО/ДПО_ФИС-ФРДО.xlsx')
    # first_sheet = template_dpo_fis_frdo.sheetnames[0]
    # start_row = 2
    # for row_data in lst_df:
    #     for col, value in enumerate(row_data, 1):
    #         template_dpo_fis_frdo[first_sheet].cell(row=start_row, column=col, value=value)
    #     start_row += 1
    # for column in template_dpo_fis_frdo[first_sheet].columns:
    #     max_length = 0
    #     column_name = get_column_letter(column[0].column)
    #     for cell in column:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2)
    #     template_dpo_fis_frdo[first_sheet].column_dimensions[column_name].width = adjusted_width
    #
    # template_dpo_fis_frdo.save(f'{path_to_end_folder}/ДПО ФИС-ФРДО для загрузки.xlsx')
    df.to_excel(f'{path_to_end_folder}/Датафрейм.xlsx', index=False)










if __name__ == '__main__':
    path_folder_template_main = 'data/example/Шаблоны'
    # data_file_main = 'data/example/ДПО_Цифровые_инструменты_в_образовательной_среде_БРИЭТ_март.xlsx'
    data_file_main = 'data/example/Тестовый вариант.xlsx'
    path_end_folder_main = 'data/example/result'


    generate_docs(path_folder_template_main,data_file_main,path_end_folder_main,'ДПО')
    print('Lindy Booth!')
